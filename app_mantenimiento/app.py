"""
Atlas Excellence App - Backend Server
2026 Atlas Senior Living
Developed by Gabriel Rosales | gabriel@gabrielrosales.org
Flask application for managing maintenance and cleaning reports
With user authentication and automatic community detection
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_from_directory, send_file
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
import os
import re
import requests
from datetime import datetime #calendarios
import json
from services.question_manager import QuestionManager
from services.inspection_service import InspectionService
from services.input_sanitizer import InputSanitizer


# Initialize Flask app
app = Flask(__name__)

# Configuration
# SECRET_KEY resolution order (never falls back to a hardcoded insecure key):
#   1. SECRET_KEY env var (preferred for production).
#   2. A persisted random key in data/.secret_key (stable across restarts so
#      sessions survive a reboot). Generated with secrets.token_hex on first run.
def _resolve_secret_key():
    env_key = os.environ.get('SECRET_KEY')
    if env_key:
        return env_key
    import secrets
    key_path = os.path.join(os.path.dirname(__file__), 'data', '.secret_key')
    try:
        if os.path.exists(key_path):
            with open(key_path, 'r', encoding='utf-8') as f:
                k = f.read().strip()
            if k:
                return k
        os.makedirs(os.path.dirname(key_path), exist_ok=True)
        k = secrets.token_hex(32)
        with open(key_path, 'w', encoding='utf-8') as f:
            f.write(k)
        try:
            os.chmod(key_path, 0o600)
        except OSError:
            pass
        app.logger.warning('SECRET_KEY env var not set — generated and persisted a '
                           'random key in data/.secret_key. Set SECRET_KEY in the '
                           'environment for full control.')
        return k
    except OSError:
        # Last resort (read-only fs): a per-process random key. Sessions won't
        # survive a restart, but the key is never a known hardcoded value.
        return secrets.token_hex(32)


app.config['SECRET_KEY'] = _resolve_secret_key()
app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 hours

# --- Session cookie hardening ---
# Secure: only send the cookie over HTTPS (disable for local http via COOKIE_SECURE=0).
# HttpOnly: JS can't read the cookie (mitigates XSS cookie theft).
# SameSite=Lax: the browser won't attach the session cookie to cross-site POST/
#   fetch requests, which blocks the bulk of CSRF, while still allowing top-level
#   navigations (so inspection-report email links still log the user in).
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('COOKIE_SECURE', '1') != '0'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Configure upload folder
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Allowed file extensions for uploads
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'webp'}

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ==================== SECURITY MIDDLEWARE ====================
from urllib.parse import urlparse as _urlparse

# Endpoints that legitimately accept cross-origin / no-Origin POSTs. (None today;
# kept for clarity / future webhooks.)
_CSRF_EXEMPT_PATHS = set()


@app.before_request
def _csrf_origin_guard():
    """Lightweight CSRF protection: for state-changing requests, require the
    Origin (or Referer) to match our own host. Combined with the SameSite=Lax
    session cookie, this blocks cross-site forged requests. Same-origin fetch/
    form posts always send a matching Origin, so legitimate traffic is unaffected."""
    if request.method not in ('POST', 'PUT', 'DELETE', 'PATCH'):
        return
    if request.path in _CSRF_EXEMPT_PATHS:
        return
    source = request.headers.get('Origin') or request.headers.get('Referer')
    if not source:
        # No Origin/Referer at all (rare). The SameSite cookie still protects us,
        # so allow rather than break non-browser clients / health checks.
        return
    if _urlparse(source).netloc != request.host:
        app.logger.warning('Blocked cross-origin %s %s from %s',
                           request.method, request.path, source)
        return jsonify({'status': 'error', 'message': 'Cross-origin request blocked'}), 403


@app.before_request
def _track_presence():
    """Note that the signed-in user is active. PresenceService throttles the
    write to roughly once a minute per user, and swallows its own errors, so
    this stays cheap and can never break a request. Static files and polling
    endpoints are skipped so background refreshes don't fake activity."""
    user = session.get('user')
    if not user:
        return
    path = request.path
    if path.startswith('/static/') or path in _PRESENCE_IGNORE:
        return
    presence_service.touch(user)


# Endpoints the dashboard polls on a timer — hitting them isn't a sign of life.
_PRESENCE_IGNORE = {'/api/activity/live', '/api/presence'}


_MUST_CHANGE_ALLOW = {'/change-password', '/api/profile/password', '/logout',
                      '/api/user-info', '/login', '/api/login', '/api/forgot-password'}


# The preview used to block every write. It doesn't any more, on purpose: the
# point of training is to show the loop working, and a demonstration that stops
# short of the result teaches half of it.
#
# Nothing is falsified by allowing it. A preview changes what an administrator
# is shown, not who they are — session['user'] is untouched, so a comment left
# during a preview is recorded and emailed under their own name, not the
# Executive Director's.
#
# Administrative actions stay out of reach regardless: is_admin() answers False
# while previewing, so every admin endpoint refuses on its own terms. What a
# preview can do is exactly what that community can do.


@app.before_request
def _must_change_guard():
    """When an admin has reset an account, the user must set a new password
    before doing anything else. Allow only the change-password page + its API,
    logout and static assets until they do."""
    if not session.get('must_change'):
        return
    path = request.path
    if path in _MUST_CHANGE_ALLOW or path.startswith('/static/'):
        return
    if path.startswith('/api/'):
        return jsonify({'status': 'error', 'must_change_password': True,
                        'message': 'Please set a new password to continue.'}), 403
    return redirect(url_for('change_password_page'))


@app.after_request
def _security_headers(resp):
    """Standard hardening headers applied to every response."""
    resp.headers.setdefault('X-Content-Type-Options', 'nosniff')
    resp.headers.setdefault('X-Frame-Options', 'DENY')
    resp.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    resp.headers.setdefault('X-XSS-Protection', '0')
    # HSTS only matters over HTTPS; harmless to always send (browsers ignore on http).
    resp.headers.setdefault('Strict-Transport-Security', 'max-age=31536000; includeSubDomains')
    return resp


# --- Simple in-memory login throttle (per IP + username) ---
import time as _time
from collections import defaultdict as _defaultdict
_LOGIN_ATTEMPTS = _defaultdict(list)   # key -> [timestamps of recent failures]
_LOGIN_WINDOW = 300                    # 5 minutes
_LOGIN_MAX_FAILS = 8                   # allowed failures per window


def _client_ip():
    """Caller's IP. Behind nginx the real address arrives in X-Forwarded-For."""
    ip = (request.headers.get('X-Forwarded-For', '') or request.remote_addr or '').split(',')[0].strip()
    return ip or 'unknown'


def _login_throttle_key():
    return _client_ip()


def login_is_throttled():
    key = _login_throttle_key()
    now = _time.time()
    fails = [t for t in _LOGIN_ATTEMPTS[key] if now - t < _LOGIN_WINDOW]
    _LOGIN_ATTEMPTS[key] = fails
    return len(fails) >= _LOGIN_MAX_FAILS


def record_login_failure():
    _LOGIN_ATTEMPTS[_login_throttle_key()].append(_time.time())


def reset_login_failures():
    _LOGIN_ATTEMPTS.pop(_login_throttle_key(), None)


# ---- Automatic cache-busting for static assets ----------------------------
# static_v('theme.css') -> '/static/theme.css?v=<file-mtime>'. The version
# changes by itself whenever the file is saved, so browsers always pick up
# edits without anyone hand-bumping a version string.
_STATIC_DIR = os.path.join(os.path.dirname(__file__), 'static')


@app.context_processor
def _inject_static_v():
    def static_v(filename):
        path = os.path.join(_STATIC_DIR, filename)
        try:
            stamp = int(os.path.getmtime(path))
        except OSError:
            stamp = 0
        return url_for('static', filename=filename) + ('?v=%d' % stamp)
    return {'static_v': static_v}


# ==================== SERVICE INITIALIZATION ====================

# Initialize data directory
DATA_FOLDER = os.path.join(os.path.dirname(__file__), 'data')
os.makedirs(DATA_FOLDER, exist_ok=True)

# Seed the live data files from data/seeds/ on first run (e.g. a fresh server).
# Live data files are git-ignored so deploys never overwrite them; the seeds
# ship in git and only populate a file that doesn't exist yet.
import shutil
SEED_FOLDER = os.path.join(DATA_FOLDER, 'seeds')
for _seed_name in ('regions.json', 'questions.json', 'survey_types.json', 'resources.json', 'movein_template.json'):
    _live = os.path.join(DATA_FOLDER, _seed_name)
    _seed = os.path.join(SEED_FOLDER, _seed_name)
    if not os.path.exists(_live) and os.path.exists(_seed):
        try:
            shutil.copyfile(_seed, _live)
        except OSError as _e:
            app.logger.error(f'Could not seed {_seed_name}: {_e}')

# Initialize QuestionManager service
QUESTIONS_FILE = os.path.join(DATA_FOLDER, 'questions.json')
question_manager = QuestionManager(QUESTIONS_FILE)

# Initialize InspectionService
INSPECTIONS_FILE = os.path.join(DATA_FOLDER, 'inspections.json')
from services.inspection_service import InspectionService
inspection_service = InspectionService(INSPECTIONS_FILE, UPLOAD_FOLDER)

# Initialize FileUploadHandler.
# If S3_BUCKET is set, photos are stored privately in S3 and served via signed
# URLs; otherwise they fall back to the local static/uploads folder. This keeps
# the app working in dev / before the bucket is configured.
from services.file_upload_handler import FileUploadHandler
S3_BUCKET = os.environ.get('S3_BUCKET', '').strip() or None
AWS_REGION = os.environ.get('AWS_REGION', 'us-east-1')
S3_URL_EXPIRY = int(os.environ.get('S3_URL_EXPIRY', '3600'))
file_upload_handler = FileUploadHandler(
    UPLOAD_FOLDER,
    s3_bucket=S3_BUCKET,
    region=AWS_REGION,
    url_expiry=S3_URL_EXPIRY,
)
if S3_BUCKET:
    app.logger.info(f'Photo storage: S3 bucket "{S3_BUCKET}" ({AWS_REGION})')
else:
    app.logger.info('Photo storage: local static/uploads (set S3_BUCKET to use S3)')

# Initialize SurveyTypeService
SURVEY_TYPES_FILE = os.path.join(DATA_FOLDER, 'survey_types.json')
from services.survey_type_service import SurveyTypeService
survey_type_service = SurveyTypeService(SURVEY_TYPES_FILE)

# Initialize QuestionFilterService
from services.question_filter import QuestionFilterService
question_filter_service = QuestionFilterService(question_manager, survey_type_service)

# Initialize RegionService (regional structure: leadership + community assignments)
REGIONS_FILE = os.path.join(DATA_FOLDER, 'regions.json')
from services.region_service import RegionService
region_service = RegionService(REGIONS_FILE)

# Corporate: a non-geographic group whose members work across the whole
# organization (they can visit any community and see every region). It lives
# alongside the regions so it shows up in the Regions view, but it never owns
# communities — its reach comes from regional_communities() below.
CORPORATE_ID = 'corporate'
CORPORATE_KIND = 'corporate'
try:
    region_service.ensure_group(CORPORATE_ID, 'Corporate', CORPORATE_KIND)
except Exception as _e:            # never block startup over this
    app.logger.error(f'Could not ensure Corporate group: {_e}')

# Initialize ActivityService (audit log) and ProfileService (per-user photo)
ACTIVITY_FILE = os.path.join(DATA_FOLDER, 'activity.json')
from services.activity_service import ActivityService
activity_service = ActivityService(ACTIVITY_FILE)

PROFILES_FILE = os.path.join(DATA_FOLDER, 'profiles.json')
from services.profile_service import ProfileService
profile_service = ProfileService(PROFILES_FILE)

# Who signed in, and who is using the app right now (writes are throttled)
PRESENCE_FILE = os.path.join(DATA_FOLDER, 'presence.json')
from services.presence_service import PresenceService
presence_service = PresenceService(PRESENCE_FILE)

# Admin-created login accounts (persisted; survives restarts/deploys)
USERS_FILE = os.path.join(DATA_FOLDER, 'users.json')
from services.user_service import UserService
user_service = UserService(USERS_FILE)

# Admin-managed resource library (guides, training, FAQ; files or links)
RESOURCES_FILE = os.path.join(DATA_FOLDER, 'resources.json')
from services.resource_service import ResourceService
resource_service = ResourceService(RESOURCES_FILE)

# Admin-uploaded community cover images (shown on dashboard cards + detail panel)
COVERS_FILE = os.path.join(DATA_FOLDER, 'community_covers.json')
from services.community_cover_service import CommunityCoverService
community_cover_service = CommunityCoverService(COVERS_FILE)

# Move-In module: editable checklist template + per-resident move-in records
MOVEIN_TEMPLATE_FILE = os.path.join(DATA_FOLDER, 'movein_template.json')
# Things a community raises for itself, as opposed to what a regional finds on
# a visit. Kept in their own store because they aren't part of any visit.
from services.raised_item_service import RaisedItemService
RAISED_FILE = os.path.join(DATA_FOLDER, 'raised_items.json')
raised_item_service = RaisedItemService(RAISED_FILE)

MOVEINS_FILE = os.path.join(DATA_FOLDER, 'moveins.json')
from services.move_in_service import MoveInTemplateService, MoveInService
movein_template_service = MoveInTemplateService(MOVEIN_TEMPLATE_FILE)
movein_service = MoveInService(MOVEINS_FILE)


def community_slug(name: str) -> str:
    """Mirror of the client-side slug: lowercase, non-alphanumerics -> '_'."""
    import re as _re
    s = (name or '').lower()
    s = _re.sub(r'[^a-z0-9]+', '_', s)
    return s.strip('_')


def cover_url_for(record):
    """Build a usable image URL for a stored cover record (S3 signed or local)."""
    if not record or not record.get('path'):
        return None
    path = record['path']
    if file_upload_handler.use_s3:
        return file_upload_handler.generate_presigned_url(path)
    return url_for('static', filename=f'uploads/{path}')

# Email (Amazon SES). Disabled until MAIL_FROM is set, so the app runs fine
# before email is configured. A failed send never blocks an inspection.
from services.email_service import EmailService
_extra = [a for a in os.environ.get('MAIL_EXTRA_RECIPIENTS', '').split(',') if a.strip()]
email_service = EmailService(
    mail_from=os.environ.get('MAIL_FROM'),
    region=os.environ.get('SES_REGION') or AWS_REGION,
    extra_recipients=_extra,
    app_base_url=os.environ.get('APP_BASE_URL'),
    configuration_set=os.environ.get('SES_CONFIGURATION_SET'),
)
# Keep the mail log at INFO so every send is recorded in journalctl with its
# SES MessageId. Delivery problems almost always live at the recipient's mail
# server, and that id is the only way to trace a message there after the fact.
import logging as _logging
_logging.getLogger('services.email_service').setLevel(_logging.INFO)
if not _logging.getLogger().handlers:
    _logging.basicConfig(level=_logging.INFO,
                         format='%(asctime)s %(levelname)s %(name)s: %(message)s')
app.logger.info('Email: SES enabled' if email_service.enabled
                else 'Email: disabled (set MAIL_FROM to enable)')

# Runtime-editable settings (email recipient lists, etc.)
SETTINGS_FILE = os.path.join(DATA_FOLDER, 'settings.json')
from services.settings_service import SettingsService
settings_service = SettingsService(SETTINGS_FILE)
# First run: seed subscribers (all-regions) from MAIL_EXTRA_RECIPIENTS if empty.
settings_service.seed_subscribers(_extra)

# Avatars folder for uploaded profile photos
AVATARS_FOLDER = os.path.join(UPLOAD_FOLDER, '..', 'avatars')
AVATARS_FOLDER = os.path.normpath(AVATARS_FOLDER)
os.makedirs(AVATARS_FOLDER, exist_ok=True)



# ==================== DATABASE & USER MANAGEMENT ====================

# Sample user database - In production, use a real database
# Format: {username: {'password_hash': hash, 'community': 'Community Name'}}
USERS_DB = {
    # Administrator. The per-community "userN" test accounts that used to live
    # here were removed before launch: they all shared one weak password and
    # nobody used them — real inspections are done by named regional and
    # corporate accounts created from the Regions view.
    'admin': {
        'password': 'admin123',
        'community': None  # Admin can see all communities
    },
}

# List of all available communities
ALL_COMMUNITIES = [
    # Georgia
    "Kelley Place, Enterprise",
    "Madison Heights Enterprise, Enterprise",
    "Monark Grove Madison",
    "Monark Grove Greystone",
    "Legacy Ridge Trussville, Trussville",
    "Madison at The Range, Madison",
    "The Goldton at Athens",
    "The Goldton at Jones Farm",
    
    # Florida
    "Madison at Clermont, Clermont",
    "Madison at Ocoee, Ocoee",
    "Madison at Oviedo, Oviedo",
    "The Goldton at Venice, Venice",
    "The Goldton at St. Petersburg, St. Petersburg",
    "Lake Howard Heights, Winter Haven",
    "The Canopy At Beacon Woods",
    "The Goldton At Lake Nona",
    
    # North Carolina
    "Madison Heights Evans, Evans",
    "Legacy at Savannah Quarters, Pooler",
    "Legacy Reserve at Old Town, Columbus",
    "Legacy Ridge at Alpharetta, Alpharetta",
    "Legacy Ridge at Buckhead, Atlanta",
    "Legacy Ridge at Marietta, Marietta",
    "The Canopy at Westridge, McDonough",
    "The Overlook at Suwanee",
    
    # Ohio
    "Legacy Reserve at Fritz Farm, Lexington",
    
    # Mississippi
    "The Goldton at Southaven, Southaven",
    "The Goldton at Adelaide, Starkville",
    
    # South Carolina
    "Oakview Park, Greenville",
    "Spring Park, Travelers Rest",
    "Legacy Reserve Fairview Park, Simpsonville",
    "Wildcat Senior Living, Summerville",
    
    # Tennessee
    "The Goldton at Spring Hill, Spring Hill",
    
    # Texas
    "The Oscar at Georgetown",
    "The Oscar at Veramendi",
    
    # Maryland
    "Tribute at Black Hill",
    "Tribute at Melford",
    
    # Virginia
    "Tribute at One Loudoun",
    "Tribute at The Glen",

    # Transitioning in (DMV)
    "The Goldton at Stuart"
]


# Default password for the auto-generated regional (per-person) accounts.
REGIONAL_DEFAULT_PASSWORD = 'atlas123'


def slugify_name(name):
    """Turn 'Keith Martin' into a stable login slug 'keith.martin'."""
    import re
    s = (name or '').strip().lower()
    s = re.sub(r'[^a-z0-9]+', '.', s)
    return s.strip('.')


def get_regional_accounts():
    """
    Build per-person regional login accounts from the region leadership.
    Each leader becomes an account: username = name slug, covering all the
    communities in their region. Computed live from regions.json so edits to
    leadership are reflected immediately.
    Returns: { username: {'display_name','role','region_id','region_name','communities'} }
    """
    accounts = {}
    for region in region_service.get_all_regions():
        if region.get('id') == 'unassigned':
            continue
        for leader in region.get('leadership', []):
            name = (leader.get('name') or '').strip()
            if not name or name.lower() == 'open':
                continue
            # Prefer the pinned username so renaming a person never moves their
            # account; fall back to the name slug for rosters not yet backfilled.
            username = (leader.get('username') or '').strip() or slugify_name(name)
            if not username:
                continue
            is_corp = region.get('kind') == CORPORATE_KIND
            accounts[username] = {
                'display_name': name,
                'role': 'regional',
                'region_id': region.get('id'),
                'region_name': region.get('name'),
                'corporate': is_corp,
                # Carried through so password resets and notifications can
                # actually reach region/corporate members.
                'email': (leader.get('email') or '').strip(),
                # Corporate members work across the whole organization.
                'communities': all_communities() if is_corp else list(region.get('communities', []))
            }
    return accounts


def authenticate_user(username, password):
    """
    Authenticate a user and return their account info on success.
    Order: admin/staff (USERS_DB), then per-person regional accounts.
    A hashed override (from change-password) takes precedence over the seed.
    Returns: (True, account_dict) on success, (False, None) on failure.
    account_dict = {role, community, region_id, display_name}
    """
    # --- Admin / staff (community accounts) ---
    if username in USERS_DB:
        user = USERS_DB[username]
        override = profile_service.get_password_hash(username)
        ok = check_password_hash(override, password) if override else (user['password'] == password)
        if not ok:
            return (False, None)
        community = user['community']
        return (True, {
            'role': 'admin' if community is None else 'staff',
            'community': community,
            'region_id': None,
            'display_name': profile_service.get_display_name(username) or username
        })

    # --- Admin-created accounts (users.json) ---
    custom = user_service.get(username)
    if custom:
        override = profile_service.get_password_hash(username)
        stored = custom.get('password_hash')
        if override:
            ok = check_password_hash(override, password)
        elif stored:
            ok = check_password_hash(stored, password)
        else:
            ok = False
        if not ok:
            return (False, None)
        return (True, {
            'role': custom.get('role', 'staff'),
            'community': custom.get('community'),
            # An account may cover several communities; `community` remains the
            # primary one so anything reading a single value still works.
            'communities': custom.get('communities')
                           or ([custom.get('community')] if custom.get('community') else []),
            'region_id': custom.get('region_id'),
            'display_name': profile_service.get_display_name(username)
                            or custom.get('display_name') or username
        })

    # --- Regional (per-person) accounts ---
    regionals = get_regional_accounts()
    if username in regionals:
        acct = regionals[username]
        override = profile_service.get_password_hash(username)
        ok = check_password_hash(override, password) if override else (password == REGIONAL_DEFAULT_PASSWORD)
        if not ok:
            return (False, None)
        return (True, {
            'role': 'regional',
            'community': None,
            'region_id': acct['region_id'],
            'display_name': acct['display_name']
        })

    return (False, None)


def username_taken(candidate):
    """True if a username already exists in any of the three account sources."""
    if candidate in USERS_DB:
        return True
    if user_service.exists(candidate):
        return True
    if candidate in get_regional_accounts():
        return True
    return False


def resolve_account_context(username):
    """Resolve an account across all three sources for reset flows.
    Returns dict {exists, display_name, context, email} or {exists: False}."""
    if username in USERS_DB:
        u = USERS_DB[username]
        comm = u.get('community')
        return {'exists': True,
                'display_name': profile_service.get_display_name(username) or username,
                'context': 'Administrator' if comm is None else f'Staff · {comm}',
                'email': None}
    custom = user_service.get(username)
    if custom:
        comm = custom.get('community')
        role = custom.get('role', 'staff')
        ctx = role.capitalize() + (f' · {comm}' if comm else '')
        return {'exists': True,
                'display_name': profile_service.get_display_name(username) or custom.get('display_name') or username,
                'context': ctx, 'email': custom.get('email')}
    regionals = get_regional_accounts()
    if username in regionals:
        acct = regionals[username]
        return {'exists': True,
                'display_name': acct.get('display_name') or username,
                'context': f"Regional · {acct.get('region_id', '')}",
                'email': acct.get('email')}
    return {'exists': False}


def backfill_leader_usernames():
    """Pin a permanent username on every region/corporate member that doesn't
    have one yet. Existing rosters were keyed off the person's name, so this
    freezes today's derived login before anyone can rename them."""
    changed = 0
    for region in region_service.get_all_regions():
        if region.get('id') == 'unassigned':
            continue
        for i, leader in enumerate(region.get('leadership', [])):
            if (leader.get('username') or '').strip():
                continue
            name = (leader.get('name') or '').strip()
            if not name or name.lower() == 'open':
                continue
            uname = slugify_name(name)
            if uname and region_service.set_leader_username(region.get('id'), i, uname):
                changed += 1
    return changed


def retire_removed_builtin_users():
    """Delete stored copies of built-in accounts that no longer exist in code.

    The per-community "userN" test logins were retired before launch (they all
    shared one weak password). This removes their migrated records so they can't
    be used, while leaving admin-created accounts untouched. Submitted
    inspections are unaffected — they keep the inspector recorded on them."""
    removed = 0
    for u in user_service.get_all():
        username = u.get('username')
        rec = user_service.get(username) or {}
        if rec.get('builtin') and username not in USERS_DB:
            if user_service.delete(username):
                removed += 1
    return removed


def migrate_builtin_users():
    """Copy the built-in accounts (admin + community users defined in code) into
    editable storage so they can be managed from the People view like everyone
    else. Idempotent: only creates records that don't exist yet."""
    created = 0
    for username, info in USERS_DB.items():
        if user_service.exists(username):
            continue
        community = info.get('community')
        made = user_service.ensure(
            username,
            display_name=profile_service.get_display_name(username) or username,
            role='admin' if community is None else 'staff',
            community=community,
            password_hash=generate_password_hash(info.get('password', '')),
            created_by='system', builtin=True)
        if made:
            created += 1
    return created


def generate_unique_username(base):
    """Slug from a name, with a numeric suffix if it collides."""
    base = slugify_name(base) or 'user'
    candidate = base
    n = 2
    while username_taken(candidate):
        candidate = f"{base}{n}"
        n += 1
    return candidate


def generate_password(length=10):
    """Readable strong password (avoids ambiguous characters)."""
    import secrets
    alphabet = 'abcdefghjkmnpqrstuvwxyzABCDEFGHJKMNPQRSTUVWXYZ23456789'
    return ''.join(secrets.choice(alphabet) for _ in range(length))


# One-time (idempotent) data upkeep so every account is editable and stable.
try:
    _pinned = backfill_leader_usernames()
    _migrated = migrate_builtin_users()
    _retired = retire_removed_builtin_users()
    if _pinned or _migrated or _retired:
        app.logger.info('People upkeep: pinned %d leader usernames, migrated %d built-in users, '
                        'retired %d obsolete test accounts', _pinned, _migrated, _retired)
except Exception as _e:
    app.logger.error(f'People upkeep failed: {_e}')


# ===================== Previewing the app as an Executive Director =====================
#
# An administrator can look at the app exactly as one community sees it —
# useful for training, and for answering "what does my ED actually see?"
# without creating a fake account that would then receive real emails and
# appear as that community's director.
#
# Two rules make it safe, and both are enforced in one place each rather than
# screen by screen:
#   - Only a real administrator can start it (real_is_admin, below).
#   - Nothing can be written while it is on (_preview_is_read_only).
#
# While previewing, is_admin() answers False on purpose: the point is to see
# what they see, and an admin menu on screen would defeat that. Stopping the
# preview therefore cannot depend on is_admin() — it reads the stored role
# directly, which the preview never touches.

def viewing_as():
    """The preview in progress, or None.

    {'communities': [...], 'label': 'Jazmyn Frazier'} — a list, because an
    Executive Director can stand in for a neighbouring community and a preview
    that showed only one of them would be showing a different job than the one
    that person actually does."""
    v = session.get('view_as')
    return v if isinstance(v, dict) and v.get('communities') else None


def real_is_admin():
    """Admin status ignoring any preview in progress.

    Everything else asks is_admin(), which deliberately says False while a
    preview is running. Starting and stopping a preview has to ask this one, or
    an administrator could enter a preview and never get out of it."""
    if session.get('role') == 'admin':
        return True
    return bool(session.get('admin_extra'))


def current_role():
    # A preview makes every downstream check — capabilities, scoping, which
    # menu items appear — behave as it does for that community.
    if viewing_as():
        return 'staff'
    return _stored_role()


def _stored_role():
    """Resolve the current session role, with backward-compatible fallback."""
    role = session.get('role')
    if role:
        return role
    # Older sessions without an explicit role: infer from community
    return 'admin' if session.get('community') is None else 'staff'


# The two roles that cover communities rather than belong to one. They behave
# identically everywhere that matters — a Corporate member is a regional whose
# region happens to be the whole company, which is what regional_communities()
# already encodes. Several checks tested only for 'regional', so an account
# stored as 'corporate' fell through to the community-account branch and ended
# up with no communities, no visits and no leaderboard.
LEADERSHIP_ROLES = ('regional', 'corporate')


def is_leadership(role=None):
    """True for regional and corporate accounts."""
    return (role or current_role()) in LEADERSHIP_ROLES


def is_native_admin():
    """True only for real Administrator accounts. The admin accessory does NOT
    count here — granting privileges is reserved for the main administrator.

    False while previewing as a community, so the preview shows what that
    community sees rather than an administrator's view of it."""
    if viewing_as():
        return False
    return current_role() == 'admin'


def is_admin():
    """True when the user may perform administrative actions — either because
    they are an Administrator, or because the main administrator granted them
    the admin accessory on top of their own role (e.g. a Corporate member)."""
    if viewing_as():
        return False
    if is_native_admin():
        return True
    return bool(session.get('admin_extra'))


def requested_communities(data, primary):
    """The communities an admin picked for a community account.

    Accepts a list, and always keeps the primary one first so anything reading
    a single value gets the main site. Unknown names are dropped rather than
    trusted — this is what stops a request from granting itself a community."""
    known = set(all_communities())
    out = []
    if primary:
        out.append(primary)
    for c in (data.get('communities') or []):
        name = InputSanitizer.sanitize_community_name(c or '')
        if name and name in known and name not in out:
            out.append(name)
    return [c for c in out if c in known]


def account_communities(u):
    """Which communities a stored account covers, tolerating older records
    that only ever had a single one."""
    comms = u.get('communities')
    if comms:
        return [c for c in comms if c]
    one = (u.get('community') or '').strip()
    return [one] if one else []


def session_communities():
    """Every community this session covers.

    Community accounts used to be one-per-site, and a dozen checks compared a
    single string. An ED can now stand in for a neighbouring community, so the
    answer is a list — with the old single value as a fallback for sessions
    created before this existed."""
    # A preview covers exactly what that account covers.
    previewing = viewing_as()
    if previewing:
        return list(previewing['communities'])
    comms = session.get('communities')
    if comms:
        return [c for c in comms if c]
    one = session.get('community')
    return [one] if one else []


def can_run_visits():
    """Who may carry out a visit.

    Regional and corporate staff only. A community must not inspect itself:
    the score comes from the most recent visit, so a self-run walkthrough
    would overwrite what the regional found — including the open items."""
    return is_leadership()


def can_verify_fixes():
    """Who may close out a failed standard.

    Community-level accounts (Executive Directors) report a fix by commenting;
    a regional, corporate member or admin reviews it and marks it addressed.
    Keeping those two apart is what makes the current score trustworthy — a
    community can never raise its own number."""
    return is_admin() or is_leadership()


def all_communities():
    """Every community assigned to any region (org-wide scope)."""
    names = []
    for r in region_service.get_all_regions():
        for c in r.get('communities', []):
            if c not in names:
                names.append(c)
    return names


def regional_communities():
    """Communities the current regional user may inspect.

    Normally that's the communities of their own region. Members of a
    corporate-style group are not tied to a geography, so they get the whole
    organization — this single function is what gives Corporate its reach."""
    region_id = session.get('region_id')
    if not region_id:
        return []
    region = next((r for r in region_service.get_all_regions() if r.get('id') == region_id), None)
    if not region:
        return []
    if region.get('kind') == CORPORATE_KIND:
        return all_communities()
    return list(region.get('communities', []))


def region_for_community(community):
    """The region dict that owns this community, or None."""
    for r in region_service.get_all_regions():
        if community in (r.get('communities') or []):
            return r
    return None


def community_account_emails(community, exclude_username=None):
    """Email addresses of the accounts that belong to a community itself —
    in practice its Executive Director. Empty until those accounts exist, so
    this is safe to call before anyone is onboarded."""
    community = (community or '').strip()
    if not community:
        return []
    out = []
    for u in user_service.get_all():
        if u.get('role') != 'staff' or community not in account_communities(u):
            continue
        if exclude_username and u.get('username') == exclude_username:
            continue
        addr = (u.get('email') or '').strip()
        if addr and addr not in out:
            out.append(addr)
    return out


def local_dt(dt=None):
    """Move a moment into the timezone Atlas actually works in.

    The server runs on UTC, so anything formatted straight from the clock
    reads hours off for the person opening the email. Timestamps stay UTC in
    storage — this is only for display. Set APP_TIMEZONE if HQ isn't Central."""
    from datetime import timezone as _tz
    dt = dt or datetime.now(_tz.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=_tz.utc)
    try:
        from zoneinfo import ZoneInfo
        return dt.astimezone(ZoneInfo(os.environ.get('APP_TIMEZONE', 'America/Chicago')))
    except Exception:
        return dt


def fmt_local(dt=None, fmt='%b %d, %Y %I:%M %p %Z'):
    return local_dt(dt).strftime(fmt).strip()


def role_label_for(role, community=None):
    """How a role reads to a person, not how it's stored.

    One place for this on purpose: the label was being built separately in each
    email, so fixing one left the others saying "Staff"."""
    if role == 'staff':
        return f"Executive Director · {community}" if community else 'Executive Director'
    return {'admin': 'Administrator',
            'regional': 'Regional',
            'corporate': 'Corporate'}.get(role, (role or '').capitalize())


def send_community_handover(community, recipients):
    """Bring a new community account up to speed on what is already open.

    The work belongs to the community, not to whoever held the role, so someone
    taking over inherits every open item and comment thread. This is the email
    that tells them so, instead of letting them find a backlog by accident."""
    if not email_service.enabled or not recipients:
        return
    subs = sorted(inspection_service.get_submissions_by_community(community),
                  key=lambda s: s.get('submitted_at', ''), reverse=True)
    latest = subs[0] if subs else None
    failed, open_items, inspector, when = [], [], '', ''
    if latest:
        failed = [r for r in (latest.get('responses') or [])
                  if r.get('condition') == 'Fail' and not r.get('addressed')]
        open_items = [i for i in (latest.get('action_items') or []) if not i.get('resolved')]
        inspector = latest.get('inspector_name') or resolve_display_name(latest.get('username', ''))
        when = (latest.get('submitted_at') or '')[:10]

    criteria_map = {}
    for q in question_manager.get_all_active_questions():
        crit = q.get('pass_criteria') or []
        if q.get('id'):
            criteria_map[q['id']] = crit
        if q.get('text'):
            criteria_map['t:' + q['text'].strip().lower()] = crit

    email_service.send_community_findings(
        recipients, community, inspector, when, failed, open_items,
        criteria_map, handover=True)


def region_leader_emails(community):
    """Email addresses of the leadership for the region that owns this community."""
    r = region_for_community(community)
    if not r:
        return []
    return [(l.get('email') or '').strip()
            for l in (r.get('leadership') or [])
            if (l.get('email') or '').strip()]


def movein_recipients(community):
    """Who is emailed about a move-in.

    The community itself, plus the administrator notification list.

    Deliberately not the region's leadership. A move-in is the community's own
    work, and a regional covering a dozen communities was receiving forty to
    fifty of these a month — which is how a mailbox teaches someone to ignore a
    sender. Regionals keep full access to every move-in in their region under
    Move-Ins, and the daily summary still reports any that are past their date
    with required items open, so nothing is hidden from them. They just aren't
    told one at a time.

    If a community has no account yet, the region's leadership is used after
    all: better a regional hears about it than nobody does."""
    community = (community or '').strip()
    admin_notify = settings_service.get_email_settings().get('admin_notify', [])
    to = community_account_emails(community)
    if not to:
        to = region_leader_emails(community)
    return list(dict.fromkeys(to + admin_notify))


def alert_password_changed(username, changed_by=''):
    """Tell the administrators that an account's password changed.

    Security notices are worth sending but never worth failing the change
    itself, so every error here is logged and swallowed."""
    try:
        admin_notify = settings_service.get_email_settings().get('admin_notify', [])
        if not admin_notify:
            return
        email_service.send_password_changed_alert(
            admin_notify,
            resolve_display_name(username), username,
            changed_by=changed_by,
            when=fmt_local(),
            ip=_client_ip() if not changed_by else '')
    except Exception as e:
        app.logger.error(f'Password-change alert failed: {e}')


def leadership_names():
    """All distinct regional leadership names (people who perform inspections)."""
    names = set()
    for r in region_service.get_all_regions():
        for l in (r.get('leadership') or []):
            n = (l.get('name') or '').strip()
            if n and n.lower() != 'open':
                names.add(n)
    return sorted(names)


def _display_name_for(username):
    """The best name we have for whoever is acting right now.

    resolve_display_name() ends by returning the username when it finds nothing,
    which would quietly store "jazmyn.frasier" where a person's name belongs —
    and that string is then read for years. The session carries the real name
    from sign-in, so try that before settling."""
    resolved = resolve_display_name(username)
    if resolved and resolved != username:
        return resolved
    return session.get('display_name') or resolved or username


def resolve_display_name(username):
    """Friendly name for a username (profile override, regional, or username)."""
    if not username:
        return ''
    name = profile_service.get_display_name(username)
    if name:
        return name
    regionals = get_regional_accounts()
    if username in regionals:
        return regionals[username]['display_name']
    return username


def login_required(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def require_admin(f):
    """
    Decorator to require admin role for routes
    Admin users have community set to None
    Non-admin users are redirected to the inspection form
    """
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        # Only admins may access admin routes
        if not is_admin():
            # API endpoints must answer with JSON 403 so the frontend's fetch()
            # can surface the error, instead of silently redirecting to a page.
            if request.path.startswith('/api/'):
                return jsonify({'status': 'error', 'message': 'Admin access required'}), 403
            return redirect(url_for('report_form'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== ROUTES ====================

@app.route('/login')
def login():
    """
    Display the login page
    """
    # If already logged in, redirect to report form
    if 'user' in session:
        return redirect(url_for('report_form'))
    return render_template('login.html')


@app.route('/api/login', methods=['POST'])
def api_login():
    """
    API endpoint for user authentication
    Expects JSON with username and password
    Returns user info and their assigned community
    
    Error Handling:
    - 400: JSON parsing errors or missing fields
    - 401: Invalid credentials
    - 500: Internal server errors
    """
    try:
        # Throttle brute-force attempts (per client IP).
        if login_is_throttled():
            return jsonify({
                'status': 'error',
                'message': 'Too many failed attempts. Please wait a few minutes and try again.'
            }), 429

        # Handle JSON parsing errors
        data = request.get_json(silent=True)

        if data is None:
            return jsonify({
                'status': 'error',
                'message': 'Invalid JSON format or Content-Type must be application/json'
            }), 400

        # Validate JSON structure
        if not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({
                'status': 'error',
                'message': 'Request body must be a JSON object'
            }), 400

        # Sanitize and validate inputs
        username = InputSanitizer.sanitize_username(data.get('username', ''))
        password = data.get('password', '')
        
        if not username:
            return jsonify({
                'status': 'error',
                'message': 'Username is required'
            }), 400
        
        if not password:
            return jsonify({
                'status': 'error',
                'message': 'Password is required'
            }), 400

        # Authenticate user
        success, account = authenticate_user(username, password)

        if success:
            reset_login_failures()
            # Store user in session
            session['user'] = username
            session['community'] = account['community']
            session['communities'] = account.get('communities') or (
                [account['community']] if account['community'] else [])
            session['role'] = account['role']
            session['region_id'] = account['region_id']
            session['display_name'] = account['display_name']
            session.permanent = True
            # Administrator privileges granted on top of their role (accessory).
            session['admin_extra'] = profile_service.get_admin_extra(username)
            # If an admin reset this account, force a password change before use.
            must_change = profile_service.get_must_change(username)
            session['must_change'] = bool(must_change)

            # Record the sign-in for the People directory and the activity feed.
            presence_service.record_login(username)
            activity_service.log(username, 'login', 'Signed in',
                                 meta={'ip': _client_ip()})

            return jsonify({
                'status': 'success',
                'message': 'Login successful',
                'username': username,
                'community': account['community'],
                'role': account['role'],
                'display_name': account['display_name'],
                'must_change_password': bool(must_change)
            }), 200
        else:
            record_login_failure()
            return jsonify({
                'status': 'error',
                'message': 'Invalid username or password'
            }), 401

    except Exception as e:
        # Log the error for debugging
        app.logger.error(f'Login error: {str(e)}')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error during login'
        }), 500


_forgot_attempts = {}


@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    """Public: a user asks for a password reset. We never reveal whether the
    account exists; if it does, the configured admins are emailed so they can
    reset it (admin-assisted flow). Lightly rate-limited per IP."""
    from datetime import datetime, timedelta
    generic = jsonify({'status': 'success',
                       'message': "If that account exists, your administrator has been notified."})
    try:
        ip = request.remote_addr or 'unknown'
        now = datetime.now()
        hits = [t for t in _forgot_attempts.get(ip, []) if now - t < timedelta(minutes=10)]
        if len(hits) >= 5:
            return jsonify({'status': 'error',
                            'message': 'Too many requests. Please try again later.'}), 429
        hits.append(now)
        _forgot_attempts[ip] = hits

        data = request.get_json(silent=True) or {}
        username = InputSanitizer.sanitize_username(data.get('username', ''))
        if not username:
            return generic, 200

        acct = resolve_account_context(username)
        if acct.get('exists'):
            admin_notify = settings_service.get_email_settings().get('admin_notify', [])
            if admin_notify:
                try:
                    email_service.send_password_reset_request(
                        admin_notify, acct.get('display_name'), username,
                        acct.get('context', ''), fmt_local(now))
                except Exception as e:
                    app.logger.error(f'Reset-request email failed: {str(e)}')
            activity_service.log(username, 'password_reset_requested',
                                 f"Password reset requested for {acct.get('display_name') or username}")
        return generic, 200
    except Exception as e:
        app.logger.error(f'forgot_password error: {str(e)}')
        return generic, 200


@app.route('/logout')
def logout():
    """
    Logout user and clear session
    """
    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password')
def change_password_page():
    """Standalone page to set a new password. Shown (and required) right after
    an admin reset, and reachable by anyone signed in who wants to change it."""
    if not session.get('user'):
        return redirect(url_for('login'))
    return render_template('change_password.html',
                           forced=bool(session.get('must_change')),
                           display_name=session.get('display_name') or session.get('user'))


@app.route('/')
def index():
    """
    Root route - redirect based on authentication and user type
    
    Behavior:
    - Not authenticated: Redirect to login
    - Admin user: Redirect to dashboard
    - Staff user: Redirect to survey type selection (start new visit)
    """
    if 'user' not in session:
        return redirect(url_for('login'))

    # Route by role
    if is_admin():
        return redirect(url_for('dashboard'))
    else:
        # Staff and regional users start a visit via survey type selection
        return redirect(url_for('select_survey_type'))


@app.route('/reporte')
@login_required
def report_form():
    """
    Render the mobile report form (reporte.html)
    This page is designed for maintenance/cleaning staff using mobile devices
    User must be logged in and their community is automatically detected
    
    Survey Type Requirement:
    - User must have selected a survey type before accessing this page
    - If no survey type in session, redirect to survey type selection
    - Admin users are redirected to dashboard (cannot submit inspections)
    """
    # Visits are carried out by regional and corporate staff. Administrators
    # don't run them, and neither do community accounts — a community cannot
    # inspect itself. Note this checks the *native* role: someone whose real
    # role is Corporate/Regional keeps inspecting even with admin privileges.
    if is_native_admin() or not can_run_visits():
        return redirect(url_for('dashboard'))

    # Check if survey type is selected
    survey_type_id = session.get('survey_type_id')
    if not survey_type_id:
        # No survey type selected, redirect to selection screen
        return redirect(url_for('select_survey_type'))
    
    # Validate survey type is still valid
    if not survey_type_service.validate_survey_type(survey_type_id):
        # Invalid survey type in session, clear and redirect
        session.pop('survey_type_id', None)
        session.pop('survey_type_name', None)
        return redirect(url_for('select_survey_type'))
    
    # Get survey type details for display
    survey_type = survey_type_service.get_survey_type_by_id(survey_type_id)
    
    # Communities the user may report on: regionals pick from their whole
    # region; staff are locked to their single community.
    if is_leadership():
        communities = regional_communities()
    else:
        communities = session_communities()

    return render_template('reporte.html',
                         community=session.get('community'),
                         communities=communities,
                         role=current_role(),
                         username=session.get('display_name') or session.get('user'),
                         survey_type=survey_type,
                         survey_type_id=survey_type_id)


@app.route('/select-survey-type')
@login_required
def select_survey_type():
    """
    Render the survey type selection screen
    User must select a survey type before starting an inspection
    Requires login - admin users are redirected to dashboard
    """
    # Same rule as the form itself: only regional and corporate staff run visits.
    if is_native_admin() or not can_run_visits():
        return redirect(url_for('dashboard'))

    # Regional and corporate users have no fixed community — they choose it on
    # the form itself. Pass an empty string so the page can say something
    # useful instead of printing "None".
    return render_template('select_survey_type.html',
                         community=session.get('community') or '',
                         username=session.get('user'))


@app.route('/dashboard')
@login_required
def dashboard():
    """
    Render the admin dashboard (dashboard.html)
    This page is designed for managers to view reports from a desktop
    Admin users can see all communities
    """
    # nav_admin decides whether the sidebar offers the admin-only sections.
    #
    # It deliberately uses is_admin() rather than the old "community is None"
    # shorthand this route used to rely on: a regional also has no community,
    # so that test called every regional an admin. Nothing read the value
    # before, which is why it went unnoticed — the sidebar reads it now.
    return render_template('dashboard.html',
                         username=session.get('user'),
                         is_admin=is_admin(),
                         nav_admin=is_admin(),
                         community=session.get('community'))


@app.route('/api/submit-report', methods=['POST'])
@login_required
def submit_report():
    """
    API endpoint to receive report submissions from the mobile form
    Handles file upload with proper image processing and storage
    
    Error Handling:
    - 400: Missing required fields, invalid file types/sizes
    - 500: Internal server errors (file system, etc.)
    """
    try:
        # Get and sanitize form data
        community = InputSanitizer.sanitize_community_name(request.form.get('community', ''))
        location = InputSanitizer.sanitize_string(request.form.get('location', ''), max_length=200)
        condition = InputSanitizer.sanitize_string(request.form.get('condition', ''), max_length=50)
        description = InputSanitizer.sanitize_description(request.form.get('description', ''))
        username = InputSanitizer.sanitize_username(session.get('user', ''))

        # Validate required fields
        if not all([community, location, condition, description]):
            return jsonify({
                'status': 'error',
                'message': 'Missing required fields'
            }), 400

        # Handle file upload
        file_path = None
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename != '':
                # Validate file
                if not allowed_file(file.filename):
                    return jsonify({
                        'status': 'error',
                        'message': 'Invalid file type. Only images are allowed.'
                    }), 400

                try:
                    # Create secure filename with timestamp
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    file_ext = secure_filename(file.filename).rsplit('.', 1)[1].lower()
                    filename = f"{username}_{community}_{timestamp}.{file_ext}"
                    
                    # Create community folder if it doesn't exist
                    community_folder = os.path.join(app.config['UPLOAD_FOLDER'], 
                                                   secure_filename(community))
                    os.makedirs(community_folder, exist_ok=True)
                    
                    # Save file
                    file_path = os.path.join(community_folder, filename)
                    file.save(file_path)
                except IOError as e:
                    app.logger.error(f'File system error saving report photo: {str(e)}')
                    return jsonify({
                        'status': 'error',
                        'message': 'Internal server error: Failed to save photo'
                    }), 500

        # Create report record (in production, save to database)
        report = {
            'id': datetime.now().isoformat(),
            'username': username,
            'community': community,
            'location': location,
            'condition': condition,
            'description': description,
            'photo': file_path,
            'timestamp': datetime.now().isoformat()
        }

        # In production, save to database
        # For now, just return success
        return jsonify({
            'status': 'success',
            'message': 'Report submitted successfully',
            'report': report
        }), 200

    except Exception as e:
        app.logger.exception('Unexpected error submitting report')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while submitting report'
        }), 500


@app.route('/api/profile', methods=['GET'])
@login_required
def get_profile():
    """
    Profile data for the current user: identity, photo, activity stats, and
    recent activity feed.
    """
    try:
        username = session.get('user')
        community = session.get('community')
        role = current_role()
        # NOTE: don't shadow the module-level is_admin() helper.
        has_admin_access = is_admin()
        role_label = role_label_for(role, community)

        # Inspection count from the source of truth (historical-safe)
        try:
            all_subs = inspection_service.get_all_submissions()
            inspections = sum(1 for s in all_subs if s.get('username') == username)
        except Exception:
            inspections = 0

        stats = {
            'inspections': inspections,
            'questions_created': activity_service.count_for_user(username, 'question_created'),
            'questions_edited': activity_service.count_for_user(username, 'question_updated'),
            'total_actions': activity_service.count_for_user(username)
        }

        # For regionals, surface their region in place of a single community
        display_community = community
        if is_leadership(role):
            display_community = session.get('region_id') and \
                next((r.get('name') for r in region_service.get_all_regions()
                      if r.get('id') == session.get('region_id')), None)
            if display_community:
                display_community = f"{display_community} region"

        return jsonify({
            'status': 'success',
            'username': username,
            'display_name': profile_service.get_display_name(username) or session.get('display_name') or '',
            'community': display_community,
            'is_admin': has_admin_access,
            'role': role_label,
            'photo': profile_service.get_photo(username),
            'last_active': activity_service.last_active(username),
            'stats': stats,
            'recent_activity': activity_service.get_for_user(username, limit=15)
        }), 200
    except Exception as e:
        app.logger.exception('Error retrieving profile')
        return jsonify({'status': 'error', 'message': 'Internal server error while retrieving profile'}), 500


@app.route('/api/profile/name', methods=['POST'])
@login_required
def update_profile_name():
    """Update the current user's display name."""
    try:
        data = request.get_json(silent=True)
        if data is None or not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({'status': 'error', 'message': 'Request body must be a JSON object'}), 400

        username = session.get('user')
        display_name = InputSanitizer.sanitize_string(data.get('display_name', ''), max_length=80)
        profile_service.set_display_name(username, display_name)
        return jsonify({'status': 'success', 'display_name': display_name}), 200
    except Exception as e:
        app.logger.exception('Error updating display name')
        return jsonify({'status': 'error', 'message': 'Internal server error while updating name'}), 500


@app.route('/api/profile/password', methods=['POST'])
@login_required
def change_password():
    """
    Change the current user's password. Verifies the current password, then
    stores a securely hashed override.
    """
    try:
        data = request.get_json(silent=True)
        if data is None or not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({'status': 'error', 'message': 'Request body must be a JSON object'}), 400

        username = session.get('user')
        current_password = data.get('current_password', '')
        new_password = data.get('new_password', '')
        forced = bool(session.get('must_change'))

        if not new_password or (not forced and not current_password):
            return jsonify({'status': 'error', 'message': 'Current and new password are required'}), 400

        if len(new_password) < 6:
            return jsonify({'status': 'error', 'message': 'New password must be at least 6 characters'}), 400

        # In the forced flow the user just authenticated with the temp password,
        # so we don't ask for it again; otherwise verify the current password.
        if not forced:
            success, _ = authenticate_user(username, current_password)
            if not success:
                return jsonify({'status': 'error', 'message': 'Current password is incorrect'}), 400

        profile_service.set_password_hash(username, generate_password_hash(new_password))
        profile_service.set_must_change(username, False)
        session['must_change'] = False
        activity_service.log(username, 'password_changed', 'Changed account password')
        alert_password_changed(username)
        return jsonify({'status': 'success', 'message': 'Password updated successfully'}), 200
    except Exception as e:
        app.logger.exception('Error changing password')
        return jsonify({'status': 'error', 'message': 'Internal server error while changing password'}), 500


@app.route('/api/profile/photo', methods=['POST'])
@login_required
def upload_profile_photo():
    """Upload/replace the current user's profile photo."""
    try:
        username = InputSanitizer.sanitize_username(session.get('user', ''))
        if 'photo' not in request.files:
            return jsonify({'status': 'error', 'message': 'No photo provided'}), 400

        file = request.files['photo']
        if not file or file.filename == '':
            return jsonify({'status': 'error', 'message': 'No photo provided'}), 400

        if not allowed_file(file.filename):
            return jsonify({'status': 'error', 'message': 'Invalid file type. Only images are allowed.'}), 400

        ext = secure_filename(file.filename).rsplit('.', 1)[1].lower()
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"{secure_filename(username)}_{timestamp}.{ext}"
        try:
            file.save(os.path.join(AVATARS_FOLDER, filename))
        except IOError as e:
            app.logger.exception('Error saving avatar')
            return jsonify({'status': 'error', 'message': 'Failed to save photo'}), 500

        relative_path = f"avatars/{filename}"
        profile_service.set_photo(username, relative_path)
        return jsonify({'status': 'success', 'photo': relative_path}), 200
    except Exception as e:
        app.logger.exception('Error uploading profile photo')
        return jsonify({'status': 'error', 'message': 'Internal server error while uploading photo'}), 500


@app.route('/api/communities')
@login_required
def get_communities():
    """
    Communities the current user is allowed to see on the dashboard:
      - admin: all communities
      - regional: only their region's communities
      - staff: only their assigned community
    """
    role = current_role()
    if is_admin():
        # Derive from the regional structure (union, de-duped) so renames and
        # additions made in the Regions view are reflected. Every community
        # lives in a region (including the Unassigned bucket), so this is the
        # complete list. Fall back to the seed list only if regions are empty.
        seen, communities = set(), []
        for r in region_service.get_all_regions():
            for c in r.get('communities', []):
                if c not in seen:
                    seen.add(c); communities.append(c)
        if not communities:
            communities = list(ALL_COMMUNITIES)
    elif is_leadership(role):
        communities = regional_communities()
    else:
        communities = session_communities()
    return jsonify({'status': 'success', 'communities': communities}), 200


@app.route('/api/community-covers', methods=['GET'])
@login_required
def list_community_covers():
    """Return { slug: image_url } for every community that has a cover image.
    Any logged-in user can read these (covers are shown on the cards)."""
    out = {}
    for slug, rec in community_cover_service.get_all().items():
        url = cover_url_for(rec)
        if url:
            out[slug] = url
    return jsonify({'status': 'success', 'covers': out}), 200


@app.route('/api/communities/cover', methods=['POST'])
@require_admin
def upload_community_cover():
    """Admin-only: upload (or replace) a community's cover image.
    Multipart form: name (community name) + photo (image file)."""
    try:
        name = InputSanitizer.sanitize_community_name(request.form.get('name', ''))
        if not name:
            return jsonify({'status': 'error', 'message': 'Community name is required'}), 400
        if 'photo' not in request.files:
            return jsonify({'status': 'error', 'message': 'No image provided'}), 400

        file = request.files['photo']
        valid, msg = file_upload_handler.validate_file(file)
        if not valid:
            return jsonify({'status': 'error', 'message': msg}), 400

        slug = community_slug(name)
        # Remove any previous cover object (extension may differ) before saving.
        old = community_cover_service.get(slug)
        if old and old.get('path'):
            file_upload_handler.delete_file(old['path'])

        rel_path, stored = file_upload_handler.save_cover(file, slug)
        rec = community_cover_service.set(slug, name, rel_path, stored)
        activity_service.log(session.get('user'), 'community_cover_set',
                             f'Set cover image for "{name}"')
        return jsonify({'status': 'success', 'slug': slug,
                        'url': cover_url_for(rec)}), 200
    except Exception as e:
        app.logger.exception('Error uploading community cover')
        return jsonify({'status': 'error', 'message': 'Internal server error while uploading cover'}), 500


@app.route('/api/communities/cover', methods=['DELETE'])
@require_admin
def delete_community_cover():
    """Admin-only: remove a community's cover image (reverts to the gradient)."""
    try:
        data = request.get_json(silent=True) or {}
        name = InputSanitizer.sanitize_community_name(data.get('name', ''))
        if not name:
            return jsonify({'status': 'error', 'message': 'Community name is required'}), 400
        slug = community_slug(name)
        rec = community_cover_service.delete(slug)
        if rec and rec.get('path'):
            file_upload_handler.delete_file(rec['path'])
        activity_service.log(session.get('user'), 'community_cover_removed',
                             f'Removed cover image for "{name}"')
        return jsonify({'status': 'success', 'slug': slug}), 200
    except Exception as e:
        app.logger.exception('Error removing community cover')
        return jsonify({'status': 'error', 'message': 'Internal server error while removing cover'}), 500


@app.route('/api/people/profile')
@login_required
def person_profile():
    """Aggregated activity profile for one team member, computed across ALL
    communities (intentionally cross-region — performance profiles are visible
    to everyone to encourage healthy competition). Any logged-in user can view."""
    from datetime import datetime as _dt
    name = InputSanitizer.sanitize_string(request.args.get('name', ''), max_length=120)
    if not name:
        return jsonify({'status': 'error', 'message': 'name is required'}), 400

    # Match by inspector display name (captured at visit time) or username.
    subs = []
    for s in inspection_service.get_all_submissions():
        disp = s.get('inspector_name') or resolve_display_name(s.get('username', ''))
        if disp == name or s.get('username') == name:
            subs.append(s)

    now = _dt.now()
    visits_this_month = 0
    last_visit = ''
    passes = fails = 0
    by_comm, by_month, weak, recent = {}, {}, {}, []
    for s in subs:
        sa = s.get('submitted_at', '') or ''
        try:
            d = _dt.strptime(sa[:19], '%Y-%m-%dT%H:%M:%S')
        except ValueError:
            d = None
        if d and d.year == now.year and d.month == now.month:
            visits_this_month += 1
        if d:
            mk = d.strftime('%Y-%m')
            by_month[mk] = by_month.get(mk, 0) + 1
        if sa > last_visit:
            last_visit = sa
        comm = s.get('community', '')
        c = by_comm.setdefault(comm, {'visits': 0, 'pass': 0, 'fail': 0})
        c['visits'] += 1
        sp = sf = 0
        for r in (s.get('responses') or []):
            cond = r.get('condition')
            if cond == 'Pass':
                passes += 1; sp += 1; c['pass'] += 1
            elif cond == 'Fail':
                fails += 1; sf += 1; c['fail'] += 1
                q = r.get('question_text', '')
                if q:
                    weak[q] = weak.get(q, 0) + 1
        sc = round(sp / (sp + sf) * 100) if (sp + sf) else None
        recent.append({'community': comm, 'date': sa[:10], 'score': sc, 'submitted_at': sa})

    pass_rate = round(passes / (passes + fails) * 100) if (passes + fails) else 0
    scores = [r['score'] for r in recent if r['score'] is not None]
    avg_score = round(sum(scores) / len(scores)) if scores else None
    recent.sort(key=lambda x: x['submitted_at'], reverse=True)

    comm_list = []
    for cn, cc in by_comm.items():
        tot = cc['pass'] + cc['fail']
        comm_list.append({'name': cn, 'visits': cc['visits'],
                          'avg_score': round(cc['pass'] / tot * 100) if tot else None})
    comm_list.sort(key=lambda x: x['visits'], reverse=True)

    months = []
    for i in range(5, -1, -1):
        m = now.month - 1 - i
        y, mm = now.year + (m // 12), (m % 12) + 1
        months.append({'label': _dt(y, mm, 1).strftime('%b'),
                       'visits': by_month.get(f"{y:04d}-{mm:02d}", 0)})

    weak_list = sorted([{'question': q, 'fails': n} for q, n in weak.items()],
                       key=lambda x: x['fails'], reverse=True)[:5]

    # --- Global leaderboard ranks (across everyone, to fuel competition) ---
    agg = {}
    for s in inspection_service.get_all_submissions():
        key = s.get('inspector_name') or resolve_display_name(s.get('username', '')) or s.get('username', '?')
        a = agg.setdefault(key, {'visits': 0, 'pass': 0, 'fail': 0})
        a['visits'] += 1
        for r in (s.get('responses') or []):
            if r.get('condition') == 'Pass':
                a['pass'] += 1
            elif r.get('condition') == 'Fail':
                a['fail'] += 1

    def _rank_of(key, pairs):
        for i, (k, _v) in enumerate(pairs):
            if k == key:
                return i + 1
        return None

    visits_sorted = sorted(((k, a['visits']) for k, a in agg.items() if a['visits'] > 0),
                           key=lambda x: x[1], reverse=True)
    pass_sorted = sorted(((k, a['pass'] / (a['pass'] + a['fail'])) for k, a in agg.items()
                          if (a['pass'] + a['fail']) > 0), key=lambda x: x[1], reverse=True)
    rank = {
        'visits': {'pos': _rank_of(name, visits_sorted), 'total': len(visits_sorted)},
        'pass_rate': {'pos': _rank_of(name, pass_sorted), 'total': len(pass_sorted)},
    }

    meta = {'role': '', 'region': '', 'photo': None, 'email': ''}
    for r in region_service.get_all_regions():
        for l in (r.get('leadership') or []):
            if (l.get('name') or '') == name:
                meta = {'role': l.get('role', ''), 'region': r.get('name', ''),
                        'photo': profile_service.get_leader_photo(r.get('id', ''), name),
                        'email': l.get('email', '')}

    return jsonify({'status': 'success', 'profile': {
        'name': name, 'role': meta['role'], 'region': meta['region'],
        'photo': meta['photo'], 'email': meta['email'],
        'total_visits': len(subs), 'visits_this_month': visits_this_month,
        'last_visit': last_visit[:10], 'avg_score': avg_score, 'pass_rate': pass_rate,
        'passes': passes, 'fails': fails, 'communities': comm_list,
        'weakest': weak_list, 'by_month': months, 'recent': recent[:8],
        'rank': rank,
    }}), 200


@app.route('/api/leaderboard')
@login_required
def leaderboard():
    """Team leaderboard: top by number of visits and by pass rate.

    Cross-region by design, to encourage competition among the people who do
    the visits. Community-level accounts (Executive Directors) are deliberately
    excluded: they only ever see their own community, and a company-wide
    ranking of regionals is not theirs to browse."""
    if not (is_admin() or is_leadership()):
        return jsonify({'status': 'error', 'message': 'Not available for this account'}), 403
    agg = {}
    for s in inspection_service.get_all_submissions():
        key = s.get('inspector_name') or resolve_display_name(s.get('username', '')) or s.get('username', '?')
        a = agg.setdefault(key, {'visits': 0, 'pass': 0, 'fail': 0})
        a['visits'] += 1
        for r in (s.get('responses') or []):
            if r.get('condition') == 'Pass':
                a['pass'] += 1
            elif r.get('condition') == 'Fail':
                a['fail'] += 1

    meta = {}
    for r in region_service.get_all_regions():
        for l in (r.get('leadership') or []):
            nm = l.get('name')
            if nm:
                meta[nm] = {'role': l.get('role', ''), 'region': r.get('name', '')}

    rows = []
    for k, a in agg.items():
        tot = a['pass'] + a['fail']
        rows.append({'name': k, 'visits': a['visits'],
                     'pass_rate': round(a['pass'] / tot * 100) if tot else None,
                     'role': meta.get(k, {}).get('role', ''),
                     'region': meta.get(k, {}).get('region', '')})

    by_visits = sorted(rows, key=lambda x: x['visits'], reverse=True)[:10]
    by_pass = sorted([r for r in rows if r['pass_rate'] is not None],
                     key=lambda x: (x['pass_rate'], x['visits']), reverse=True)[:10]
    return jsonify({'status': 'success', 'by_visits': by_visits, 'by_pass_rate': by_pass}), 200


# ==================== MOVE-IN MODULE ====================

def _movein_progress(rec, item_ids):
    """Count completed items for a move-in record against the template items."""
    comps = rec.get('completions') or {}
    done = sum(1 for iid in item_ids if (comps.get(iid) or {}).get('done'))
    total = len(item_ids)
    return done, total


def _movein_attachment_url(entry):
    path = (entry or {}).get('attachment_path')
    if not path:
        return None
    if file_upload_handler.use_s3:
        return file_upload_handler.generate_presigned_url(path, download_name=entry.get('attachment_name'))
    return url_for('static', filename=f'uploads/{path}')


def _movein_blockers(rec):
    """Required ('gate') items not yet completed for this move-in: [(id, text)]."""
    comps = rec.get('completions') or {}
    return [(iid, text) for iid, text in movein_template_service.required_items()
            if not (comps.get(iid) or {}).get('done')]


def _movein_allowed_communities():
    """Communities the current user may see move-ins for, or None = all (admin)."""
    role = current_role()
    if is_admin():
        return None
    if is_leadership(role):
        return set(regional_communities())
    # community account — may cover more than one site
    return set(session_communities())


def _can_access_movein(rec):
    allowed = _movein_allowed_communities()
    return allowed is None or (rec and rec.get('community') in allowed)


def _scoped_moveins():
    """All move-in records visible to the current user (role-scoped)."""
    allowed = _movein_allowed_communities()
    everything = movein_service.get_all()
    if allowed is None:
        return everything
    return [m for m in everything if m.get('community') in allowed]


@app.route('/api/moveins', methods=['GET'])
@login_required
def list_moveins():
    """List move-ins with computed progress, scoped to the user's role
    (admin = all, regional = their region's communities, staff = their community)."""
    item_ids = movein_template_service.all_item_ids()
    out = []
    for rec in _scoped_moveins():
        done, total = _movein_progress(rec, item_ids)
        out.append({
            'id': rec.get('id'),
            'resident_name': rec.get('resident_name'),
            'community': rec.get('community'),
            'target_date': rec.get('target_date'),
            'status': rec.get('status', 'active'),
            'created_at': rec.get('created_at'),
            'done': done, 'total': total,
            'blockers': len(_movein_blockers(rec)),
        })
    return jsonify({'status': 'success', 'moveins': out}), 200


@app.route('/api/moveins', methods=['POST'])
@login_required
def create_movein():
    """Create a new resident move-in record."""
    data = request.get_json(silent=True) or {}
    resident = InputSanitizer.sanitize_string(data.get('resident_name', ''), max_length=120)
    community = InputSanitizer.sanitize_community_name(data.get('community', ''))
    target_date = InputSanitizer.sanitize_string(data.get('target_date', ''), max_length=20)
    if not resident:
        return jsonify({'status': 'error', 'message': 'Resident name is required'}), 400
    if not community:
        return jsonify({'status': 'error', 'message': 'Community is required'}), 400
    # Non-admins can only create move-ins for communities in their own scope.
    allowed = _movein_allowed_communities()
    if allowed is not None and community not in allowed:
        return jsonify({'status': 'error', 'message': 'You can only create move-ins for your own communities'}), 403
    rec = movein_service.create(resident, community, target_date, created_by=session.get('user'))
    activity_service.log(session.get('user'), 'movein_created', f'Started move-in for {resident} ({community})')
    return jsonify({'status': 'success', 'movein': rec}), 200


@app.route('/api/moveins/<mv_id>', methods=['GET'])
@login_required
def get_movein(mv_id):
    """Return a move-in record merged with the template (phases + items + completion)."""
    rec = movein_service.get(mv_id)
    if rec is None or not _can_access_movein(rec):
        return jsonify({'status': 'error', 'message': 'Move-in not found'}), 404
    template = movein_template_service.get_template()
    comps = rec.get('completions') or {}
    item_ids = []
    for ph in template['phases']:
        for it in ph.get('items', []):
            item_ids.append(it['id'])
            entry = comps.get(it['id']) or {}
            it['done'] = bool(entry.get('done'))
            it['date'] = entry.get('date', '')
            it['initials'] = entry.get('initials', '')
            it['note'] = entry.get('note', '')
            it['required'] = bool(it.get('required'))
            it['department'] = it.get('department', '')
            it['updated_by'] = entry.get('updated_by')
            it['updated_at'] = entry.get('updated_at')
            it['attachment_name'] = entry.get('attachment_name')
            it['attachment_url'] = _movein_attachment_url(entry)
    done, total = _movein_progress(rec, item_ids)
    blockers = [{'id': i, 'text': t} for i, t in _movein_blockers(rec)]
    return jsonify({'status': 'success', 'movein': rec, 'template': template,
                    'done': done, 'total': total, 'blockers': blockers}), 200


@app.route('/api/moveins/<mv_id>/item', methods=['POST'])
@login_required
def update_movein_item(mv_id):
    """Update one checklist item's completion (done / date / initials)."""
    data = request.get_json(silent=True) or {}
    item_id = InputSanitizer.sanitize_string(data.get('item_id', ''), max_length=60)
    if not item_id:
        return jsonify({'status': 'error', 'message': 'item_id is required'}), 400
    if not _can_access_movein(movein_service.get(mv_id)):
        return jsonify({'status': 'error', 'message': 'Move-in not found'}), 404
    rec = movein_service.update_item(
        mv_id, item_id,
        done=data.get('done'),
        date=data.get('date'),
        initials=data.get('initials'),
        note=data.get('note'),
        updated_by=session.get('user'))
    if rec is None:
        return jsonify({'status': 'error', 'message': 'Move-in not found'}), 404
    return jsonify({'status': 'success'}), 200


@app.route('/api/moveins/<mv_id>/item/attachment', methods=['POST'])
@login_required
def upload_movein_attachment(mv_id):
    """Attach a file (signed form, etc.) to a checklist item."""
    item_id = InputSanitizer.sanitize_string(request.form.get('item_id', ''), max_length=60)
    if not item_id:
        return jsonify({'status': 'error', 'message': 'item_id is required'}), 400
    if not _can_access_movein(movein_service.get(mv_id)):
        return jsonify({'status': 'error', 'message': 'Move-in not found'}), 404
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'status': 'error', 'message': 'No file provided'}), 400
    file = request.files['file']
    try:
        rel, name = file_upload_handler.save_movein_attachment(file, mv_id, item_id)
        movein_service.set_attachment(mv_id, item_id, rel, name)
        entry = (movein_service.get(mv_id).get('completions') or {}).get(item_id, {})
        return jsonify({'status': 'success', 'attachment_name': name,
                        'attachment_url': _movein_attachment_url(entry)}), 200
    except Exception as e:
        app.logger.exception('Error uploading move-in attachment')
        return jsonify({'status': 'error', 'message': 'Internal server error while uploading attachment'}), 500


@app.route('/api/moveins/<mv_id>', methods=['DELETE'])
@login_required
def delete_movein(mv_id):
    """Delete a move-in record (and best-effort remove its attachments)."""
    rec = movein_service.get(mv_id)
    if rec is None or not _can_access_movein(rec):
        return jsonify({'status': 'error', 'message': 'Move-in not found'}), 404
    for entry in (rec.get('completions') or {}).values():
        if entry.get('attachment_path'):
            file_upload_handler.delete_file(entry['attachment_path'])
    movein_service.delete(mv_id)
    activity_service.log(session.get('user'), 'movein_deleted', f"Deleted move-in for {rec.get('resident_name')}")
    return jsonify({'status': 'success'}), 200


@app.route('/api/moveins/<mv_id>/status', methods=['POST'])
@login_required
def set_movein_status(mv_id):
    """Mark a move-in active / completed / archived."""
    data = request.get_json(silent=True) or {}
    status = (data.get('status') or '').strip().lower()
    if status not in ('active', 'completed', 'archived'):
        return jsonify({'status': 'error', 'message': 'Invalid status'}), 400
    rec = movein_service.get(mv_id)
    if rec is None or not _can_access_movein(rec):
        return jsonify({'status': 'error', 'message': 'Move-in not found'}), 404
    # Compliance gate: can't mark complete while required items are unchecked.
    if status == 'completed':
        blockers = _movein_blockers(rec)
        if blockers:
            return jsonify({
                'status': 'error',
                'message': 'Cannot complete: required items are still pending.',
                'blockers': [{'id': i, 'text': t} for i, t in blockers],
            }), 409
    movein_service.set_status(mv_id, status)
    # Send a summary email when a move-in is completed (best-effort).
    if status == 'completed':
        try:
            community = rec.get('community', '')
            recipients = movein_recipients(community)
            if recipients:
                item_ids = movein_template_service.all_item_ids()
                done, total = _movein_progress(rec, item_ids)
                email_service.send_movein_completed(
                    recipients, rec.get('resident_name', ''), community,
                    rec.get('target_date', ''), done, total)
        except Exception as e:
            app.logger.error(f'Move-in completion email failed: {str(e)}')
    return jsonify({'status': 'success'}), 200


@app.route('/api/moveins/template', methods=['GET'])
@login_required
def get_movein_template():
    return jsonify({'status': 'success', 'template': movein_template_service.get_template()}), 200


@app.route('/api/moveins/template', methods=['POST'])
@require_admin
def save_movein_template():
    """Admin-only: replace the move-in checklist template (phases + items)."""
    data = request.get_json(silent=True) or {}
    phases = data.get('phases')
    if not isinstance(phases, list):
        return jsonify({'status': 'error', 'message': 'phases must be a list'}), 400
    tmpl = movein_template_service.save_template(phases)
    activity_service.log(session.get('user'), 'movein_template_saved', 'Updated move-in checklist template')
    return jsonify({'status': 'success', 'template': tmpl}), 200


def survey_type_coverage():
    """How many active standards each survey type would put on a form.

    A standard with no survey types is used by every one of them, which is the
    rule the form itself applies — so the count has to follow that, or a type
    would look empty while working perfectly.

    This exists because a survey type can be emptied without anyone noticing:
    unticking it on the last standard is an ordinary-looking edit, and the only
    symptom appears later, to a regional who has already driven to a community."""
    try:
        questions = question_manager.get_all_active_questions()
    except Exception:
        app.logger.exception('Could not read standards for survey type coverage')
        return {}

    counts = {}
    for st in survey_type_service.get_all_survey_types():
        tid = st.get('id')
        counts[tid] = sum(1 for q in questions
                          if not q.get('survey_types') or tid in q.get('survey_types'))
    return counts


def backup_status(stale_after_hours=36):
    """When the data was last backed up, and whether that is recent enough.

    The nightly backup failed silently for two weeks: cron could not execute
    the script, said so in a log file, and nobody reads log files. The backup
    now leaves a receipt on success; this reads it. Anything that guards the
    data has to report to somewhere a person actually looks.

    `stale_after_hours` is 36 rather than 24 so a single late or skipped night
    doesn't cry wolf — two missed nights does."""
    from datetime import timedelta as _timedelta
    path = os.path.join(DATA_FOLDER, '.last_backup')
    try:
        with open(path, encoding='utf-8') as f:
            stamp = f.read().strip().split(' ')[0]
        when = datetime.fromisoformat(stamp.replace('Z', ''))
        if when.tzinfo is not None:
            when = when.astimezone().replace(tzinfo=None)
        age = datetime.now() - when
        return {
            'known': True,
            'when': when.isoformat(),
            'age_hours': round(age.total_seconds() / 3600, 1),
            'stale': age > _timedelta(hours=stale_after_hours),
        }
    except FileNotFoundError:
        # No receipt at all: either it has never succeeded, or this is the
        # first run since the receipt was introduced.
        return {'known': False, 'when': None, 'age_hours': None, 'stale': True}
    except Exception as e:
        app.logger.warning(f'Could not read the backup receipt: {e}')
        return {'known': False, 'when': None, 'age_hours': None, 'stale': True}


def recent_errors(hours=24):
    """How many requests failed in the last `hours`, grouped by kind.

    Failures were invisible until a user wrote in — someone did, but the next
    person may just give up. Reading the service log costs nothing and turns a
    silent break into a line in the summary that already goes out."""
    import subprocess
    try:
        out = subprocess.run(
            ['journalctl', '-u', 'atlas', '--since', f'{hours} hours ago', '--no-pager'],
            capture_output=True, text=True, timeout=20).stdout
    except Exception:
        return []          # not on the server, or journald unavailable
    counts = {}
    for line in out.splitlines():
        # The catch-all handlers log "... ERROR ... <what failed>"; the traceback
        # lines that follow are indented, so only the headline is counted.
        if ' ERROR ' not in line:
            continue
        msg = line.split(' ERROR ', 1)[1].strip()
        if ':' in msg:
            msg = msg.split(':', 1)[1].strip() or msg
        msg = msg[:90]
        counts[msg] = counts.get(msg, 0) + 1
    return sorted(({'what': k, 'count': v} for k, v in counts.items()),
                  key=lambda x: -x['count'])[:6]


def build_activity_digest(hours=24):
    """Summarize what happened in the app over the last `hours`.

    Returns plain data (no email, no I/O beyond the log) so it can be rendered,
    tested or previewed without sending anything."""
    from datetime import timedelta as _timedelta
    since_dt = datetime.now() - _timedelta(hours=hours)
    since = since_dt.isoformat()
    events = [e for e in activity_service.get_recent(limit=5000)
              if e.get('timestamp', '') >= since]

    def of_type(*types):
        return [e for e in events if e.get('type') in types]

    # Who signed in, and how many times each
    signins = {}
    for e in of_type('login'):
        u = e.get('username')
        signins[u] = signins.get(u, 0) + 1
    signed_in = sorted(
        ({'username': u, 'name': resolve_display_name(u) or u, 'count': n}
         for u, n in signins.items()),
        key=lambda x: (-x['count'], x['name'].lower()))

    visits = [{'name': resolve_display_name(e.get('username')) or e.get('username'),
               'detail': e.get('detail', ''),
               'community': (e.get('meta') or {}).get('community', '')}
              for e in of_type('inspection_submitted')]

    addressed = [{'name': resolve_display_name(e.get('username')) or e.get('username'),
                  'detail': e.get('detail', '')}
                 for e in of_type('standard_addressed')]

    # The report-and-verify loop is the point of the whole thing, so the day's
    # comments belong here. What matters isn't the wording — it's whether the
    # communities that spoke up got an answer.
    comments = [{'name': resolve_display_name(e.get('username')) or e.get('username'),
                 'detail': e.get('detail', ''),
                 'community': (e.get('meta') or {}).get('community', '')}
                for e in of_type('standard_commented')]
    # Communities that reported something today with nothing closed out for
    # them — that's the queue somebody needs to work through.
    closed_for = {c for e in of_type('standard_addressed')
                  for c in [(e.get('meta') or {}).get('community', '')] if c}
    awaiting = sorted({c['community'] for c in comments
                       if c['community'] and c['community'] not in closed_for})

    security = [{'name': resolve_display_name(e.get('username')) or e.get('username'),
                 'type': e.get('type'), 'detail': e.get('detail', '')}
                for e in of_type('password_changed', 'password_reset',
                                 'password_reset_requested')]

    accounts = [{'name': resolve_display_name(e.get('username')) or e.get('username'),
                 'type': e.get('type'), 'detail': e.get('detail', '')}
                for e in of_type('user_created', 'person_created', 'person_updated',
                                 'person_removed', 'admin_privileges_granted',
                                 'admin_privileges_revoked')]

    # People with an account who have never signed in at all.
    never = []
    try:
        presence = presence_service.all()
        for reg in region_service.get_all_regions():
            for leader in (reg.get('leadership') or []):
                u = (leader.get('username') or '').strip()
                nm = (leader.get('name') or '').strip()
                if u and nm.lower() != 'open' and not presence.get(u, {}).get('last_login'):
                    never.append(nm or u)
        for u in user_service.get_all():
            uname = u.get('username')
            if uname and not presence.get(uname, {}).get('last_login'):
                never.append(u.get('display_name') or uname)
    except Exception as e:
        app.logger.error(f'Digest never-signed-in step failed: {e}')

    # Move-ins whose date has passed with required items still open. Not an
    # event from the last 24h — a standing condition that otherwise nobody is
    # told about, since the reminder script only looks forward.
    overdue = []
    try:
        from datetime import date as _date
        today = _date.today()
        for rec in movein_service.get_all():
            if rec.get('status') != 'active':
                continue
            td = (rec.get('target_date') or '').strip()
            if not td:
                continue
            try:
                target = datetime.strptime(td, '%Y-%m-%d').date()
            except ValueError:
                continue
            late = (today - target).days
            if late <= 0:
                continue
            missing = _movein_blockers(rec)
            if missing:
                overdue.append({'resident': rec.get('resident_name', ''),
                                'community': rec.get('community', ''),
                                'days': late, 'pending': len(missing)})
        overdue.sort(key=lambda x: -x['days'])
    except Exception as e:
        app.logger.error(f'Digest overdue move-ins step failed: {e}')

    return {
        'since': fmt_local(since_dt),
        'errors': recent_errors(hours),
        'backup': backup_status(),
        'overdue_moveins': overdue,
        'hours': hours,
        'signed_in': signed_in,
        'visits': visits,
        'addressed': addressed,
        'comments': comments,
        'awaiting_review': awaiting,
        'security': security,
        'accounts': accounts,
        'never_signed_in': sorted(set(never)),
        'total_events': len(events),
    }


def run_activity_digest(hours=24):
    """Build and email the activity digest to the admin-notify list.
    Returns (sent, detail, digest) so the cron script can report what happened."""
    digest = build_activity_digest(hours=hours)
    admin_notify = settings_service.get_email_settings().get('admin_notify', [])
    if not admin_notify:
        return (False, 'no admin recipients configured', digest)
    if not email_service.enabled:
        return (False, 'email disabled', digest)
    try:
        ok, detail = email_service.send_activity_digest(admin_notify, digest)
        return (bool(ok), detail, digest)
    except Exception as e:
        app.logger.error(f'Activity digest failed: {e}')
        return (False, str(e), digest)


def run_movein_reminders(days_ahead=3, other_cap=6):
    """Email a reminder for every ACTIVE move-in whose target date is within
    `days_ahead` days and still has open checklist items. Recipients come from
    movein_recipients(): the community itself plus the administrator list, not
    the region's leadership. Returns a summary list (also used by the cron
    script). Safe to call when email is disabled."""
    from datetime import date as _date
    today = _date.today()
    template = movein_template_service.get_template()
    all_items = [(it['id'], it.get('text', ''), bool(it.get('required')))
                 for ph in template['phases'] for it in ph.get('items', [])]
    sent = []
    for rec in movein_service.get_all():
        if rec.get('status') != 'active':
            continue
        td = (rec.get('target_date') or '').strip()
        try:
            target = datetime.strptime(td, '%Y-%m-%d').date()
        except ValueError:
            continue
        days_left = (target - today).days
        if days_left < 0 or days_left > days_ahead:
            continue
        comps = rec.get('completions') or {}
        missing_req = [t for (iid, t, req) in all_items if req and not (comps.get(iid) or {}).get('done')]
        missing_other = [t for (iid, t, req) in all_items if not req and not (comps.get(iid) or {}).get('done')]
        if not missing_req and not missing_other:
            continue  # everything done — no need to nag
        community = rec.get('community', '')
        recipients = movein_recipients(community)
        if not recipients:
            continue
        shown_other = missing_other[:other_cap]
        if len(missing_other) > other_cap:
            shown_other = shown_other + [f"...and {len(missing_other) - other_cap} more"]
        ok, detail = email_service.send_movein_reminder(
            recipients, rec.get('resident_name', ''), community, td, days_left,
            missing_req, shown_other)
        sent.append({'resident': rec.get('resident_name'), 'community': community,
                     'days_left': days_left, 'recipients': recipients, 'sent': ok, 'detail': detail})
    return sent


_MOVEIN_EXPORT_HEADERS = ['Resident', 'Community', 'Target date', 'Status',
                          'Completed', 'Total', 'Percent', 'Required pending', 'Created']


def _movein_export_rows():
    item_ids = movein_template_service.all_item_ids()
    rows = []
    for rec in _scoped_moveins():
        done, total = _movein_progress(rec, item_ids)
        pct = round(done / total * 100) if total else 0
        rows.append([
            rec.get('resident_name', ''), rec.get('community', ''),
            rec.get('target_date', ''), rec.get('status', 'active'),
            done, total, f"{pct}%", len(_movein_blockers(rec)),
            (rec.get('created_at', '') or '')[:10],
        ])
    return rows


@app.route('/api/moveins/export.csv')
@login_required
def export_moveins_csv():
    import csv
    import io
    from flask import Response
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(_MOVEIN_EXPORT_HEADERS)
    w.writerows(_movein_export_rows())
    return Response(buf.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename="atlas-moveins.csv"'})


@app.route('/api/moveins/export.xlsx')
@login_required
def export_moveins_xlsx():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = 'Move-Ins'
    hf = PatternFill('solid', fgColor='00285C')
    hfont = Font(bold=True, color='FFFFFF')
    for col, name in enumerate(_MOVEIN_EXPORT_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = hf
        c.font = hfont
    for r in _movein_export_rows():
        ws.append(r)
    for i, wdt in enumerate([26, 24, 14, 12, 11, 8, 9, 16, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = wdt
    ws.freeze_panes = 'A2'
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name='atlas-moveins.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _build_movein_pdf(resident, community, target_date, phases, filled):
    """Render a move-in checklist PDF. `phases` = [{name, items:[{text, required,
    done, date, initials}]}]. filled=True shows the resident's progress; False
    produces a blank form for the binder."""
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image)

    styles = getSampleStyleSheet()
    cell = ParagraphStyle('mi_cell', parent=styles['Normal'], fontSize=9, leading=12)
    req = ParagraphStyle('mi_req', parent=cell, textColor=colors.HexColor('#b42318'),
                         fontName='Helvetica-Bold', fontSize=7)
    head = ParagraphStyle('mi_head', parent=styles['Normal'], fontSize=9, leading=11,
                          textColor=colors.white, fontName='Helvetica-Bold')
    phead = ParagraphStyle('mi_phead', parent=styles['Normal'], fontSize=12,
                           textColor=colors.white, fontName='Helvetica-Bold')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=0.55 * inch,
                            rightMargin=0.55 * inch, topMargin=0.55 * inch, bottomMargin=0.55 * inch)
    story = []
    logo = os.path.join(os.path.dirname(__file__), 'static', 'atlas-logo.png')
    if os.path.exists(logo):
        try:
            img = Image(logo)
            ratio = img.imageWidth / float(img.imageHeight)
            img.drawHeight = 0.5 * inch
            img.drawWidth = 0.5 * inch * ratio
            img.hAlign = 'CENTER'
            story += [img, Spacer(1, 6)]
        except Exception:
            pass
    story += [
        Paragraph('Move-In Checklist', styles['Title']),
        Paragraph('Atlas Senior Living &mdash; New Resident Move-In', styles['Normal']),
        Spacer(1, 8),
    ]
    # Resident info line (filled) or blank lines (template)
    if filled:
        info = (f"<b>Resident:</b> {resident or '—'} &nbsp;&nbsp; "
                f"<b>Community:</b> {community or '—'} &nbsp;&nbsp; "
                f"<b>Target move-in date:</b> {target_date or '—'}")
    else:
        info = ("<b>Resident:</b> ______________________   "
                "<b>Community:</b> ______________________   "
                "<b>Move-in date:</b> ____________")
    story += [Paragraph(info, cell), Spacer(1, 12)]

    col_widths = [0.4 * inch, 4.5 * inch, 1.1 * inch, 0.9 * inch]
    for ph in phases:
        # phase header bar
        ph_tbl = Table([[Paragraph(ph['name'], phead)]], colWidths=[sum(col_widths)])
        ph_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#00285c')),
            ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(ph_tbl)

        rows = [[Paragraph('<b>&#10003;</b>', head), Paragraph('Item', head),
                 Paragraph('Date', head), Paragraph('Initials', head)]]
        for it in ph.get('items', []):
            mark = 'X' if (filled and it.get('done')) else ''
            txt = it.get('text', '')
            if it.get('required'):
                txt += " &nbsp;<font color='#b42318'><b>[REQUIRED]</b></font>"
            date_v = it.get('date', '') if filled else ''
            init_v = it.get('initials', '') if filled else ''
            rows.append([Paragraph(f"<b>{mark}</b>", cell), Paragraph(txt, cell),
                         Paragraph(date_v or '', cell), Paragraph(init_v or '', cell)])
        tbl = Table(rows, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f6fe5')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#c7d0db')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f6f8fb')]),
            ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story += [tbl, Spacer(1, 14)]

    # Sign-off block for the physical binder: staff + resident/family signatures.
    from reportlab.platypus import KeepTogether
    sig_label = ParagraphStyle('mi_sig', parent=cell, fontSize=9,
                               textColor=colors.HexColor('#475569'))
    sig_head = ParagraphStyle('mi_sighead', parent=cell, fontSize=10,
                              fontName='Helvetica-Bold', textColor=colors.HexColor('#00285c'))
    line = "________________________________"
    sig_rows = [
        [Paragraph(line, cell), Paragraph(line, cell)],
        [Paragraph('Staff signature', sig_label), Paragraph('Resident / Family signature', sig_label)],
        [Paragraph('&nbsp;', cell), Paragraph('&nbsp;', cell)],
        [Paragraph('Printed name: ______________________', sig_label),
         Paragraph('Printed name: ______________________', sig_label)],
        [Paragraph('Date: ____________', sig_label),
         Paragraph('Date: ____________', sig_label)],
    ]
    sig_tbl = Table(sig_rows, colWidths=[3.2 * inch, 3.2 * inch])
    sig_tbl.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
    ]))
    story += [KeepTogether([
        Spacer(1, 10),
        Paragraph('Sign-off', sig_head),
        Spacer(1, 8),
        sig_tbl,
    ])]

    doc.build(story)
    buf.seek(0)
    return buf


@app.route('/api/moveins/template/pdf')
@login_required
def movein_template_pdf():
    """Blank printable checklist (for the binder)."""
    template = movein_template_service.get_template()
    phases = [{'name': ph['name'],
               'items': [{'text': it.get('text', ''), 'required': bool(it.get('required'))}
                         for it in ph.get('items', [])]}
              for ph in template['phases']]
    buf = _build_movein_pdf('', '', '', phases, filled=False)
    return send_file(buf, as_attachment=True, download_name='move-in-checklist-blank.pdf',
                     mimetype='application/pdf')


@app.route('/api/moveins/<mv_id>/pdf')
@login_required
def movein_pdf(mv_id):
    """Printable checklist for one resident, showing current progress."""
    rec = movein_service.get(mv_id)
    if rec is None or not _can_access_movein(rec):
        return jsonify({'status': 'error', 'message': 'Move-in not found'}), 404
    template = movein_template_service.get_template()
    comps = rec.get('completions') or {}
    phases = []
    for ph in template['phases']:
        items = []
        for it in ph.get('items', []):
            entry = comps.get(it['id']) or {}
            items.append({'text': it.get('text', ''), 'required': bool(it.get('required')),
                          'done': bool(entry.get('done')), 'date': entry.get('date', ''),
                          'initials': entry.get('initials', '')})
        phases.append({'name': ph['name'], 'items': items})
    buf = _build_movein_pdf(rec.get('resident_name', ''), rec.get('community', ''),
                            rec.get('target_date', ''), phases, filled=True)
    safe = ''.join(c if c.isalnum() else '-' for c in (rec.get('resident_name') or 'resident')).strip('-').lower()
    return send_file(buf, as_attachment=True, download_name=f'move-in-{safe or "resident"}.pdf',
                     mimetype='application/pdf')


@app.route('/api/moveins/run-reminders', methods=['POST'])
@require_admin
def trigger_movein_reminders():
    """Admin-only: run the move-in reminder sweep on demand (also used to test)."""
    data = request.get_json(silent=True) or {}
    try:
        days = int(data.get('days_ahead', 3))
    except (TypeError, ValueError):
        days = 3
    result = run_movein_reminders(days_ahead=days)
    return jsonify({'status': 'success', 'emails': result, 'count': len(result)}), 200


@app.route('/api/activity/digest', methods=['POST'])
@require_admin
def trigger_activity_digest():
    """Admin-only: send the daily activity digest now, or preview it without
    sending (pass {"preview": true}) to see exactly what would go out."""
    data = request.get_json(silent=True) or {}
    try:
        hours = max(1, min(int(data.get('hours', 24)), 168))
    except (TypeError, ValueError):
        hours = 24
    if data.get('preview'):
        return jsonify({'status': 'success', 'preview': True,
                        'digest': build_activity_digest(hours=hours)}), 200
    sent, detail, digest = run_activity_digest(hours=hours)
    return jsonify({'status': 'success' if sent else 'error',
                    'sent': sent, 'detail': str(detail),
                    'digest': digest}), (200 if sent else 400)


@app.route('/api/users', methods=['GET'])
@login_required
def list_users():
    """Admin-only: list admin-created login accounts."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    return jsonify({'status': 'success', 'users': user_service.get_all()}), 200


@app.route('/api/users', methods=['POST'])
@login_required
def create_user():
    """
    Admin-only: create a new login account.
    Body: { display_name, role: admin|staff|regional, community?, region_id? }
    The username is generated from the name; a strong password is generated and
    returned ONCE (it is stored only as a hash).
    """
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403

    data = request.get_json(silent=True) or {}
    display_name = (data.get('display_name') or '').strip()
    role = (data.get('role') or '').strip().lower()
    community = (data.get('community') or '').strip() or None
    region_id = (data.get('region_id') or '').strip() or None
    email = (data.get('email') or '').strip() or None

    if not display_name:
        return jsonify({'status': 'error', 'message': 'Name is required'}), 400
    if role not in ('admin', 'staff', 'regional'):
        return jsonify({'status': 'error', 'message': 'Invalid role'}), 400

    if role == 'staff':
        valid = set()
        for r in region_service.get_all_regions():
            valid.update(r.get('communities', []))
        valid.update(ALL_COMMUNITIES)
        if not community or community not in valid:
            return jsonify({'status': 'error', 'message': 'A valid community is required for staff'}), 400
        region_id = None
    elif is_leadership(role):
        valid_ids = {r.get('id') for r in region_service.get_all_regions() if r.get('id') != 'unassigned'}
        if not region_id or region_id not in valid_ids:
            return jsonify({'status': 'error', 'message': 'A valid region is required for a regional account'}), 400
        community = None
    else:  # admin
        community = None
        region_id = None

    username = generate_unique_username(display_name)
    password = generate_password()
    user_service.create(
        username=username,
        display_name=display_name,
        role=role,
        password_hash=generate_password_hash(password),
        community=community,
        region_id=region_id,
        created_by=session.get('user'),
        email=email,
    )
    # Same rule as People > Add person: the temporary password must be replaced
    # by one only they know, at first sign-in.
    profile_service.set_must_change(username, True)
    try:
        activity_service.log(session.get('user'), 'user_created',
                             f'Created {role} account for {display_name}',
                             meta={'username': username})
    except Exception:
        pass

    # Emails (best-effort; never block account creation):
    #  - welcome the new user with their login (if an email was given)
    #  - alert the configured admin-notify list
    role_label = role_label_for(role, community)
    emailed = False
    if email_service.enabled:
        try:
            if email:
                ok, _ = email_service.send_welcome(email, display_name, username, password, role_label)
                emailed = bool(ok)
            admin_notify = settings_service.get_email_settings().get('admin_notify', [])
            if admin_notify:
                email_service.send_new_user_alert(admin_notify, display_name, username,
                                                  role_label, session.get('user'))
        except Exception as e:
            app.logger.error(f'User-creation email step failed: {e}')

    return jsonify({
        'status': 'success',
        'message': 'User created',
        'username': username,
        'password': password,
        'display_name': display_name,
        'role': role,
        'emailed': emailed,
    }), 201


@app.route('/api/users/<username>/reset-password', methods=['POST'])
@login_required
def reset_user_password(username):
    """Admin-only: generate a new password for a created user (returned once)."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    if not user_service.exists(username):
        return jsonify({'status': 'error', 'message': 'User not found'}), 404
    password = generate_password()
    user_service.set_password_hash(username, generate_password_hash(password))
    # Clear any per-user profile override so the new password takes effect,
    # and force a password change on next login.
    try:
        profile_service.set_password_hash(username, '')
        profile_service.set_must_change(username, True)
    except Exception:
        pass
    return jsonify({'status': 'success', 'username': username, 'password': password}), 200


@app.route('/api/admin/reset-password', methods=['POST'])
@login_required
def admin_reset_password():
    """Admin-only: reset ANY account (admin/staff, created user, or regional)
    to a temporary password and force a change on next login. The temp password
    is returned once so the admin can share it with the user."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    data = request.get_json(silent=True) or {}
    username = InputSanitizer.sanitize_username(data.get('username', ''))
    if not username:
        return jsonify({'status': 'error', 'message': 'Username is required'}), 400
    acct = resolve_account_context(username)
    if not acct.get('exists'):
        return jsonify({'status': 'error', 'message': 'No account with that username'}), 404
    password = generate_password()
    # A profile override takes precedence over every account source, so this
    # works uniformly for admin/staff, created users and regional leaders.
    profile_service.set_password_hash(username, generate_password_hash(password))
    profile_service.set_must_change(username, True)
    # Security notice to the administrators: someone else's password was reset.
    alert_password_changed(username, changed_by=session.get('user') or 'an admin')

    # Email them the temporary password when we have an address on file.
    # Whether it actually went out matters as much as the reset itself, so we
    # keep the reason and record it in the audit log instead of dropping it.
    emailed = False
    reason = ''
    if not acct.get('email'):
        reason = 'no email address on file'
    else:
        try:
            ok, detail = email_service.send_password_reset(
                acct['email'], acct.get('display_name') or username, username, password)
            emailed = bool(ok)
            if not emailed:
                reason = str(detail)
                app.logger.error('Password-reset email to %s not sent: %s',
                                 acct['email'], detail)
        except Exception as e:
            reason = str(e)
            app.logger.error(f'Password-reset email failed: {e}')

    who = acct.get('display_name') or username
    activity_service.log(
        session.get('user'), 'password_reset',
        f"Reset password for {who} — "
        + (f"emailed to {acct['email']}" if emailed else f"NOT emailed ({reason})"))

    return jsonify({'status': 'success', 'username': username,
                    'display_name': who,
                    'password': password, 'emailed': emailed,
                    'reason': reason,
                    'email': acct.get('email') or ''}), 200


def _can_see_community(community):
    """Everyone sees their own scope: admins everywhere, regional/corporate
    across their communities, a community account only its own."""
    if is_admin():
        return True
    if is_leadership():
        return community in set(regional_communities())
    return community in set(session_communities())


def visible_communities():
    """Every community the current session may see, in one call.

    The same three-way answer as _can_see_community, but as a list — anything
    that has to iterate rather than test a single name should use this so the
    scoping rule stays in one place."""
    if is_admin():
        return all_communities()
    if is_leadership():
        return regional_communities()
    return session_communities()


def _community_account_usernames():
    """Usernames belonging to communities themselves (Executive Directors).

    Used to tell apart the two sides of a comment thread: a community saying
    "this is fixed" is a request, leadership replying is an answer."""
    return {u.get('username') for u in user_service.get_all()
            if u.get('role') == 'staff' and u.get('username')}


def _days_since(iso_ts):
    """Whole days between an ISO timestamp and now, or None if unparseable."""
    if not iso_ts:
        return None
    try:
        ts = datetime.fromisoformat(str(iso_ts).replace('Z', ''))
    except (TypeError, ValueError):
        return None
    return max(0, (datetime.now() - ts).days)


def build_attention(limit_per_group=6):
    """What needs the signed-in person, today.

    The dashboard has always reported: averages, totals, charts. None of it
    says which three things are waiting on *you*, so the report-and-verify loop
    depended on somebody remembering to go looking. Everything here is derived
    from data the app already holds — this only surfaces it.

    Scope is whatever the session may see, so a regional gets their region and
    an Executive Director gets their own community and nothing else."""
    role = current_role()
    scope = set(visible_communities())
    cadence = settings_service.get_visit_cadence_days()
    community_accounts = _community_account_usernames()
    groups = []

    # Only visits inside the caller's scope, newest first.
    subs = [s for s in inspection_service.get_all_submissions()
            if s.get('community') in scope]
    subs.sort(key=lambda s: s.get('submitted_at', ''), reverse=True)

    def failed_open(sub):
        """Failed standards on a visit that nobody has closed out yet.

        A standard is yielded once even if the visit holds more than one row
        for it. Older submissions do contain duplicates, and since a comment
        or a close only ever lands on the first row, the extras would show up
        as phantom work that can never be cleared."""
        seen_q = set()
        for r in (sub.get('responses') or []):
            if r.get('condition') != 'Fail' or r.get('addressed'):
                continue
            key = r.get('question_id') or ('t:' + (r.get('question_text') or '').strip().lower())
            if key in seen_q:
                continue
            seen_q.add(key)
            yield r

    if can_verify_fixes():
        # 1. A community reported a fix and is waiting for someone to confirm.
        #    This is the one thing that stalls silently: the ED has done the
        #    work and the score stays down until a regional looks.
        waiting = []
        for s in subs:
            for r in failed_open(s):
                comments = r.get('comments') or []
                if not comments:
                    continue
                last = comments[-1]
                if last.get('username') not in community_accounts:
                    continue  # leadership spoke last — the ball isn't here
                waiting.append({
                    'community': s.get('community', ''),
                    'title': r.get('question_text', '') or 'Standard',
                    'detail': f"Reported fixed by {last.get('author') or 'the community'}",
                    'days': _days_since(last.get('at')),
                    'submission_id': s.get('id'),
                    'question_id': r.get('question_id'),
                    'view': 'action-items',
                })
        waiting.sort(key=lambda x: -(x['days'] or 0))
        if waiting:
            groups.append({
                'key': 'verify', 'tone': 'act',
                'title': 'Fixes waiting on your confirmation',
                'note': 'A community says these are done. The score only moves once you agree.',
                'total': len(waiting), 'items': waiting[:limit_per_group],
            })

        # 2. Communities falling behind on visits. Never-visited ones sort
        #    first: they are the easiest to overlook precisely because they
        #    have nothing on the dashboard.
        seen = {}
        for s in subs:
            c = s.get('community')
            if c and c not in seen:
                seen[c] = s.get('submitted_at')
        behind = []
        for c in sorted(scope):
            d = _days_since(seen.get(c)) if seen.get(c) else None
            if c not in seen:
                behind.append({'community': c, 'title': c, 'detail': 'No visit on record',
                               'days': None, 'never': True, 'view': 'communities'})
            elif d is not None and d >= cadence:
                behind.append({'community': c, 'title': c,
                               'detail': f'Last visit {d} days ago',
                               'days': d, 'never': False, 'view': 'communities'})
        behind.sort(key=lambda x: (not x['never'], -(x['days'] or 0)))
        if behind:
            groups.append({
                'key': 'overdue', 'tone': 'plan',
                'title': 'Due for a visit',
                'note': f'Target is every {cadence} days.',
                'total': len(behind), 'items': behind[:limit_per_group],
            })

        # 3. Open findings nobody has said anything about. Not overdue in any
        #    formal sense — just quietly ageing, which is how a backlog forms.
        quiet = []
        for s in subs:
            age = _days_since(s.get('submitted_at'))
            if age is None or age < cadence:
                continue
            for r in failed_open(s):
                if r.get('comments'):
                    continue
                quiet.append({
                    'community': s.get('community', ''),
                    'title': r.get('question_text', '') or 'Standard',
                    'detail': f'Open {age} days, no follow-up',
                    'days': age,
                    'submission_id': s.get('id'),
                    'question_id': r.get('question_id'),
                    'view': 'action-items',
                })
        quiet.sort(key=lambda x: -(x['days'] or 0))
        if quiet:
            groups.append({
                'key': 'quiet', 'tone': 'watch',
                'title': 'Open with no follow-up',
                'note': 'Nothing has been said about these since the visit.',
                'total': len(quiet), 'items': quiet[:limit_per_group],
            })
    else:
        # An Executive Director's side of the same loop.
        mine = set(session_communities())
        me = session.get('user')
        to_answer, awaiting = [], []
        for s in subs:
            if s.get('community') not in mine:
                continue
            age = _days_since(s.get('submitted_at'))
            for r in failed_open(s):
                comments = r.get('comments') or []
                spoke = any(c.get('username') in community_accounts for c in comments)
                row = {
                    'community': s.get('community', ''),
                    'title': r.get('question_text', '') or 'Standard',
                    'days': age,
                    'submission_id': s.get('id'),
                    'question_id': r.get('question_id'),
                    'view': 'action-items',
                }
                if not spoke:
                    row['detail'] = (f'Open {age} days — tell your regional what you have done'
                                     if age else 'Tell your regional what you have done')
                    to_answer.append(row)
                else:
                    last = comments[-1]
                    row['detail'] = 'Waiting on your regional to confirm'
                    row['days'] = _days_since(last.get('at'))
                    awaiting.append(row)
        to_answer.sort(key=lambda x: -(x['days'] or 0))
        awaiting.sort(key=lambda x: -(x['days'] or 0))
        if to_answer:
            groups.append({
                'key': 'respond', 'tone': 'act',
                'title': 'Needs an update from you',
                'note': 'Add a comment and a photo once the work is done.',
                'total': len(to_answer), 'items': to_answer[:limit_per_group],
            })
        if awaiting:
            groups.append({
                'key': 'awaiting', 'tone': 'wait',
                'title': 'With your regional',
                'note': 'You have reported these. Only a regional can close them.',
                'total': len(awaiting), 'items': awaiting[:limit_per_group],
            })
        _ = me  # scope is by community, not by who happens to be signed in

    return {
        'role': role,
        'cadence_days': cadence,
        'groups': groups,
        'total': sum(g['total'] for g in groups),
    }


@app.route('/api/attention')
@login_required
def attention():
    """The short list of things waiting on whoever is signed in."""
    try:
        return jsonify({'status': 'success', **build_attention()}), 200
    except Exception:
        app.logger.exception('Building the attention list failed')
        # Never let this take the dashboard down with it — an empty list just
        # means the strip doesn't render.
        return jsonify({'status': 'success', 'role': current_role(),
                        'groups': [], 'total': 0, 'cadence_days': 0}), 200


# ===================== Items a community raises for itself =====================

@app.route('/api/raised-items', methods=['GET'])
@login_required
def list_raised_items():
    """Everything raised by the communities this account covers."""
    include = request.args.get('resolved') in ('1', 'true', 'yes')
    items = raised_item_service.for_communities(visible_communities(),
                                               include_resolved=include)
    return jsonify({'status': 'success', 'items': items}), 200


@app.route('/api/raised-items', methods=['POST'])
@login_required
def create_raised_item():
    """Raise something for your own community.

    Open to anyone who can see the community — an Executive Director raising
    what they need, a regional noting something between visits. It is not part
    of any visit and never touches a score."""
    community = InputSanitizer.sanitize_community_name(
        (request.form.get('community') if request.files
         else (request.get_json(silent=True) or {}).get('community', '')) or '')
    if not community:
        # Someone covering a single community shouldn't have to name it.
        mine = session_communities()
        community = mine[0] if len(mine) == 1 else ''
    if not community or not _can_see_community(community):
        return jsonify({'status': 'error', 'message': 'Pick a community you cover'}), 400

    if request.files:
        text = InputSanitizer.sanitize_description(request.form.get('text', ''))
        priority = InputSanitizer.sanitize_string(request.form.get('priority', 'medium'),
                                                  max_length=10)
    else:
        data = request.get_json(silent=True) or {}
        text = InputSanitizer.sanitize_description(data.get('text', ''))
        priority = InputSanitizer.sanitize_string(data.get('priority', 'medium'),
                                                  max_length=10)
    if not text.strip():
        return jsonify({'status': 'error', 'message': 'Say what needs attention'}), 400

    photo_path = ''
    f = request.files.get('photo')
    if f and f.filename:
        ok_file, why = file_upload_handler.validate_file(f)
        if not ok_file:
            return jsonify({'status': 'error', 'message': why}), 400
        try:
            photo_path = file_upload_handler.save_file(f, session.get('user', 'user'), community)
        except Exception:
            # A photo is never worth losing the item over.
            app.logger.exception('Could not save a raised item photo')

    username = session.get('user')
    item = raised_item_service.create(
        community, text, username,
        # Resolved once, at write time, so the item still reads correctly years
        # later even if the person leaves. resolve_display_name() falls back to
        # the username itself, so the session's own name is tried first — it is
        # the one the person actually signed in under.
        _display_name_for(username),
        priority=priority, photo=photo_path)
    if not item:
        return jsonify({'status': 'error', 'message': 'Could not raise this'}), 400

    activity_service.log(username, 'item_raised',
                         f'Raised an item at {community}: {text[:60]}',
                         meta={'community': community})
    notify_raised_item(item)
    return jsonify({'status': 'success', 'item': item}), 201


@app.route('/api/raised-items/<item_id>/resolve', methods=['POST'])
@login_required
def resolve_raised_item(item_id):
    """Close one out — or reopen it.

    Unlike a failed standard, a community may close its own. Closing a finding
    moves the score, which is why that stays with a regional; this doesn't
    touch any score, and the person who asked for the furniture is the one who
    knows it arrived."""
    item = raised_item_service.get(item_id)
    if not item:
        return jsonify({'status': 'error', 'message': 'Not found'}), 404
    if not _can_see_community(item.get('community', '')):
        return jsonify({'status': 'error', 'message': 'Not allowed for this community'}), 403

    data = request.get_json(silent=True) or {}
    resolved = data.get('resolved', True) is not False
    note = InputSanitizer.sanitize_description(data.get('note', ''))
    updated = raised_item_service.resolve(item_id, session.get('user'), note, resolved)
    activity_service.log(session.get('user'),
                         'item_resolved' if resolved else 'item_reopened',
                         f"{'Closed' if resolved else 'Reopened'} an item at "
                         f"{item.get('community', '')}: {item.get('text', '')[:60]}",
                         meta={'community': item.get('community', '')})
    return jsonify({'status': 'success', 'item': updated}), 200


def notify_raised_item(item):
    """Tell the region's leadership that a community has raised something.

    Best-effort, like every other notification here: a failed email must never
    lose the item that was just raised."""
    if not email_service.enabled:
        return
    try:
        community = item.get('community', '')
        recipients = region_leader_emails(community)
        # Whoever raised it doesn't need to be told about their own item.
        recipients = [a for a in recipients if a]
        if not recipients:
            return
        email_service.send_raised_item(recipients, item)
    except Exception:
        app.logger.exception('Could not send the raised-item notification')


@app.route('/api/communities/<path:community>/history')
@login_required
def community_history(community):
    """Everything this community has been through, on demand.

    Kept out of the bulk /api/inspections payload on purpose: that one is
    fetched on every dashboard load, and history only grows. Asking for it per
    community keeps the app fast however many years of visits pile up."""
    community = InputSanitizer.sanitize_community_name(community or '')
    if not community:
        return jsonify({'status': 'error', 'message': 'Community is required'}), 400
    if not _can_see_community(community):
        return jsonify({'status': 'error', 'message': 'Not allowed for this community'}), 403

    try:
        track_limit = max(1, min(int(request.args.get('track', 6)), 24))
    except (TypeError, ValueError):
        track_limit = 6

    subs = sorted(inspection_service.get_submissions_by_community(community),
                  key=lambda s: s.get('submitted_at', ''), reverse=True)

    visits = []
    for s in subs:
        responses = s.get('responses') or []
        passed = sum(1 for r in responses if r.get('condition') == 'Pass')
        failed = sum(1 for r in responses if r.get('condition') == 'Fail')
        fixed = sum(1 for r in responses
                    if r.get('condition') == 'Fail' and r.get('addressed'))
        total = passed + failed
        manual = s.get('action_items') or []
        # A visit that only covered part of the survey. Older visits have no
        # total recorded, so they are reported as unknown rather than complete
        # — claiming they were whole would be inventing a fact.
        declared_total = s.get('standards_total')
        answered = len(responses)
        partial = bool(declared_total) and answered < declared_total
        visits.append({
            'id': s.get('id'),
            'submitted_at': s.get('submitted_at'),
            'answered': answered,
            'standards_total': declared_total,
            'partial': partial,
            'inspector': s.get('inspector_name') or resolve_display_name(s.get('username', '')),
            'survey_type': survey_type_service.get_survey_type_name(s.get('survey_type_id')) or '',
            # Two numbers, same meaning as everywhere else: what was found that
            # day, and where it stands now that fixes have been verified.
            'visit_score': round(passed / total * 100) if total else None,
            'current_score': round((passed + fixed) / total * 100) if total else None,
            'passed': passed, 'failed': failed, 'fixed': fixed,
            'action_items': len(manual),
            'action_items_open': sum(1 for i in manual if not i.get('resolved')),
            'comments': sum(len(r.get('comments') or []) for r in responses),
            # The conversations themselves, not just how many there were.
            # Counting them told you something happened and then made you go
            # find it; on a timeline you are scanning, that is the whole cost.
            # Capped so a long-running argument on one standard can't bloat a
            # payload that is fetched every time a community is opened.
            'comment_list': [
                {
                    'standard': r.get('question_text', ''),
                    'question_id': r.get('question_id', ''),
                    'author': c.get('author') or c.get('username', ''),
                    'text': c.get('text', ''),
                    'at': c.get('at', ''),
                    'photo': c.get('photo', ''),
                }
                for r in responses
                for c in (r.get('comments') or [])
            ][:40],
            # The regional's own words about the visit. Read far more often
            # than any per-standard comment, because it explains the number.
            'notes': s.get('notes', ''),
            'notes_photo': s.get('notes_photo', ''),
        })

    # Per-standard record across the most recent visits, newest first. This is
    # what turns a pile of visits into "these three keep failing".
    recent = subs[:track_limit]
    track = {}
    for idx, s in enumerate(recent):
        for r in (s.get('responses') or []):
            cond = r.get('condition')
            if cond not in ('Pass', 'Fail'):
                continue
            key = r.get('question_id') or ('t:' + (r.get('question_text') or '').strip().lower())
            rec = track.setdefault(key, {
                'question_id': r.get('question_id', ''),
                'question_text': r.get('question_text', ''),
                'results': [None] * len(recent),
                'fails': 0,
                'seen': 0,
            })
            rec['results'][idx] = 'fixed' if (cond == 'Fail' and r.get('addressed')) else cond.lower()
            rec['seen'] += 1
            if cond == 'Fail':
                rec['fails'] += 1

    # Worst offenders first: most failures, then most recently failing.
    standards = sorted(track.values(),
                       key=lambda t: (-t['fails'], t['question_text'].lower()))

    return jsonify({
        'status': 'success',
        'community': community,
        'visits': visits,
        'total_visits': len(visits),
        'standards': standards,
        'track_dates': [s.get('submitted_at') for s in recent],
    }), 200


@app.route('/api/action-items/<submission_id>/standard/<question_id>/comments', methods=['POST'])
@login_required
def add_standard_comment(submission_id, question_id):
    """Post a comment on a failed standard, optionally with a photo.

    This is how a community reports that something has been fixed. It never
    changes the verdict or the score — a regional still has to review it and
    mark the item as addressed."""
    sub = next((s for s in inspection_service.get_all_submissions()
                if s.get('id') == submission_id), None)
    if not sub:
        return jsonify({'status': 'error', 'message': 'Visit not found'}), 404
    if not _can_see_community(sub.get('community')):
        return jsonify({'status': 'error', 'message': 'Not allowed for this community'}), 403

    if request.content_type and 'multipart/form-data' in request.content_type:
        text = InputSanitizer.sanitize_description(request.form.get('text', ''))
    else:
        text = InputSanitizer.sanitize_description((request.get_json(silent=True) or {}).get('text', ''))

    photo_path = ''
    if 'photo' in request.files:
        f = request.files['photo']
        if f and f.filename:
            is_valid, error_message = file_upload_handler.validate_file(f)
            if not is_valid:
                return jsonify({'status': 'error', 'message': error_message}), 400
            try:
                photo_path = file_upload_handler.save_file(
                    f, session.get('user', 'user'), sub.get('community', ''))
            except Exception as e:
                app.logger.error(f'Comment photo upload failed: {e}')

    if not text.strip() and not photo_path:
        return jsonify({'status': 'error', 'message': 'Write a comment or attach a photo.'}), 400

    comment = inspection_service.add_comment(
        submission_id, question_id, session.get('user'),
        resolve_display_name(session.get('user')), text, photo_path)
    if comment is None:
        return jsonify({'status': 'error', 'message': 'Standard not found on this visit'}), 404

    if photo_path and file_upload_handler.use_s3:
        comment = {**comment, 'photo_url': file_upload_handler.generate_presigned_url(photo_path)}

    activity_service.log(session.get('user'), 'standard_commented',
                         f"Commented on {next((r.get('question_text', '') for r in sub.get('responses', []) if r.get('question_id') == question_id), '')[:60]} "
                         f"at {sub.get('community')}",
                         meta={'community': sub.get('community'),
                               'submission_id': submission_id, 'question_id': question_id})

    # Tell the person who ran the visit that the community replied.
    try:
        notify_comment(sub, question_id, comment)
    except Exception as e:
        app.logger.error(f'Comment notification failed: {e}')

    return jsonify({'status': 'success', 'comment': comment}), 201


@app.route('/api/action-items/<submission_id>/item/<item_id>/comments', methods=['POST'])
@login_required
def add_item_comment(submission_id, item_id):
    """Comment on an item the inspector raised by hand. Same contract as the
    standards thread, so the community has one way to report progress."""
    sub = next((s for s in inspection_service.get_all_submissions()
                if s.get('id') == submission_id), None)
    if not sub:
        return jsonify({'status': 'error', 'message': 'Visit not found'}), 404
    if not _can_see_community(sub.get('community')):
        return jsonify({'status': 'error', 'message': 'Not allowed for this community'}), 403

    if request.content_type and 'multipart/form-data' in request.content_type:
        text = InputSanitizer.sanitize_description(request.form.get('text', ''))
    else:
        text = InputSanitizer.sanitize_description((request.get_json(silent=True) or {}).get('text', ''))

    photo_path = ''
    if 'photo' in request.files:
        f = request.files['photo']
        if f and f.filename:
            is_valid, error_message = file_upload_handler.validate_file(f)
            if not is_valid:
                return jsonify({'status': 'error', 'message': error_message}), 400
            try:
                photo_path = file_upload_handler.save_file(
                    f, session.get('user', 'user'), sub.get('community', ''))
            except Exception as e:
                app.logger.error(f'Comment photo upload failed: {e}')

    if not text.strip() and not photo_path:
        return jsonify({'status': 'error', 'message': 'Write a comment or attach a photo.'}), 400

    comment = inspection_service.add_item_comment(
        submission_id, item_id, session.get('user'),
        resolve_display_name(session.get('user')), text, photo_path)
    if comment is None:
        return jsonify({'status': 'error', 'message': 'Item not found on this visit'}), 404

    if photo_path and file_upload_handler.use_s3:
        comment = {**comment, 'photo_url': file_upload_handler.generate_presigned_url(photo_path)}

    item_text = next((i.get('text', '') for i in (sub.get('action_items') or [])
                      if i.get('id') == item_id), '')
    activity_service.log(session.get('user'), 'standard_commented',
                         f"Commented on {item_text[:60]} at {sub.get('community')}",
                         meta={'community': sub.get('community'),
                               'submission_id': submission_id, 'item_id': item_id})
    try:
        notify_item_comment(sub, item_text, comment)
    except Exception as e:
        app.logger.error(f'Comment notification failed: {e}')

    return jsonify({'status': 'success', 'comment': comment}), 201


@app.route('/api/action-items/<submission_id>/item/<item_id>/comments/<comment_id>',
           methods=['DELETE'])
@login_required
def delete_item_comment(submission_id, item_id, comment_id):
    """Delete a comment on an ad-hoc item. Authors and admins only."""
    sub = next((s for s in inspection_service.get_all_submissions()
                if s.get('id') == submission_id), None)
    if not sub:
        return jsonify({'status': 'error', 'message': 'Visit not found'}), 404
    if not _can_see_community(sub.get('community')):
        return jsonify({'status': 'error', 'message': 'Not allowed for this community'}), 403
    ok = inspection_service.delete_item_comment(submission_id, item_id, comment_id,
                                                session.get('user'), is_admin())
    if not ok:
        return jsonify({'status': 'error', 'message': 'You can only delete your own comments.'}), 403
    return jsonify({'status': 'success'}), 200


def notify_item_comment(sub, item_text, comment):
    """Same audience as a comment on a standard: whoever is on the other side."""
    if not email_service.enabled:
        return
    author = session.get('user')
    recipients = []
    inspector = sub.get('username')
    if inspector and inspector != author:
        acct = resolve_account_context(inspector)
        if acct.get('email'):
            recipients.append(acct['email'])
    for addr in region_leader_emails(sub.get('community', '')):
        if addr not in recipients:
            recipients.append(addr)
    for addr in community_account_emails(sub.get('community', ''), exclude_username=author):
        if addr not in recipients:
            recipients.append(addr)
    if not recipients:
        return
    email_service.send_standard_comment(
        recipients, sub.get('community', ''), item_text,
        comment.get('author', ''), comment.get('text', ''),
        bool(comment.get('photo')))


@app.route('/api/action-items/<submission_id>/standard/<question_id>/comments/<comment_id>',
           methods=['DELETE'])
@login_required
def delete_standard_comment(submission_id, question_id, comment_id):
    """Delete a comment. Authors can remove their own; admins can remove any."""
    sub = next((s for s in inspection_service.get_all_submissions()
                if s.get('id') == submission_id), None)
    if not sub:
        return jsonify({'status': 'error', 'message': 'Visit not found'}), 404
    if not _can_see_community(sub.get('community')):
        return jsonify({'status': 'error', 'message': 'Not allowed for this community'}), 403
    ok = inspection_service.delete_comment(submission_id, question_id, comment_id,
                                           session.get('user'), is_admin())
    if not ok:
        return jsonify({'status': 'error', 'message': 'You can only delete your own comments.'}), 403
    return jsonify({'status': 'success'}), 200


def notify_comment(sub, question_id, comment):
    """Tell everyone on the other side of the conversation.

    A comment travels both ways: when the community reports a fix, the
    inspector and the region's leaders hear about it; when leadership replies
    or asks for more, the community hears about it. Without the second half an
    Executive Director would have to keep checking the app to notice a
    question, which nobody does. Best-effort: never blocks the comment."""
    if not email_service.enabled:
        return
    standard = next((r.get('question_text', '') for r in sub.get('responses', [])
                     if r.get('question_id') == question_id), '')
    author = session.get('user')
    recipients = []
    inspector = sub.get('username')
    if inspector and inspector != author:
        acct = resolve_account_context(inspector)
        if acct.get('email'):
            recipients.append(acct['email'])
    for addr in region_leader_emails(sub.get('community', '')):
        if addr not in recipients:
            recipients.append(addr)
    # The community's own account, unless they are the one who just wrote.
    for addr in community_account_emails(sub.get('community', ''), exclude_username=author):
        if addr not in recipients:
            recipients.append(addr)
    if not recipients:
        return
    email_service.send_standard_comment(
        recipients, sub.get('community', ''), standard,
        comment.get('author', ''), comment.get('text', ''),
        bool(comment.get('photo')))


@app.route('/api/action-items/<submission_id>/standard/<question_id>/resolve', methods=['POST'])
@login_required
def resolve_failed_standard(submission_id, question_id):
    """Mark a failed standard as addressed between visits.

    This never edits the inspection: the item still reads Fail and the score of
    that visit is unchanged. It records who verified the fix, when, an optional
    note and a photo, so the team has a follow-up trail without rewriting
    history. A later visit that passes the standard still clears it the usual
    way.

    Closing an item is deliberately reserved for leadership. A community can
    report that something is fixed by commenting, but somebody from outside the
    community confirms it — otherwise the "current" score would be self-awarded."""
    sub = next((s for s in inspection_service.get_all_submissions()
                if s.get('id') == submission_id), None)
    if not sub:
        return jsonify({'status': 'error', 'message': 'Inspection not found'}), 404
    if not can_verify_fixes():
        return jsonify({'status': 'error',
                        'message': 'Only a regional, corporate or admin can mark an item as '
                                   'addressed. Add a comment to report that it is fixed.'}), 403
    if not is_admin():
        allowed = set(regional_communities())
        if sub.get('community') not in allowed:
            return jsonify({'status': 'error', 'message': 'Not allowed for this community'}), 403

    # Accepts JSON, or multipart when a photo of the fix is attached.
    if request.content_type and 'multipart/form-data' in request.content_type:
        resolved = (request.form.get('resolved', 'true').lower() != 'false')
        note = InputSanitizer.sanitize_description(request.form.get('note', ''))
    else:
        data = request.get_json(silent=True) or {}
        resolved = bool(data.get('resolved', True))
        note = InputSanitizer.sanitize_description(data.get('note', ''))

    # Photo of the fix. Same handler as inspection photos, so it lands in S3
    # when S3 is configured and on disk otherwise.
    photo_path = ''
    if resolved and 'photo' in request.files:
        f = request.files['photo']
        if f and f.filename:
            is_valid, error_message = file_upload_handler.validate_file(f)
            if not is_valid:
                return jsonify({'status': 'error', 'message': error_message}), 400
            try:
                photo_path = file_upload_handler.save_file(
                    f, session.get('user', 'user'), sub.get('community', ''))
            except Exception as e:
                app.logger.error(f'Fix photo upload failed: {e}')

    resp = inspection_service.resolve_response(
        submission_id, question_id, session.get('user'),
        note=note, photo=photo_path, resolved=resolved)
    if resp is None:
        return jsonify({'status': 'error', 'message': 'Standard not found on this visit'}), 404

    activity_service.log(session.get('user'),
                         'standard_addressed' if resolved else 'standard_reopened',
                         f'{"Marked addressed" if resolved else "Reopened"}: '
                         f'{resp.get("question_text", "")[:60]} at {sub.get("community")}',
                         meta={'community': sub.get('community'),
                               'submission_id': submission_id, 'question_id': question_id})
    return jsonify({'status': 'success', 'response': resp,
                    'message': 'Marked as addressed.' if resolved else 'Reopened.'}), 200


@app.route('/api/action-items/<submission_id>/<item_id>/resolve', methods=['POST'])
@login_required
def resolve_action_item(submission_id, item_id):
    """Mark a manual action item as done (or reopen it), with an optional note.

    The visit itself is never edited — it stays a faithful record of what was
    seen that day. We only track the follow-up on top of it.

    Closing is leadership's call, the same as for a failed standard. These items
    don't affect the score, so the reason isn't integrity — it's that people
    should only have to remember one rule: the community reports, a regional
    closes. An exception here would be one more thing to explain."""
    data = request.get_json(silent=True) or {}
    resolved = bool(data.get('resolved', True))
    note = InputSanitizer.sanitize_description(data.get('note', ''))

    sub = next((s for s in inspection_service.get_all_submissions()
                if s.get('id') == submission_id), None)
    if not sub:
        return jsonify({'status': 'error', 'message': 'Visit not found'}), 404
    if not can_verify_fixes():
        return jsonify({'status': 'error',
                        'message': 'Only a regional, corporate or admin can close an item. '
                                   'Add a comment to report that it is done.'}), 403
    if not is_admin():
        allowed = set(regional_communities())
        if sub.get('community') not in allowed:
            return jsonify({'status': 'error', 'message': 'Not allowed for this community'}), 403

    item = inspection_service.resolve_action_item(
        submission_id, item_id, session.get('user'), note=note, resolved=resolved)
    if item is None:
        return jsonify({'status': 'error', 'message': 'Action item not found'}), 404

    activity_service.log(session.get('user'),
                         'action_item_resolved' if resolved else 'action_item_reopened',
                         f'{"Resolved" if resolved else "Reopened"} action item at {sub.get("community")}',
                         meta={'community': sub.get('community')})
    return jsonify({'status': 'success', 'item': item,
                    'message': 'Marked as done.' if resolved else 'Reopened.'}), 200


@app.route('/api/people', methods=['GET'])
@login_required
def list_people():
    """Admin-only: everyone with access to Atlas Excellence, from all sources —
    stored users (admins and community staff) and region/corporate members —
    in one list with their role, scope, photo and activity."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403

    regions = region_service.get_all_regions()
    region_name = {r.get('id'): r.get('name', r.get('id')) for r in regions}

    # Inspection counts per person (matched on the name shown on submissions).
    counts, last_seen = {}, {}
    for s in inspection_service.get_all_submissions():
        key = (s.get('inspector_name') or s.get('username') or '').strip().lower()
        if not key:
            continue
        counts[key] = counts.get(key, 0) + 1
        ts = s.get('submitted_at') or ''
        if ts > last_seen.get(key, ''):
            last_seen[key] = ts

    def activity_for(name, username):
        for k in ((name or '').strip().lower(), (username or '').strip().lower()):
            if k in counts:
                return counts[k], last_seen.get(k, '')
        return 0, ''

    people = []

    # 1) Region + corporate members (their login comes from the roster)
    for region in regions:
        if region.get('id') == 'unassigned':
            continue
        is_corp = region.get('kind') == CORPORATE_KIND
        for idx, leader in enumerate(region.get('leadership', [])):
            name = (leader.get('name') or '').strip()
            if not name or name.lower() == 'open':
                continue
            username = (leader.get('username') or '').strip() or slugify_name(name)
            n, last = activity_for(name, username)
            people.append({
                'username': username, 'name': name,
                'email': (leader.get('email') or '').strip(),
                'title': (leader.get('role') or '').strip(),
                'role': 'corporate' if is_corp else 'regional',
                'scope': 'All communities' if is_corp else region_name.get(region.get('id'), ''),
                'region_id': region.get('id'),
                'source': 'region', 'index': idx,
                'photo': profile_service.get_leader_photo(region.get('id', ''), name),
                'inspections': n, 'last_visit': last,
                'must_change': profile_service.get_must_change(username),
                'admin_extra': profile_service.get_admin_extra(username),
            })

    # 2) Stored users (admins, community staff, admin-created accounts)
    for u in user_service.get_all():
        username = u.get('username')
        name = profile_service.get_display_name(username) or u.get('display_name') or username
        n, last = activity_for(name, username)
        role = u.get('role', 'staff')
        people.append({
            'username': username, 'name': name,
            'email': (u.get('email') or '').strip(),
            'title': '',
            'role': role,
            'scope': ('All communities' if role == 'admin'
                      else (' · '.join(account_communities(u))
                            or region_name.get(u.get('region_id'), '') or '—')),
            'region_id': u.get('region_id'),
            'community': u.get('community'),
            'communities': account_communities(u),
            'source': 'user',
            'photo': profile_service.get_photo(username),
            'inspections': n, 'last_visit': last,
            'must_change': profile_service.get_must_change(username),
            'admin_extra': profile_service.get_admin_extra(username),
        })

    # Sign-in activity: who is using the app right now, who never has.
    for p in people:
        pres = presence_service.get(p['username'])
        p['last_login'] = pres['last_login']
        p['last_active'] = pres['last_seen']
        p['logins'] = pres['logins']
        p['online'] = pres['active']

    people.sort(key=lambda p: ((p['name'] or p['username']).lower()))

    # Two entries sharing a username would fight over the same login, so flag
    # them for the admin instead of silently letting one win.
    seen = {}
    for p in people:
        seen.setdefault(p['username'].lower(), []).append(p)
    conflicts = []
    for uname, group in seen.items():
        if len(group) > 1:
            for p in group:
                p['duplicate'] = True
            conflicts.append({'username': uname,
                              'where': [g['scope'] or g['role'] for g in group],
                              'name': group[0]['name']})

    counts_by_role = {}
    for p in people:
        counts_by_role[p['role']] = counts_by_role.get(p['role'], 0) + 1
    return jsonify({'status': 'success', 'people': people, 'counts': counts_by_role,
                    'can_grant_admin': is_native_admin(),
                    'online_now': sum(1 for p in people if p['online']),
                    'never_signed_in': [p['username'] for p in people if not p['last_login']],
                    'conflicts': conflicts,
                    'missing_email': [p['username'] for p in people
                                      if not p['email'] and p['role'] in ('regional', 'corporate')],
                    'regions': [{'id': r.get('id'), 'name': r.get('name'), 'kind': r.get('kind')}
                                for r in regions if r.get('id') != 'unassigned'],
                    'communities': all_communities()}), 200


@app.route('/api/people', methods=['POST'])
@login_required
def create_person():
    """Admin-only: add anyone — staff, regional, corporate or administrator —
    from the People view. Creates their login, sets a temporary password they
    must change at first sign-in, and emails it to them when we have an
    address."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    data = request.get_json(silent=True) or {}
    name = InputSanitizer.sanitize_string(data.get('name', ''), max_length=120)
    email = InputSanitizer.sanitize_string(data.get('email', ''), max_length=160)
    title = InputSanitizer.sanitize_string(data.get('title', ''), max_length=80)
    role = data.get('role') if data.get('role') in ('admin', 'staff', 'regional', 'corporate') else None
    region_id = InputSanitizer.sanitize_string(data.get('region_id', ''), max_length=50)
    community = InputSanitizer.sanitize_community_name(data.get('community') or '') or None

    if not name:
        return jsonify({'status': 'error', 'message': 'Name is required'}), 400
    if not role:
        return jsonify({'status': 'error', 'message': 'Pick a role'}), 400
    if email and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return jsonify({'status': 'error', 'message': 'Enter a valid email address'}), 400
    if role == 'staff' and not community:
        return jsonify({'status': 'error', 'message': 'Pick a community for staff'}), 400

    username = generate_unique_username(name)
    password = generate_password()
    pw_hash = generate_password_hash(password)

    if role in ('regional', 'corporate'):
        dest = next((r for r in region_service.get_all_regions() if r.get('id') == region_id), None)
        if not dest or dest.get('id') == 'unassigned':
            return jsonify({'status': 'error', 'message': 'Pick the region or group'}), 400
        is_corp = dest.get('kind') == CORPORATE_KIND
        if role == 'corporate' and not is_corp:
            return jsonify({'status': 'error', 'message': 'Corporate members go in the Corporate group.'}), 400
        if role == 'regional' and is_corp:
            return jsonify({'status': 'error', 'message': 'Pick a region — Corporate is company-wide.'}), 400
        if not region_service.add_leader(region_id, name, title or '', email, username=username):
            return jsonify({'status': 'error', 'message': 'Could not add this person'}), 400
        profile_service.set_password_hash(username, pw_hash)
        scope = dest.get('name')
    else:
        user_service.create(username, display_name=name, role=role, password_hash=pw_hash,
                            community=community if role == 'staff' else None,
                            communities=(requested_communities(data, community)
                                         if role == 'staff' else None),
                            region_id=None, created_by=session.get('user'), email=email or None)
        scope = community if role == 'staff' else 'All communities'

    profile_service.set_display_name(username, name)
    profile_service.set_must_change(username, True)

    emailed = False
    if email and email_service.enabled:
        try:
            ok, _ = email_service.send_welcome(email, name, username, password,
                                               role_label_for(role, community))
            emailed = bool(ok)
        except Exception as e:
            app.logger.error(f'Welcome email failed: {e}')
    try:
        admin_notify = settings_service.get_email_settings().get('admin_notify', [])
        if admin_notify:
            email_service.send_new_user_alert(admin_notify, name, username,
                                              role_label_for(role, community),
                                              session.get('user'))
    except Exception:
        pass

    # A community account inherits whatever is already open there. Turnover is
    # constant, so somebody's first day is usually mid-conversation: send the
    # current findings straight away rather than leaving them to discover a
    # backlog on their own.
    if role == 'staff' and community and email:
        try:
            for comm in requested_communities(data, community):
                send_community_handover(comm, [email])
        except Exception as e:
            app.logger.error(f'Handover email failed: {e}')

    activity_service.log(session.get('user'), 'person_created',
                         f'Added {name} ({username}) as {role}')
    return jsonify({'status': 'success', 'username': username, 'password': password,
                    'display_name': name, 'role': role, 'scope': scope, 'emailed': emailed,
                    'message': f'{name} added.'}), 200


@app.route('/api/people/<username>', methods=['PUT'])
@login_required
def update_person(username):
    """Admin-only: edit a person's name, email, title, role and scope.
    The username (their login) is never changed, so history stays intact."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    data = request.get_json(silent=True) or {}
    name = InputSanitizer.sanitize_string(data.get('name', ''), max_length=120)
    email = InputSanitizer.sanitize_string(data.get('email', ''), max_length=160)
    title = InputSanitizer.sanitize_string(data.get('title', ''), max_length=80)
    if not name:
        return jsonify({'status': 'error', 'message': 'Name is required'}), 400
    if email and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return jsonify({'status': 'error', 'message': 'Enter a valid email address'}), 400

    requested_role = data.get('role') if data.get('role') in ('admin', 'staff', 'regional', 'corporate') else None
    target_region = InputSanitizer.sanitize_string(data.get('region_id', ''), max_length=50)
    community = InputSanitizer.sanitize_community_name(data.get('community') or '') or None
    communities = requested_communities(data, community)

    def preserve_password(user_rec):
        """Carry an account's password across storage types so moving someone
        between roles never changes how they sign in."""
        existing = profile_service.get_password_hash(username)
        if not existing and user_rec and user_rec.get('password_hash'):
            profile_service.set_password_hash(username, user_rec['password_hash'])

    region, index, leader = region_service.find_leader_by_username(username)

    # ---- Currently a roster member (regional / corporate) ----
    if leader is not None:
        current_region = region.get('id')
        old_name = leader.get('name', '')
        moving_off_roster = requested_role in ('admin', 'staff')

        if moving_off_roster:
            # Becomes a stored account; keep their login and password working.
            region_service.remove_leader(current_region, index)
            if not user_service.exists(username):
                user_service.ensure(username, display_name=name, role=requested_role,
                                    community=community if requested_role == 'staff' else None,
                                    email=email, created_by=session.get('user'))
            else:
                user_service.update(username, display_name=name, role=requested_role,
                                    community=community if requested_role == 'staff' else None,
                                    communities=(communities if requested_role == 'staff' else []),
                                    email=email)
            profile_service.set_display_name(username, name)
            activity_service.log(session.get('user'), 'person_updated',
                                 f'Changed {name} to {requested_role}')
            return jsonify({'status': 'success', 'message': f'{name} is now {requested_role}.'}), 200

        dest = target_region or current_region
        if dest != current_region:
            region_service.remove_leader(current_region, index)
            region_service.add_leader(dest, name, title or leader.get('role', ''), email, username=username)
        else:
            region_service.update_leader(current_region, index, name, title or leader.get('role', ''), email)
        if old_name and old_name != name:
            photo = profile_service.get_leader_photo(current_region, old_name)
            if photo:
                profile_service.set_leader_photo(dest, name, photo)
        profile_service.set_display_name(username, name)
        activity_service.log(session.get('user'), 'person_updated', f'Updated {name} ({username})')
        return jsonify({'status': 'success', 'message': f'{name} updated.'}), 200

    # ---- Currently a stored user (admin / staff) ----
    if user_service.exists(username):
        rec = user_service.get(username) or {}

        if requested_role in ('regional', 'corporate'):
            # Moving onto a region/corporate roster. Keep the same login and
            # password by pinning the username and copying the hash across.
            if not target_region:
                return jsonify({'status': 'error',
                                'message': 'Pick the region or group for this person.'}), 400
            dest = next((r for r in region_service.get_all_regions() if r.get('id') == target_region), None)
            if not dest or dest.get('id') == 'unassigned':
                return jsonify({'status': 'error', 'message': 'Unknown region or group'}), 400
            is_corp_dest = dest.get('kind') == CORPORATE_KIND
            if requested_role == 'corporate' and not is_corp_dest:
                return jsonify({'status': 'error',
                                'message': 'Corporate members must be placed in the Corporate group.'}), 400
            if requested_role == 'regional' and is_corp_dest:
                return jsonify({'status': 'error',
                                'message': 'Pick a region — Corporate is company-wide.'}), 400
            preserve_password(rec)
            region_service.add_leader(target_region, name, title or '', email, username=username)
            user_service.delete(username)
            profile_service.set_display_name(username, name)
            activity_service.log(session.get('user'), 'person_updated',
                                 f'Moved {name} to {dest.get("name")} ({requested_role})')
            return jsonify({'status': 'success',
                            'message': f'{name} moved to {dest.get("name")}.'}), 200

        fields = {'display_name': name, 'email': email}
        if requested_role:
            fields['role'] = requested_role
            if requested_role == 'admin':
                fields['community'] = None
                fields['communities'] = []
        if 'community' in data and requested_role != 'admin':
            fields['community'] = community
            fields['communities'] = communities
        user_service.update(username, **fields)
        profile_service.set_display_name(username, name)
        activity_service.log(session.get('user'), 'person_updated', f'Updated {name} ({username})')
        return jsonify({'status': 'success', 'message': f'{name} updated.'}), 200

    return jsonify({'status': 'error', 'message': 'Person not found'}), 404


@app.route('/api/people/<username>/admin-privileges', methods=['POST'])
@login_required
def set_admin_privileges(username):
    """Grant or revoke administrator privileges on top of someone's own role.

    Reserved for the main administrator: people who merely hold the accessory
    cannot hand it out, which prevents privileges from spreading quietly."""
    if not is_native_admin():
        return jsonify({'status': 'error',
                        'message': 'Only the main administrator can change admin privileges.'}), 403
    data = request.get_json(silent=True) or {}
    grant = bool(data.get('grant'))

    if username == session.get('user'):
        return jsonify({'status': 'error',
                        'message': 'You cannot change your own privileges.'}), 400

    acct = resolve_account_context(username)
    if not acct.get('exists'):
        return jsonify({'status': 'error', 'message': 'Person not found'}), 404

    # Administrators already have everything; the accessory is for other roles.
    rec = user_service.get(username) or {}
    if rec.get('role') == 'admin':
        return jsonify({'status': 'error',
                        'message': 'This person is already an Administrator.'}), 400

    profile_service.set_admin_extra(username, grant)
    activity_service.log(session.get('user'),
                         'admin_privileges_granted' if grant else 'admin_privileges_revoked',
                         f'{"Granted" if grant else "Revoked"} admin privileges for '
                         f'{acct.get("display_name") or username}')
    return jsonify({'status': 'success', 'grant': grant,
                    'message': f'{acct.get("display_name") or username} '
                               f'{"now has" if grant else "no longer has"} admin privileges.'}), 200


@app.route('/api/people/<username>', methods=['DELETE'])
@login_required
def delete_person(username):
    """Admin-only: remove someone's access. Roster members (region/corporate)
    come off their roster; stored users are deleted. Submitted inspections are
    never touched — they keep the inspector recorded on them."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    if username == session.get('user'):
        return jsonify({'status': 'error', 'message': 'You cannot remove your own account'}), 400

    region, index, leader = region_service.find_leader_by_username(username)
    if leader is not None:
        name = leader.get('name', username)
        if not region_service.remove_leader(region.get('id'), index):
            return jsonify({'status': 'error', 'message': 'Could not remove this person'}), 400
        activity_service.log(session.get('user'), 'person_removed',
                             f'Removed {name} from {region.get("name")}')
        presence_service.forget(username)
        return jsonify({'status': 'success', 'message': f'{name} removed.'}), 200

    if user_service.exists(username):
        rec = user_service.get(username) or {}
        name = rec.get('display_name') or username
        if not user_service.delete(username):
            return jsonify({'status': 'error', 'message': 'Could not remove this account'}), 400
        activity_service.log(session.get('user'), 'person_removed', f'Removed account {username}')
        presence_service.forget(username)
        return jsonify({'status': 'success', 'message': f'{name} removed.'}), 200

    return jsonify({'status': 'error', 'message': 'Person not found'}), 404


@app.route('/api/activity/live', methods=['GET'])
@login_required
def activity_live():
    """Admin-only feed of what's happening in the app: who signed in, who
    submitted a visit, what changed. Polled by the dashboard panel."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    try:
        limit = max(1, min(int(request.args.get('limit', 25)), 100))
    except (TypeError, ValueError):
        limit = 25
    events = []
    for e in activity_service.get_recent(limit=limit):
        events.append({
            'id': e.get('id'),
            'username': e.get('username'),
            'name': resolve_display_name(e.get('username')) or e.get('username'),
            'type': e.get('type'),
            'detail': e.get('detail', ''),
            'timestamp': e.get('timestamp'),
            'community': (e.get('meta') or {}).get('community', ''),
            # Enough to open the exact item the event is about.
            'submission_id': (e.get('meta') or {}).get('submission_id', ''),
            'question_id': (e.get('meta') or {}).get('question_id', ''),
            'item_id': (e.get('meta') or {}).get('item_id', ''),
        })
    online = [{'username': u, 'name': resolve_display_name(u) or u}
              for u in presence_service.active_usernames()]
    return jsonify({'status': 'success', 'events': events, 'online': online}), 200


@app.route('/api/settings/email/summary', methods=['GET'])
@login_required
def email_notification_summary():
    """Admin-only: a consolidated 'who receives what' view. Rolls up every
    source of outgoing email — inspection subscribers, region leadership (who
    are copied automatically on their own region), admin alerts, and the
    Clinical/Ops routing lists — into one row per email address."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403

    s = settings_service.get_email_settings()
    regions = region_service.get_all_regions()
    region_name = {r.get('id'): r.get('name', r.get('id')) for r in regions}
    people = {}   # lowercase email -> record

    def add(email, name, label, detail='', kind='info'):
        addr = (email or '').strip()
        if not addr:
            return
        key = addr.lower()
        rec = people.setdefault(key, {'email': addr, 'name': (name or '').strip(), 'items': []})
        if name and not rec['name']:
            rec['name'] = name.strip()
        rec['items'].append({'label': label, 'detail': detail, 'kind': kind})

    # 1) Inspection report subscribers (explicitly configured)
    for sub in s.get('subscribers', []):
        regs = sub.get('regions') or []
        insp = sub.get('inspectors') or []
        if not regs and not insp:
            add(sub['email'], sub.get('name'), 'Inspection reports',
                'All inspections, every region', 'all')
        else:
            if regs:
                add(sub['email'], sub.get('name'), 'Inspection reports',
                    'Regions: ' + ', '.join(region_name.get(r, r) for r in regs), 'scoped')
            if insp:
                add(sub['email'], sub.get('name'), 'Inspection reports',
                    'Inspectors: ' + ', '.join(insp), 'scoped')

    # 2) Region leadership — always copied on their own region's inspections,
    #    plus move-in reminders and completion summaries for their communities.
    for r in regions:
        if r.get('id') == 'unassigned':
            continue
        for leader in (r.get('leadership') or []):
            addr = (leader.get('email') or '').strip()
            if not addr:
                continue
            add(addr, leader.get('name'), 'Inspection reports',
                f"{r.get('name')} region (automatic)", 'auto')
            add(addr, leader.get('name'), 'Move-in emails',
                f"Reminders and completion summaries — {r.get('name')}", 'auto')

    # 3) Admin alerts
    for addr in s.get('admin_notify', []):
        add(addr, '', 'Admin alerts',
            'New users, password changes and resets, move-in summaries', 'admin')
        add(addr, '', 'Daily activity digest',
            'Once a day: sign-ins, visits, items addressed, account changes', 'admin')
    # 4) Routed comments — one configurable list per company-level team
    for route in settings_service.ROUTES:
        label = settings_service.ROUTE_LABELS[route]
        for addr in s.get(route, []):
            add(addr, '', f'{label} comments',
                f'Visit comments directed to {label}', 'route')

    # A recipient can be removed from here when it came from a configurable list
    # (subscribers / admin alerts / clinical / ops). Region leaders are copied
    # automatically from their region, so those are managed in Regions instead.
    for rec in people.values():
        kinds = {i['kind'] for i in rec['items']}
        rec['removable'] = bool(kinds - {'auto'})
        rec['auto_only'] = kinds == {'auto'}

    rows = sorted(people.values(), key=lambda p: (p['name'] or p['email']).lower())
    # Leaders without an email on file can't be reached — surface that too.
    missing = []
    for r in regions:
        if r.get('id') == 'unassigned':
            continue
        for leader in (r.get('leadership') or []):
            if not (leader.get('email') or '').strip():
                nm = (leader.get('name') or '').strip()
                if nm and nm.lower() != 'open':
                    missing.append({'name': nm, 'region': r.get('name')})
    return jsonify({'status': 'success', 'people': rows, 'missing_email': missing,
                    'email_enabled': bool(email_service.enabled)}), 200


@app.route('/api/settings/email/subscriber', methods=['POST'])
@login_required
def add_email_subscriber():
    """Admin-only: add ONE inspection-report subscriber, leaving every existing
    one untouched. If the address is already subscribed, its scope is updated."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    import re as _re
    data = request.get_json(silent=True) or {}
    email = (data.get('email') or '').strip()
    if not email or not _re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return jsonify({'status': 'error', 'message': 'Enter a valid email address.'}), 400

    valid_ids = {r.get('id') for r in region_service.get_all_regions() if r.get('id') != 'unassigned'}
    valid_names = set(leadership_names())
    regions = [r for r in (data.get('regions') or []) if r in valid_ids]
    inspectors = [n for n in (data.get('inspectors') or []) if n in valid_names]

    current = settings_service.get_email_settings()['subscribers']
    subs, replaced = [], False
    for s in current:
        if (s.get('email') or '').lower() == email.lower():
            subs.append({'email': email, 'name': (data.get('name') or s.get('name') or '').strip(),
                         'regions': regions, 'inspectors': inspectors})
            replaced = True
        else:
            subs.append(s)
    if not replaced:
        subs.append({'email': email, 'name': (data.get('name') or '').strip(),
                     'regions': regions, 'inspectors': inspectors})

    saved = settings_service.set_email_settings(subscribers=subs)
    activity_service.log(session.get('user'), 'email_subscriber_added',
                         f'{"Updated" if replaced else "Added"} notification subscriber {email}')
    return jsonify({'status': 'success', 'updated': replaced,
                    'count': len(saved['subscribers']),
                    'message': f'{email} {"updated" if replaced else "added"}.'}), 200


@app.route('/api/settings/email/recipient', methods=['DELETE'])
@login_required
def remove_email_recipient():
    """Admin-only: stop sending to an address. Removes it from every
    configurable list (inspection subscribers, admin alerts, Clinical, Ops).
    Region leadership is NOT touched here — those addresses live on the region
    and are managed in Regions."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    data = request.get_json(silent=True) or {}
    target = (data.get('email') or '').strip().lower()
    if not target:
        return jsonify({'status': 'error', 'message': 'Email is required'}), 400

    s = settings_service.get_email_settings()
    subs = [x for x in s['subscribers'] if (x.get('email') or '').lower() != target]
    admin_notify = [a for a in s['admin_notify'] if a.lower() != target]
    routed = {r: [a for a in s[r] if a.lower() != target]
              for r in settings_service.ROUTES}

    removed = (len(subs) != len(s['subscribers']) or len(admin_notify) != len(s['admin_notify'])
               or any(len(routed[r]) != len(s[r]) for r in settings_service.ROUTES))
    if not removed:
        return jsonify({'status': 'error',
                        'message': 'That address is not in a list you can edit here. '
                                   'Region leaders are managed in Regions.'}), 404

    settings_service.set_email_settings(subscribers=subs, admin_notify=admin_notify,
                                        **routed)
    activity_service.log(session.get('user'), 'email_recipient_removed',
                         f'Removed {target} from notification lists')
    return jsonify({'status': 'success', 'message': f'{target} will no longer receive notifications.'}), 200


@app.route('/api/admin/reset-inspections', methods=['POST'])
@login_required
def admin_reset_inspections():
    """Admin-only: wipe every submitted inspection so the app can go live with a
    clean slate. Scores, action items and the dashboard's recent activity all
    derive from these records, so they clear too. Users, regions, communities,
    standards, survey types and move-ins are NOT touched.

    A timestamped snapshot of the affected files is written to data/backups/
    first, so the reset can be undone by restoring those files on the server.
    Requires the caller to type the word RESET to confirm."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    data = request.get_json(silent=True) or {}
    if (data.get('confirm') or '').strip().upper() != 'RESET':
        return jsonify({'status': 'error',
                        'message': 'Type RESET to confirm this action.'}), 400
    try:
        import shutil
        from datetime import datetime as _dt
        stamp = _dt.now().strftime('%Y%m%d-%H%M%S')
        backup_dir = os.path.join(DATA_FOLDER, 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        saved = []
        for name in ('inspections.json', 'activity.json'):
            src = os.path.join(DATA_FOLDER, name)
            if os.path.exists(src):
                dst = os.path.join(backup_dir, f'{name.rsplit(".", 1)[0]}-{stamp}.json')
                shutil.copy2(src, dst)
                saved.append(os.path.basename(dst))

        removed = inspection_service.reset_all()
        purged = activity_service.purge_types(['inspection_submitted'])

        activity_service.log(session.get('user'), 'data_reset',
                             f'Reset visit data — {removed} visits cleared')
        app.logger.warning('INSPECTION DATA RESET by %s — %d submissions removed (backup: %s)',
                           session.get('user'), removed, ', '.join(saved) or 'none')
        return jsonify({'status': 'success', 'removed': removed,
                        'activity_removed': purged, 'backups': saved,
                        'message': f'{removed} inspection(s) cleared. Backup saved on the server.'}), 200
    except Exception as e:
        app.logger.error(f'Reset inspections failed: {str(e)}')
        return jsonify({'status': 'error',
                        'message': 'Internal server error during reset. Nothing was changed.'}), 500


@app.route('/api/users/<username>', methods=['DELETE'])
@login_required
def delete_user(username):
    """Admin-only: remove a created user. Cannot remove yourself."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    if username == session.get('user'):
        return jsonify({'status': 'error', 'message': 'You cannot delete your own account'}), 400
    if not user_service.delete(username):
        return jsonify({'status': 'error', 'message': 'User not found'}), 404
    return jsonify({'status': 'success', 'message': 'User removed'}), 200


RESOURCE_ALLOWED_EXT = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx',
                        'txt', 'csv', 'png', 'jpg', 'jpeg', 'gif', 'webp', 'zip'}


@app.route('/standards/print')
@login_required
def standards_print():
    """Printable one-pager of all standards (text + pass criteria + guideline)."""
    questions = question_manager.get_all_active_questions()
    return render_template('standards_print.html', questions=questions)


@app.route('/standards/pdf')
@login_required
def standards_pdf():
    """Server-generated PDF of all standards — reliable download (no browser print)."""
    import os
    import html as _html
    from io import BytesIO
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    ListFlowable, ListItem, Image, HRFlowable)

    questions = question_manager.get_all_active_questions()
    navy = colors.HexColor('#00285c')
    styles = getSampleStyleSheet()
    s_title = ParagraphStyle('t', parent=styles['Title'], textColor=navy, fontSize=18, alignment=0, spaceAfter=2)
    s_sub = ParagraphStyle('s', parent=styles['Normal'], textColor=colors.HexColor('#6b7280'), fontSize=10, spaceAfter=14)
    s_h = ParagraphStyle('h', parent=styles['Heading2'], textColor=navy, fontSize=13, spaceBefore=14, spaceAfter=4)
    s_lbl = ParagraphStyle('l', parent=styles['Normal'], textColor=colors.HexColor('#94a3b8'), fontSize=8, spaceBefore=6, spaceAfter=2)
    s_body = ParagraphStyle('b', parent=styles['Normal'], fontSize=10, leading=14, textColor=colors.HexColor('#334155'))

    def esc(x):
        return _html.escape(str(x or '')).replace('\n', '<br/>')

    story = []
    logo = os.path.join(os.path.dirname(__file__), 'static', 'atlas-logo.png')
    if os.path.exists(logo):
        try:
            img = Image(logo)
            ratio = img.imageWidth / float(img.imageHeight)
            img.drawHeight = 0.45 * inch
            img.drawWidth = 0.45 * inch * ratio
            img.hAlign = 'LEFT'
            story += [img, Spacer(1, 8)]
        except Exception:
            pass
    story.append(Paragraph('Inspection Standards &amp; Pass Criteria', s_title))
    story.append(Paragraph('Atlas Senior Living &mdash; Communities Standards', s_sub))

    for i, q in enumerate(questions, 1):
        story.append(Paragraph(f"{i}. {esc(q.get('text'))}", s_h))
        crit = q.get('pass_criteria') or []
        if crit:
            story.append(Paragraph('TO PASS, MUST INCLUDE', s_lbl))
            story.append(ListFlowable(
                [ListItem(Paragraph(esc(c), s_body), leftIndent=12) for c in crit],
                bulletType='bullet', start='•', leftIndent=16))
        g = (q.get('interpretive_guideline') or '').strip()
        if g:
            story.append(Paragraph('INTERPRETIVE GUIDELINE', s_lbl))
            story.append(Paragraph(esc(g), s_body))
        story.append(Spacer(1, 8))
        story.append(HRFlowable(width='100%', color=colors.HexColor('#eef1f6')))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=LETTER, title='Atlas Excellence',
                            topMargin=0.6 * inch, bottomMargin=0.6 * inch,
                            leftMargin=0.7 * inch, rightMargin=0.7 * inch)
    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    from flask import Response
    return Response(pdf, mimetype='application/pdf',
                    headers={'Content-Disposition': 'attachment; filename="Atlas-Excellence-Standards.pdf"'})


@app.route('/api/resources', methods=['GET'])
@login_required
def list_resources():
    """Everyone signed in can see the resource library."""
    items = []
    for r in resource_service.get_all():
        items.append({
            'id': r.get('id'),
            'title': r.get('title'),
            'description': r.get('description'),
            'kind': r.get('kind'),
            'url': r.get('url') if r.get('kind') == 'link' else None,
            'filename': r.get('filename') if r.get('kind') == 'file' else None,
            'created_at': r.get('created_at'),
        })
    return jsonify({'status': 'success', 'resources': items, 'is_admin': is_admin()}), 200


@app.route('/api/resources', methods=['POST'])
@login_required
def add_resource():
    """Admin-only: add a resource (an uploaded file or an external link)."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403

    title = InputSanitizer.sanitize_string(request.form.get('title', ''), max_length=120).strip()
    description = InputSanitizer.sanitize_string(request.form.get('description', ''), max_length=400).strip()
    kind = (request.form.get('kind') or '').strip().lower()
    if not title:
        return jsonify({'status': 'error', 'message': 'Title is required'}), 400

    if kind == 'link':
        url = (request.form.get('url') or '').strip()
        if not (url.startswith('http://') or url.startswith('https://')):
            return jsonify({'status': 'error', 'message': 'Enter a valid http(s) link'}), 400
        rec = resource_service.add(title, description, 'link', url=url, created_by=session.get('user'))
    elif kind == 'file':
        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'status': 'error', 'message': 'Choose a file to upload'}), 400
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if ext not in RESOURCE_ALLOWED_EXT:
            return jsonify({'status': 'error', 'message': f'Unsupported file type .{ext}'}), 400
        try:
            rel, stored = file_upload_handler.save_resource(file)
        except Exception as e:
            app.logger.error(f'Resource upload failed: {e}')
            return jsonify({'status': 'error', 'message': 'Could not save the file'}), 500
        rec = resource_service.add(title, description, 'file', file_path=rel,
                                   filename=secure_filename(file.filename),
                                   content_type=file.mimetype, created_by=session.get('user'))
    else:
        return jsonify({'status': 'error', 'message': 'Invalid resource type'}), 400

    return jsonify({'status': 'success', 'id': rec['id']}), 201


@app.route('/api/resources/<resource_id>', methods=['DELETE'])
@login_required
def delete_resource(resource_id):
    """Admin-only: remove a resource (and its file)."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    rec = resource_service.delete(resource_id)
    if not rec:
        return jsonify({'status': 'error', 'message': 'Resource not found'}), 404
    if rec.get('kind') == 'file' and rec.get('file_path'):
        file_upload_handler.delete_file(rec['file_path'])
    return jsonify({'status': 'success'}), 200


@app.route('/api/resources/<resource_id>/attach', methods=['POST'])
@login_required
def attach_resource(resource_id):
    """Admin-only: attach a file or link to an existing (e.g. pending) resource."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    existing = resource_service.get(resource_id)
    if not existing:
        return jsonify({'status': 'error', 'message': 'Resource not found'}), 404

    kind = (request.form.get('kind') or '').strip().lower()
    if kind == 'link':
        url = (request.form.get('url') or '').strip()
        if not (url.startswith('http://') or url.startswith('https://')):
            return jsonify({'status': 'error', 'message': 'Enter a valid http(s) link'}), 400
        resource_service.attach(resource_id, 'link', url=url)
    elif kind == 'file':
        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'status': 'error', 'message': 'Choose a file to upload'}), 400
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if ext not in RESOURCE_ALLOWED_EXT:
            return jsonify({'status': 'error', 'message': f'Unsupported file type .{ext}'}), 400
        try:
            rel, stored = file_upload_handler.save_resource(file)
        except Exception as e:
            app.logger.error(f'Resource attach upload failed: {e}')
            return jsonify({'status': 'error', 'message': 'Could not save the file'}), 500
        # remove the previous file if this resource already had one
        if existing.get('kind') == 'file' and existing.get('file_path'):
            file_upload_handler.delete_file(existing['file_path'])
        resource_service.attach(resource_id, 'file', file_path=rel,
                                filename=secure_filename(file.filename),
                                content_type=file.mimetype)
    else:
        return jsonify({'status': 'error', 'message': 'Invalid resource type'}), 400

    return jsonify({'status': 'success'}), 200


def _normalize_photo_path(value):
    """Reduce whatever the page had to hand down to a stored path.

    It is not always the stored path. Several views keep photo_url — the signed
    S3 link — in the same field, because that is what the <img> tag needs; the
    download button then sent a whole "https://...?X-Amz-Signature=..." where a
    key belonged, and every download came back as a 404. The key is inside that
    link, so take it here rather than making eleven call sites carry a second
    field and trusting that none is ever missed.
    """
    rel = (value or '').strip()
    if rel[:7] == 'http://' or rel[:8] == 'https://':
        from urllib.parse import urlparse, unquote
        rel = unquote(urlparse(rel).path or '')     # drops the signature query
    rel = rel.lstrip('/')

    # Local mode serves from /static/uploads/, S3 keys are uploads/<...>, and a
    # path-style URL puts the bucket first. All of them agree from the last
    # "uploads/" onward.
    marker = rel.rfind('uploads/')
    if marker != -1:
        rel = rel[marker + len('uploads/'):]
    return rel


def _photo_context(relative_path):
    """Work out which community a stored photo belongs to, and its story.

    Everything needed is already in the path: uploads are written as
    "<safe community>/<username>_<YYYYMMDD-HHMMSS>_<hex>.<ext>". Reading it
    back beats threading community, date and author through ten call sites in
    the template — and it cannot drift out of step with the stored file.

    Returns (community or None, taken_on or None, username or None).
    """
    rel = _normalize_photo_path(relative_path)
    if '/' not in rel:
        return None, None, None
    folder, filename = rel.split('/', 1)

    # secure_filename() flattened the real name on the way in ("Foo, TX" ->
    # "Foo_TX"), so match by putting the roster through the same funnel rather
    # than trying to reverse it.
    community = next((c for c in all_communities()
                      if secure_filename(c) == folder), None)

    stem = filename.rsplit('.', 1)[0]
    parts = stem.split('_')
    username, taken = None, None
    if len(parts) >= 3 and re.fullmatch(r'\d{8}-\d{6}', parts[-2]):
        username = '_'.join(parts[:-2])
        try:
            taken = datetime.strptime(parts[-2], '%Y%m%d-%H%M%S')
        except ValueError:
            taken = None
    elif len(parts) >= 2 and parts[-1].isdigit():
        # The older "<user>_<unix seconds>" names, from before the collision fix.
        username = '_'.join(parts[:-1])
        try:
            taken = datetime.utcfromtimestamp(int(parts[-1]))
        except (ValueError, OverflowError):
            taken = None
    return community, taken, username


@app.route('/api/photo/download')
@login_required
def download_photo():
    """Hand back a photo with its context printed underneath it.

    A photo pasted into an email loses everything that made it mean something —
    which community, which visit, who was standing there. The strip is added to
    the copy being sent; the stored original is never rewritten.
    """
    rel = _normalize_photo_path(request.args.get('path'))
    if not rel:
        return jsonify({'status': 'error', 'message': 'No photo'}), 400

    community, taken, username = _photo_context(rel)

    # The path arrives from the browser, so trusting it would let any signed-in
    # account read any community's photos by naming one. An unknown folder is
    # refused outright rather than served without a caption.
    if not community:
        return jsonify({'status': 'error', 'message': 'Photo not found'}), 404
    if community not in visible_communities():
        return jsonify({'status': 'error', 'message': 'Not your community'}), 403

    try:
        raw = file_upload_handler.read_bytes(rel)
    except Exception as e:
        app.logger.error(f'Could not read photo {rel}: {e}')
        return jsonify({'status': 'error', 'message': 'Photo not found'}), 404

    # Filenames are stamped in UTC by the upload handler; showing that clock
    # raw would push an evening visit onto the next day.
    when = fmt_local(taken, '%b %d, %Y') if taken else ''
    who = _display_name_for(username) if username else ''
    secondary = ' · '.join(p for p in (when, who) if p)

    import io
    from services.photo_stamp import stamp
    out, ext = stamp(raw, community, secondary)

    ext = ext or (rel.rsplit('.', 1)[-1].lower() if '.' in rel else 'jpg')
    mime = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png',
            'gif': 'image/gif', 'webp': 'image/webp'}.get(ext, 'image/jpeg')
    slug = community_slug(community) or 'photo'
    datepart = (taken.strftime('%Y-%m-%d') if taken else 'photo')
    name = f'{slug}-{datepart}.{ext}'
    return send_file(io.BytesIO(out), mimetype=mime,
                     as_attachment=True, download_name=name)


@app.route('/api/resources/<resource_id>/download')
@login_required
def download_resource(resource_id):
    """Stream/redirect to a resource for any signed-in user."""
    rec = resource_service.get(resource_id)
    if not rec:
        return jsonify({'status': 'error', 'message': 'Resource not found'}), 404
    if rec.get('kind') == 'link':
        return redirect(rec.get('url') or '/')
    rel = rec.get('file_path')
    if not rel:
        return jsonify({'status': 'error', 'message': 'No file'}), 404
    if file_upload_handler.use_s3:
        signed = file_upload_handler.generate_presigned_url(rel, download_name=rec.get('filename'))
        if not signed:
            return jsonify({'status': 'error', 'message': 'Could not generate download'}), 500
        return redirect(signed)
    # local: rel is like "resources/<stored>"
    directory = os.path.join(UPLOAD_FOLDER, os.path.dirname(rel))
    return send_from_directory(directory, os.path.basename(rel),
                               as_attachment=True, download_name=rec.get('filename'))


@app.route('/api/settings/email', methods=['GET'])
@login_required
def get_email_settings():
    """Admin-only: read subscribers + admin-notify list + the regions to pick from."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    s = settings_service.get_email_settings()
    regions = [{'id': r.get('id'), 'name': r.get('name')}
               for r in region_service.get_all_regions() if r.get('id') != 'unassigned']
    return jsonify({'status': 'success', 'subscribers': s['subscribers'],
                    'admin_notify': s['admin_notify'], 'regions': regions,
                    'inspectors': leadership_names(),
                    'clinical': s['clinical'], 'ops': s['ops'], 'sales': s['sales'],
                    'email_enabled': email_service.enabled}), 200


@app.route('/api/settings/email', methods=['POST'])
@login_required
def save_email_settings():
    """Admin-only: update subscribers + admin-notify list."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    data = request.get_json(silent=True) or {}
    # Subscribers are managed individually (add / update / remove one at a time),
    # so this endpoint only rewrites them when the caller explicitly sends the
    # key. That way saving the other fields can never drop someone silently.
    subs = None
    if 'subscribers' in data:
        valid_ids = {r.get('id') for r in region_service.get_all_regions() if r.get('id') != 'unassigned'}
        valid_names = set(leadership_names())
        subs = []
        for s in (data.get('subscribers') or []):
            if not isinstance(s, dict):
                continue
            regions = [rid for rid in (s.get('regions') or []) if rid in valid_ids]
            inspectors = [n for n in (s.get('inspectors') or []) if n in valid_names]
            subs.append({'email': s.get('email', ''), 'name': s.get('name', ''),
                         'regions': regions, 'inspectors': inspectors})
    saved = settings_service.set_email_settings(
        subscribers=subs, admin_notify=data.get('admin_notify', ''),
        clinical=data.get('clinical', ''), ops=data.get('ops', ''),
        sales=data.get('sales', ''))
    regions = [{'id': r.get('id'), 'name': r.get('name')}
               for r in region_service.get_all_regions() if r.get('id') != 'unassigned']
    return jsonify({'status': 'success', 'regions': regions, 'inspectors': leadership_names(), **saved}), 200


@app.route('/api/settings/visit-cadence', methods=['POST'])
@login_required
def save_visit_cadence():
    """Admin-only: how many days a community may go between visits.

    Nothing is enforced by this — it only decides which communities are shown
    as falling behind, on the dashboard and on the community cards."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    data = request.get_json(silent=True) or {}
    days = settings_service.set_visit_cadence_days(data.get('days'))
    activity_service.log(session.get('user'), 'settings_updated',
                         f'Set the visit target to every {days} days')
    return jsonify({'status': 'success', 'visit_cadence_days': days}), 200


@app.route('/api/view-as', methods=['POST'])
@login_required
def start_view_as():
    """Look at the app as one Executive Director sees it.

    Takes the account rather than a community name, so the preview inherits
    every community that person covers — including the second one when they
    are standing in for a neighbour."""
    if not real_is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    data = request.get_json(silent=True) or {}

    username = InputSanitizer.sanitize_string(data.get('username', ''), max_length=80)
    if username:
        account = user_service.get(username)
        if not account or account.get('role') != 'staff':
            return jsonify({'status': 'error',
                            'message': 'That is not an Executive Director account'}), 400
        communities = [c for c in account_communities(account) if c]
        label = account.get('display_name') or username
    else:
        # Straight to a community, for when there is no account yet.
        one = InputSanitizer.sanitize_community_name(data.get('community', ''))
        communities = [one] if one else []
        label = one

    known = set(all_communities())
    communities = [c for c in communities if c in known]
    if not communities:
        return jsonify({'status': 'error',
                        'message': 'That account has no community assigned'}), 400

    session['view_as'] = {'communities': communities, 'label': label}
    session.modified = True
    activity_service.log(session.get('user'), 'view_as_started',
                         f"Started previewing as {label} ({', '.join(communities)})")
    return jsonify({'status': 'success', 'label': label,
                    'communities': communities}), 200


@app.route('/api/view-as/stop', methods=['POST'])
@login_required
def stop_view_as():
    """Leave the preview. Deliberately does not ask is_admin(), which answers
    False while a preview is running — that check would lock the door from the
    inside."""
    was = session.pop('view_as', None)
    session.modified = True
    if was:
        activity_service.log(session.get('user'), 'view_as_stopped',
                             f"Stopped previewing as {was.get('label', '')}")
    return jsonify({'status': 'success'}), 200

@app.route('/api/user-info')
@login_required
def get_user_info():
    """
    Get current user's information
    """
    username = session.get('user')
    role = current_role()
    region_id = session.get('region_id')
    region_name = None
    if is_leadership(role):
        communities = regional_communities()
        region_name = next((r.get('name') for r in region_service.get_all_regions()
                            if r.get('id') == region_id), None)
    elif session_communities():
        communities = session_communities()
    else:
        communities = []
    return jsonify({
        'username': username,
        'display_name': profile_service.get_display_name(username) or session.get('display_name') or '',
        'photo': profile_service.get_photo(username),
        'community': session.get('community'),
        'role': role,
        # Unlocks the admin UI: true for Administrators AND for people granted
        # admin privileges on top of their own role.
        'is_admin': is_admin(),
        'is_native_admin': is_native_admin(),
        'admin_extra': bool(session.get('admin_extra')),
        'can_inspect': not is_native_admin(),
        # What this account may do. The dashboard hides the "Start a visit" and
        # "Mark as addressed" controls on these, so they have to travel here —
        # they were added to the profile endpoint by mistake, which left every
        # role looking like a community account.
        'can_run_visits': can_run_visits(),
        'can_verify_fixes': can_verify_fixes(),
        'region_id': region_id,
        'region_name': region_name,
        # How often a community is meant to be visited. The community cards use
        # it to flag the ones falling behind, so it has to be here rather than
        # only in the attention payload — that one is dashboard-only.
        'visit_cadence_days': settings_service.get_visit_cadence_days(),
        # Set only while an administrator is previewing as a community. The
        # banner and the way out are driven from this.
        'view_as': (viewing_as() or {}).get('label'),
        'view_as_communities': (viewing_as() or {}).get('communities') or [],
        'communities': communities
    }), 200


# ==================== SURVEY TYPE API ====================

@app.route('/api/survey-types', methods=['GET'])
@login_required
def get_survey_types():
    """
    Get all active survey types
    
    Returns:
        200: JSON with status and survey_types array
        500: Internal server error
        
    Requirements: Task 4.1
    """
    try:
        survey_types = survey_type_service.get_all_survey_types()
        # How many standards each type would actually put on the form. A type
        # with none is unusable: whoever picks it gets "No questions available"
        # after they have already chosen a community — in practice, standing in
        # a building. The pickers use this to rule those out up front.
        coverage = survey_type_coverage()
        survey_types = [{**st, 'standards': coverage.get(st.get('id'), 0)}
                        for st in survey_types]

        return jsonify({
            'status': 'success',
            'survey_types': survey_types
        }), 200
        
    except Exception as e:
        app.logger.exception('Error retrieving survey types')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while retrieving survey types'
        }), 500


@app.route('/api/survey-types', methods=['POST'])
@login_required
def create_survey_type():
    """Admin-only: create a new survey type (a question group / checklist)."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    data = request.get_json(silent=True) or {}
    name = InputSanitizer.sanitize_string(data.get('name', ''), max_length=80).strip()
    if not name:
        return jsonify({'status': 'error', 'message': 'Name is required'}), 400
    description = InputSanitizer.sanitize_string(data.get('description', ''), max_length=300)
    icon = InputSanitizer.sanitize_string(data.get('icon', ''), max_length=40) or 'fa-clipboard-list'
    color = InputSanitizer.sanitize_string(data.get('color', ''), max_length=20) or '#1f6fe5'
    try:
        st = survey_type_service.create_survey_type(name, description, icon, color)
    except Exception as e:
        app.logger.exception('Error creating survey type')
        return jsonify({'status': 'error', 'message': 'Could not create survey type'}), 500
    return jsonify({'status': 'success', 'survey_type': st}), 201


@app.route('/api/survey-types/<survey_type_id>', methods=['PUT'])
@login_required
def update_survey_type(survey_type_id):
    """Admin-only: edit a survey type's name/description/icon/color."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    data = request.get_json(silent=True) or {}
    st = survey_type_service.update_survey_type(
        survey_type_id,
        name=InputSanitizer.sanitize_string(data.get('name', ''), max_length=80) if 'name' in data else None,
        description=InputSanitizer.sanitize_string(data.get('description', ''), max_length=300) if 'description' in data else None,
        icon=InputSanitizer.sanitize_string(data.get('icon', ''), max_length=40) if 'icon' in data else None,
        color=InputSanitizer.sanitize_string(data.get('color', ''), max_length=20) if 'color' in data else None,
    )
    if not st:
        return jsonify({'status': 'error', 'message': 'Survey type not found'}), 404
    return jsonify({'status': 'success', 'survey_type': st}), 200


@app.route('/api/survey-types/<survey_type_id>', methods=['DELETE'])
@login_required
def delete_survey_type(survey_type_id):
    """Admin-only: remove a survey type."""
    if not is_admin():
        return jsonify({'status': 'error', 'message': 'Admins only'}), 403
    if survey_type_service.delete_survey_type(survey_type_id):
        return jsonify({'status': 'success'}), 200
    return jsonify({'status': 'error', 'message': 'Survey type not found'}), 404


@app.route('/api/select-survey-type', methods=['POST'])
@login_required
def api_select_survey_type():
    """
    Store selected survey type in session
    
    Expects JSON with survey_type_id
    
    Returns:
        200: Success with survey type info
        400: Invalid survey type ID or missing field
        500: Internal server error
        
    Requirements: Task 4.2
    """
    try:
        # Handle JSON parsing errors
        data = request.get_json(silent=True)
        
        if data is None:
            return jsonify({
                'status': 'error',
                'message': 'Invalid JSON format or Content-Type must be application/json'
            }), 400
        
        # Validate JSON structure
        if not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({
                'status': 'error',
                'message': 'Request body must be a JSON object'
            }), 400
        
        # Get and sanitize survey_type_id
        survey_type_id = data.get('survey_type_id', '')
        survey_type_id = InputSanitizer.sanitize_string(survey_type_id, max_length=50)
        
        if not survey_type_id:
            return jsonify({
                'status': 'error',
                'message': 'survey_type_id is required'
            }), 400
        
        # Validate survey type exists
        if not survey_type_service.validate_survey_type(survey_type_id):
            return jsonify({
                'status': 'error',
                'message': f'Invalid survey type: {survey_type_id}'
            }), 400
        
        # Get survey type details
        survey_type = survey_type_service.get_survey_type_by_id(survey_type_id)
        
        if not survey_type:
            return jsonify({
                'status': 'error',
                'message': f'Survey type not found: {survey_type_id}'
            }), 400
        
        # Store in session
        session['survey_type_id'] = survey_type_id
        session['survey_type_name'] = survey_type['name']
        session.modified = True
        
        return jsonify({
            'status': 'success',
            'message': 'Survey type selected successfully',
            'survey_type': survey_type
        }), 200
        
    except Exception as e:
        app.logger.exception('Error selecting survey type')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while selecting survey type'
        }), 500


@app.route('/questions/manage')
@require_admin
def question_manager_ui():
    """
    Render the Question Manager UI (question_manager.html)
    This page is designed for admin users to create, edit, and manage inspection questions
    Requires admin authentication - non-admin users are redirected to inspection form
    
    Requirements: 1.1, 7.1
    """
    # The community checkboxes come from the region rosters, not the fixed list
    # below. The two had drifted apart on two names, so a standard could be
    # assigned to "The Overlook at Suwanee, Suwanee" while a visit for "The
    # Overlook at Suwanee" found nothing — the checkbox looked ticked and the
    # regional got "No questions available" standing in the building.
    return render_template('question_manager.html',
                         username=session.get('user'),
                         # Only admins reach this page at all, so the sidebar
                         # here always shows the full menu.
                         nav_admin=True,
                         active='standards',
                         communities=all_communities() or ALL_COMMUNITIES)


@app.route('/api/questions', methods=['GET'])
@login_required
def get_questions():
    """
    Get active questions with community and survey type filtering
    
    Query Parameters:
        community (optional): Filter questions by community name
        survey_type (optional): Filter questions by survey type ID
        
    Behavior:
        - For staff users: Automatically filters by their assigned community
        - For admin users: Returns all active questions, or filters by community parameter if provided
        - If survey_type is provided, filters questions by that survey type
        
    Returns:
        200: JSON with status and questions array
        400: Invalid survey type
        500: Internal server error
    """
    try:
        # Sanitize community filter from query parameter
        community_filter = request.args.get('community')
        if community_filter:
            community_filter = InputSanitizer.sanitize_community_name(community_filter)
        
        # Get survey type filter from query parameter
        survey_type_filter = request.args.get('survey_type')
        if survey_type_filter:
            survey_type_filter = InputSanitizer.sanitize_string(survey_type_filter, max_length=50)
            # Validate survey type
            if not survey_type_service.validate_survey_type(survey_type_filter):
                return jsonify({
                    'status': 'error',
                    'message': f'Invalid survey type: {survey_type_filter}'
                }), 400
        
        # Determine which questions to return. Standards are configuration, so
        # anyone with admin privileges (including the accessory) sees them all —
        # otherwise they couldn't manage the very standards they're allowed to edit.
        role = current_role()
        if is_admin():
            if community_filter:
                questions = question_manager.get_questions_for_community(community_filter)
            else:
                questions = question_manager.get_all_active_questions()
        elif is_leadership(role):
            # Regionals must pick a community within their region
            allowed = regional_communities()
            if community_filter and community_filter in allowed:
                questions = question_manager.get_questions_for_community(community_filter)
            else:
                questions = []
        else:
            # Community account — the standards for whichever site it covers.
            comms = session_communities()
            questions = question_manager.get_questions_for_community(comms[0]) if comms else []
        
        # Apply survey type filter if provided
        if survey_type_filter:
            questions = question_filter_service.filter_by_survey_type(questions, survey_type_filter)
        
        return jsonify({
            'status': 'success',
            'questions': questions
        }), 200
        
    except Exception as e:
        # Log the error for debugging
        app.logger.exception('Error retrieving questions')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while retrieving questions'
        }), 500


# ==================== QUESTION MANAGEMENT API ====================

@app.route('/api/questions', methods=['POST'])
@require_admin
def create_question():
    """
    API endpoint to create a new inspection question
    Requires admin authentication
    Expects JSON with text, photo_required, and communities array
    
    Error Handling:
    - 400: JSON parsing errors, validation errors, missing fields
    - 500: Internal server errors (file system, etc.)
    """
    try:
        # Handle JSON parsing errors
        data = request.get_json(silent=True)
        
        if data is None:
            return jsonify({
                'status': 'error',
                'message': 'Invalid JSON format or Content-Type must be application/json'
            }), 400
        
        # Validate JSON structure
        if not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({
                'status': 'error',
                'message': 'Request body must be a JSON object'
            }), 400
        
        # Sanitize input data
        sanitized_data = InputSanitizer.sanitize_question_data(data)
        
        # Extract fields
        text = sanitized_data.get('text', '')
        interpretive_guideline = sanitized_data.get('interpretive_guideline', '')
        photo_required = sanitized_data.get('photo_required', False)
        communities = sanitized_data.get('communities', [])
        survey_types = sanitized_data.get('survey_types', [])

        # Validate text is non-empty
        if not text or not text.strip():
            return jsonify({
                'status': 'error',
                'message': 'Question text cannot be empty'
            }), 400

        # Validate communities array is non-empty
        if not communities or len(communities) == 0:
            return jsonify({
                'status': 'error',
                'message': 'At least one community must be selected'
            }), 400

        pass_criteria = sanitized_data.get('pass_criteria', [])

        # Create question using QuestionManager
        question = question_manager.create_question(text, photo_required, communities, survey_types, interpretive_guideline, pass_criteria)

        activity_service.log(session.get('user'), 'question_created', text)

        return jsonify({
            'status': 'success',
            'question': question
        }), 201
        
    except ValueError as e:
        # Validation errors from QuestionManager
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 400
    except IOError as e:
        # File system errors
        app.logger.error(f'File system error creating question: {str(e)}')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error: Failed to save question'
        }), 500
    except Exception as e:
        # Unexpected errors
        app.logger.exception('Unexpected error creating question')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while creating question'
        }), 500


@app.route('/api/questions/<question_id>', methods=['PUT'])
@require_admin
def update_question(question_id):
    """
    API endpoint to update an existing inspection question
    Requires admin authentication
    Expects JSON with text, photo_required, and communities array
    
    Error Handling:
    - 400: JSON parsing errors, validation errors, missing fields
    - 404: Question not found
    - 500: Internal server errors (file system, etc.)
    """
    try:
        # Sanitize question_id
        question_id = InputSanitizer.sanitize_string(question_id, max_length=100)
        
        # Handle JSON parsing errors
        data = request.get_json(silent=True)
        
        if data is None:
            return jsonify({
                'status': 'error',
                'message': 'Invalid JSON format or Content-Type must be application/json'
            }), 400
        
        # Validate JSON structure
        if not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({
                'status': 'error',
                'message': 'Request body must be a JSON object'
            }), 400
        
        # Sanitize input data
        sanitized_data = InputSanitizer.sanitize_question_data(data)
        
        # Extract fields
        text = sanitized_data.get('text', '')
        interpretive_guideline = sanitized_data.get('interpretive_guideline', '')
        photo_required = sanitized_data.get('photo_required', False)
        communities = sanitized_data.get('communities', [])
        survey_types = sanitized_data.get('survey_types', [])

        # Validate text is non-empty
        if not text or not text.strip():
            return jsonify({
                'status': 'error',
                'message': 'Question text cannot be empty'
            }), 400

        # Validate communities array is non-empty
        if not communities or len(communities) == 0:
            return jsonify({
                'status': 'error',
                'message': 'At least one community must be selected'
            }), 400

        pass_criteria = sanitized_data.get('pass_criteria', [])

        # Update question using QuestionManager
        question = question_manager.update_question(question_id, text, photo_required, communities, survey_types, interpretive_guideline, pass_criteria)
        
        # Check if question was found
        if question is None:
            return jsonify({
                'status': 'error',
                'message': 'Question not found'
            }), 404

        activity_service.log(session.get('user'), 'question_updated', text)

        return jsonify({
            'status': 'success',
            'question': question
        }), 200
        
    except ValueError as e:
        # Validation errors from QuestionManager
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 400
    except IOError as e:
        # File system errors
        app.logger.error(f'File system error updating question: {str(e)}')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error: Failed to save question'
        }), 500
    except Exception as e:
        # Unexpected errors
        app.logger.exception('Unexpected error updating question')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while updating question'
        }), 500


@app.route('/api/questions/<question_id>', methods=['DELETE'])
@require_admin
def delete_question(question_id):
    """
    API endpoint to delete an inspection question (soft delete)
    Requires admin authentication
    Performs soft delete by setting is_active to False
    
    Error Handling:
    - 404: Question not found
    - 500: Internal server errors (file system, etc.)
    """
    try:
        # Sanitize question_id
        question_id = InputSanitizer.sanitize_string(question_id, max_length=100)
        
        # Delete question using QuestionManager (soft delete)
        success = question_manager.delete_question(question_id)
        
        # Check if question was found
        if not success:
            return jsonify({
                'status': 'error',
                'message': 'Question not found'
            }), 404

        activity_service.log(session.get('user'), 'question_deleted', f'Question {question_id}')

        return jsonify({
            'status': 'success',
            'message': 'Question deleted successfully'
        }), 200
        
    except IOError as e:
        # File system errors
        app.logger.error(f'File system error deleting question: {str(e)}')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error: Failed to save changes'
        }), 500
    except Exception as e:
        # Unexpected errors
        app.logger.exception('Unexpected error deleting question')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while deleting question'
        }), 500


@app.route('/api/questions/bulk-delete', methods=['POST'])
@require_admin
def bulk_delete_questions():
    """
    API endpoint to soft-delete multiple questions in one request
    Requires admin authentication
    Expects JSON: { "question_ids": ["id1", "id2", ...] }

    Error Handling:
    - 400: Invalid JSON, missing/empty question_ids
    - 500: Internal server errors
    """
    try:
        data = request.get_json(silent=True)

        if data is None or not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({
                'status': 'error',
                'message': 'Request body must be a JSON object'
            }), 400

        question_ids = data.get('question_ids')
        if not isinstance(question_ids, list) or len(question_ids) == 0:
            return jsonify({
                'status': 'error',
                'message': 'question_ids must be a non-empty array'
            }), 400

        deleted = 0
        not_found = []
        for raw_id in question_ids:
            qid = InputSanitizer.sanitize_string(str(raw_id), max_length=100)
            if not qid:
                continue
            if question_manager.delete_question(qid):
                deleted += 1
            else:
                not_found.append(qid)

        if deleted:
            activity_service.log(session.get('user'), 'question_deleted', f'Deleted {deleted} question(s)')

        return jsonify({
            'status': 'success',
            'message': f'Deleted {deleted} question(s)',
            'deleted': deleted,
            'not_found': not_found
        }), 200

    except IOError as e:
        app.logger.error(f'File system error during bulk delete: {str(e)}')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error: Failed to save changes'
        }), 500
    except Exception as e:
        app.logger.exception('Unexpected error during bulk delete')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while deleting questions'
        }), 500


# ==================== REGIONS API ====================

@app.route('/api/regions', methods=['GET'])
@login_required
def get_regions():
    """
    Get the regional structure: each region with its leadership roster and
    assigned communities. Used by the dashboard Regions view.
    """
    try:
        # Enrich a copy of each leader with their uploaded photo (if any),
        # without mutating the stored region data.
        enriched = []
        for region in region_service.get_all_regions():
            r = dict(region)
            r['leadership'] = [
                {**leader, 'photo': profile_service.get_leader_photo(region.get('id', ''), leader.get('name', ''))}
                for leader in region.get('leadership', [])
            ]
            enriched.append(r)

        # Who runs each community day to day. Sent alongside the regions so the
        # view can show it — and, more usefully, show where nobody is assigned.
        # A community without an account stops receiving its findings email, and
        # that failure is silent, so it needs to be visible somewhere.
        directors = {}
        for u in user_service.get_all():
            if u.get('role') != 'staff':
                continue
            for comm in account_communities(u):
                directors.setdefault(comm, []).append({
                    'username': u.get('username'),
                    'name': profile_service.get_display_name(u.get('username'))
                            or u.get('display_name') or u.get('username'),
                    'email': (u.get('email') or '').strip(),
                })

        return jsonify({
            'status': 'success',
            'regions': enriched,
            'directors': directors
        }), 200
    except Exception as e:
        app.logger.exception('Error retrieving regions')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while retrieving regions'
        }), 500


@app.route('/api/regions/assign', methods=['POST'])
@require_admin
def assign_region_community():
    """
    Assign a community to a region (move existing, add new, or restore from
    Unassigned). The community ends up belonging only to the target region.
    Expects JSON: { "community": "...", "region_id": "..." }
    """
    try:
        data = request.get_json(silent=True)
        if data is None or not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({'status': 'error', 'message': 'Request body must be a JSON object'}), 400

        community = InputSanitizer.sanitize_community_name(data.get('community', ''))
        region_id = InputSanitizer.sanitize_string(data.get('region_id', ''), max_length=50)

        if not community:
            return jsonify({'status': 'error', 'message': 'community is required'}), 400
        if not region_id:
            return jsonify({'status': 'error', 'message': 'region_id is required'}), 400
        if region_id == CORPORATE_ID:
            return jsonify({'status': 'error',
                            'message': 'Corporate is a company-wide group, not a region — '
                                       'it does not own communities.'}), 400

        if not region_service.assign_community(community, region_id):
            return jsonify({'status': 'error', 'message': f'Unknown region: {region_id}'}), 400

        activity_service.log(session.get('user'), 'region_assigned', f'Assigned {community} to {region_id}')

        return jsonify({'status': 'success', 'regions': region_service.get_all_regions()}), 200
    except IOError as e:
        app.logger.error(f'File system error assigning community: {str(e)}')
        return jsonify({'status': 'error', 'message': 'Internal server error: Failed to save changes'}), 500
    except Exception as e:
        app.logger.exception('Unexpected error assigning community')
        return jsonify({'status': 'error', 'message': 'Internal server error while assigning community'}), 500


@app.route('/api/regions/rename', methods=['POST'])
@require_admin
def rename_region():
    """Rename a region's display name (admin only). The region id is unchanged,
    so user scoping is unaffected. Expects JSON: { region_id, name }."""
    try:
        data = request.get_json(silent=True)
        if data is None or not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({'status': 'error', 'message': 'Request body must be a JSON object'}), 400

        region_id = InputSanitizer.sanitize_string(data.get('region_id', ''), max_length=50)
        new_name = InputSanitizer.sanitize_string(data.get('name', ''), max_length=80)

        if not region_id or not new_name:
            return jsonify({'status': 'error', 'message': 'region_id and name are required'}), 400
        if region_id == 'unassigned':
            return jsonify({'status': 'error', 'message': 'The Unassigned group cannot be renamed'}), 400

        if not region_service.rename_region(region_id, new_name):
            return jsonify({'status': 'error', 'message': 'Could not rename region (unknown id or unchanged name)'}), 400

        activity_service.log(session.get('user'), 'region_renamed', f'Renamed region {region_id} to "{new_name}"')
        return jsonify({'status': 'success', 'regions': region_service.get_all_regions()}), 200
    except Exception as e:
        app.logger.exception('Unexpected error renaming region')
        return jsonify({'status': 'error', 'message': 'Internal server error while renaming region'}), 500


@app.route('/api/regions/rename-community', methods=['POST'])
@require_admin
def rename_region_community():
    """
    Rename a community everywhere it's stored: the regional structure,
    the inspection questions, and historical submissions.
    Expects JSON: { "old_name": "...", "new_name": "..." }
    """
    try:
        data = request.get_json(silent=True)
        if data is None or not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({'status': 'error', 'message': 'Request body must be a JSON object'}), 400

        old_name = InputSanitizer.sanitize_community_name(data.get('old_name', ''))
        new_name = InputSanitizer.sanitize_community_name(data.get('new_name', ''))
        if not old_name or not new_name:
            return jsonify({'status': 'error', 'message': 'old_name and new_name are required'}), 400
        if old_name == new_name:
            return jsonify({'status': 'success', 'regions': region_service.get_all_regions()}), 200

        region_service.rename_community(old_name, new_name)
        try:
            question_manager.rename_community(old_name, new_name)
            inspection_service.rename_community(old_name, new_name)
            # Keep move-ins and the cover photo pointing at the renamed community
            movein_service.rename_community(old_name, new_name)
            # Items the community raised for itself travel with it too. Left
            # out, they would point at a name that no longer exists — the same
            # way a community once lost its standards and a regional drove to a
            # building she couldn't inspect.
            raised_item_service.rename_community(old_name, new_name)
            community_cover_service.rename(
                community_slug(old_name), community_slug(new_name), new_name)
        except Exception as e:
            app.logger.error(f'Partial error during community rename: {str(e)}')

        activity_service.log(session.get('user'), 'community_renamed',
                             f'Renamed "{old_name}" to "{new_name}"')

        return jsonify({'status': 'success', 'regions': region_service.get_all_regions()}), 200
    except IOError as e:
        app.logger.error(f'File system error renaming community: {str(e)}')
        return jsonify({'status': 'error', 'message': 'Internal server error: Failed to save changes'}), 500
    except Exception as e:
        app.logger.exception('Unexpected error renaming community')
        return jsonify({'status': 'error', 'message': 'Internal server error while renaming community'}), 500


@app.route('/api/regions/leader-photo', methods=['POST'])
@require_admin
def upload_leader_photo():
    """Upload a photo for a region leadership member (admin only)."""
    try:
        region_id = InputSanitizer.sanitize_string(request.form.get('region_id', ''), max_length=50)
        leader_name = InputSanitizer.sanitize_string(request.form.get('leader_name', ''), max_length=120)

        if not region_id or not leader_name:
            return jsonify({'status': 'error', 'message': 'region_id and leader_name are required'}), 400
        if 'photo' not in request.files:
            return jsonify({'status': 'error', 'message': 'No photo provided'}), 400

        file = request.files['photo']
        if not file or file.filename == '':
            return jsonify({'status': 'error', 'message': 'No photo provided'}), 400
        if not allowed_file(file.filename):
            return jsonify({'status': 'error', 'message': 'Invalid file type. Only images are allowed.'}), 400

        ext = secure_filename(file.filename).rsplit('.', 1)[1].lower()
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"leader_{secure_filename(region_id)}_{secure_filename(leader_name)}_{timestamp}.{ext}"
        try:
            file.save(os.path.join(AVATARS_FOLDER, filename))
        except IOError as e:
            app.logger.exception('Error saving leader photo')
            return jsonify({'status': 'error', 'message': 'Failed to save photo'}), 500

        relative_path = f"avatars/{filename}"
        profile_service.set_leader_photo(region_id, leader_name, relative_path)
        return jsonify({'status': 'success', 'photo': relative_path}), 200
    except Exception as e:
        app.logger.exception('Error uploading leader photo')
        return jsonify({'status': 'error', 'message': 'Internal server error while uploading photo'}), 500


@app.route('/api/regions/leader', methods=['POST'])
@require_admin
def manage_region_leader():
    """
    Add, update, or delete a leadership member in a region (admin only).
    Expects JSON:
      { "action": "add",    "region_id": "...", "name": "...", "role": "...", "email": "..." }
      { "action": "update", "region_id": "...", "index": N, "name": "...", "role": "...", "email": "..." }
      { "action": "delete", "region_id": "...", "index": N }
    """
    try:
        data = request.get_json(silent=True)
        if data is None or not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({'status': 'error', 'message': 'Request body must be a JSON object'}), 400

        action = InputSanitizer.sanitize_string(data.get('action', ''), max_length=20)

        # Move/reorder uses from_/to_ region ids rather than a single region_id
        if action == 'move':
            from_region = InputSanitizer.sanitize_string(data.get('from_region_id', ''), max_length=50)
            to_region = InputSanitizer.sanitize_string(data.get('to_region_id', ''), max_length=50)
            from_index = data.get('from_index')
            to_index = data.get('to_index')
            if not from_region or not to_region:
                return jsonify({'status': 'error', 'message': 'from_region_id and to_region_id are required'}), 400
            if not isinstance(from_index, int):
                return jsonify({'status': 'error', 'message': 'from_index is required'}), 400

            moved = region_service.move_leader(
                from_region, from_index, to_region,
                to_index if isinstance(to_index, int) else None
            )
            if not moved:
                return jsonify({'status': 'error', 'message': 'Could not move leader (bad region or index)'}), 400

            # Carry the leader's photo to the new region key
            name = moved.get('name', '')
            if from_region != to_region and name:
                photo = profile_service.get_leader_photo(from_region, name)
                if photo:
                    profile_service.set_leader_photo(to_region, name, photo)

            activity_service.log(session.get('user'), 'leader_moved',
                                 f'Moved {name} from {from_region} to {to_region}')
            return jsonify({'status': 'success'}), 200

        region_id = InputSanitizer.sanitize_string(data.get('region_id', ''), max_length=50)
        if not region_id:
            return jsonify({'status': 'error', 'message': 'region_id is required'}), 400

        if action == 'delete':
            index = data.get('index')
            if not isinstance(index, int):
                return jsonify({'status': 'error', 'message': 'index is required'}), 400
            if not region_service.remove_leader(region_id, index):
                return jsonify({'status': 'error', 'message': 'Could not delete leader (bad region or index)'}), 400
            activity_service.log(session.get('user'), 'leader_removed', f'Removed a leader from {region_id}')

        elif action in ('add', 'update'):
            name = InputSanitizer.sanitize_string(data.get('name', ''), max_length=120)
            role = InputSanitizer.sanitize_string(data.get('role', ''), max_length=40)
            email = InputSanitizer.sanitize_string(data.get('email', ''), max_length=120)
            if not name:
                return jsonify({'status': 'error', 'message': 'Leader name is required'}), 400

            if action == 'add':
                # Pin a permanent login now, so later name edits never move it.
                pinned = slugify_name(name)
                if pinned and pinned not in get_regional_accounts() and not username_taken(pinned):
                    pass          # free to use as-is
                elif pinned:
                    pinned = generate_unique_username(name)
                if not region_service.add_leader(region_id, name, role, email, username=pinned):
                    return jsonify({'status': 'error', 'message': f'Unknown region: {region_id}'}), 400
                activity_service.log(session.get('user'), 'leader_added', f'Added {name} ({role}) to {region_id}')
            else:
                index = data.get('index')
                if not isinstance(index, int):
                    return jsonify({'status': 'error', 'message': 'index is required'}), 400
                if not region_service.update_leader(region_id, index, name, role, email):
                    return jsonify({'status': 'error', 'message': 'Could not update leader (bad region or index)'}), 400
                activity_service.log(session.get('user'), 'leader_updated', f'Updated {name} ({role}) in {region_id}')
        else:
            return jsonify({'status': 'error', 'message': 'action must be add, update, delete, or move'}), 400

        return jsonify({'status': 'success'}), 200
    except IOError as e:
        app.logger.error(f'File system error managing leader: {str(e)}')
        return jsonify({'status': 'error', 'message': 'Internal server error: Failed to save changes'}), 500
    except Exception as e:
        app.logger.exception('Unexpected error managing leader')
        return jsonify({'status': 'error', 'message': 'Internal server error while managing leadership'}), 500


@app.route('/api/regions/remove-community', methods=['POST'])
@require_admin
def remove_region_community():
    """
    Remove a community from the regional structure entirely.
    Expects JSON: { "community": "..." }
    """
    try:
        data = request.get_json(silent=True)
        if data is None or not InputSanitizer.validate_json_structure(data, dict):
            return jsonify({'status': 'error', 'message': 'Request body must be a JSON object'}), 400

        community = InputSanitizer.sanitize_community_name(data.get('community', ''))
        if not community:
            return jsonify({'status': 'error', 'message': 'community is required'}), 400

        found = region_service.remove_community(community)
        if not found:
            return jsonify({'status': 'error', 'message': 'Community not found in any region'}), 404

        activity_service.log(session.get('user'), 'region_removed', f'Removed {community} from regions')

        return jsonify({'status': 'success', 'regions': region_service.get_all_regions()}), 200
    except IOError as e:
        app.logger.error(f'File system error removing community: {str(e)}')
        return jsonify({'status': 'error', 'message': 'Internal server error: Failed to save changes'}), 500
    except Exception as e:
        app.logger.exception('Unexpected error removing community')
        return jsonify({'status': 'error', 'message': 'Internal server error while removing community'}), 500


# ==================== INSPECTION SUBMISSION API ====================

@app.route('/api/inspections', methods=['POST'])
@login_required
def submit_inspection():
    """
    API endpoint for inspection submission
    Requires authentication using @login_required decorator
    Accepts multipart/form-data with responses array
    Validates each response has question_id and condition
    Handles optional photo uploads for each response
    Saves photos using FileUploadHandler
    Creates submission using InspectionService
    Returns 201 with submission data on success
    Returns 400 with error message on validation failure
    
    Error Handling:
    - 400: JSON parsing errors, validation errors, missing fields, invalid file types/sizes
    - 500: Internal server errors (file system, etc.)
    
    Requirements: 3.1, 3.3, 4.1, 4.2, 4.3, 4.4, 5.1, 5.7, 5.8
    """
    try:
        # Get user info from session
        username = session.get('user')
        survey_type_id = session.get('survey_type_id')
        role = current_role()

        # Sanitize user info
        username = InputSanitizer.sanitize_username(username)

        # Resolve the community being inspected, by role:
        #  - regional / corporate: a community they pick (within their scope)
        #  - admin: not allowed
        #  - community accounts: not allowed either — see below
        # Native Administrators don't submit visits; Corporate/Regional users
        # keep inspecting even when they hold admin privileges.
        if is_native_admin():
            return jsonify({'status': 'error', 'message': 'Admin users cannot submit visits'}), 400
        elif is_leadership(role):
            community = InputSanitizer.sanitize_community_name(request.form.get('community', ''))
            if not community or community not in regional_communities():
                return jsonify({'status': 'error', 'message': 'Select a valid community in your region'}), 400
        else:
            # A community cannot inspect itself. The score comes from the most
            # recent visit, so a self-run walkthrough marking everything Pass
            # would silently replace a regional's findings and wipe the open
            # items with it. Communities report progress by commenting; the
            # visit itself stays with the regional who performs it.
            return jsonify({
                'status': 'error',
                'message': 'Visits are carried out by regional and corporate staff. '
                           'To report progress on an item, add a comment to it.'
            }), 403
        
        # Validate survey type is selected
        if not survey_type_id:
            return jsonify({
                'status': 'error',
                'message': 'Survey type must be selected before submitting inspection'
            }), 400
        
        # Validate survey type is valid
        if not survey_type_service.validate_survey_type(survey_type_id):
            return jsonify({
                'status': 'error',
                'message': f'Invalid survey type in session: {survey_type_id}'
            }), 400
        
        # Parse responses from form data
        # Expected format: responses as JSON string in form data
        responses_json = request.form.get('responses')
        
        if not responses_json:
            return jsonify({
                'status': 'error',
                'message': 'No responses provided'
            }), 400
        
        # Parse JSON responses with error handling
        try:
            responses_data = json.loads(responses_json)
        except json.JSONDecodeError as e:
            return jsonify({
                'status': 'error',
                'message': f'Invalid JSON format for responses: {str(e)}'
            }), 400
        
        # Validate responses is a list
        if not InputSanitizer.validate_json_structure(responses_data, list):
            return jsonify({
                'status': 'error',
                'message': 'Responses must be an array'
            }), 400
        
        # Process each response
        processed_responses = []
        
        for idx, response in enumerate(responses_data):
            # Validate response is a dictionary
            if not isinstance(response, dict):
                return jsonify({
                    'status': 'error',
                    'message': f'Response {idx}: must be a JSON object'
                }), 400
            
            # Sanitize response data
            sanitized_response = InputSanitizer.sanitize_response_data(response)
            
            # Validate required fields: question_id and condition
            question_id = sanitized_response.get('question_id')
            condition = sanitized_response.get('condition')
            
            if not question_id:
                return jsonify({
                    'status': 'error',
                    'message': f'Response {idx}: question_id is required'
                }), 400
            
            if not condition:
                return jsonify({
                    'status': 'error',
                    'message': f'Response {idx}: condition is required'
                }), 400
            
            # Validate condition value (Pass/Fail)
            if condition not in ['Pass', 'Fail']:
                return jsonify({
                    'status': 'error',
                    'message': f'Response {idx}: condition must be "Pass" or "Fail"'
                }), 400
            
            # Get optional fields
            question_text = sanitized_response.get('question_text', '')
            description = sanitized_response.get('description', '')

            # A Fail must explain what was found — the follow-up depends on it.
            # The form enforces this too; this is the backstop.
            if condition == 'Fail' and not description.strip():
                label = question_text or question_id
                return jsonify({
                    'status': 'error',
                    'message': f'"{label}" is marked Fail — please add a comment describing what you found.'
                }), 400


            # Handle optional photo upload. Keyed by the standard, not by the
            # position in the list — those two stopped lining up whenever an
            # earlier standard had no photo, and evidence landed on the wrong
            # item. The positional name is still read so a page left open from
            # before this change still submits correctly.
            photo_path = None
            photo_field_name = f'photo_q_{question_id}'
            if photo_field_name not in request.files:
                photo_field_name = f'photo_{idx}'
            
            if photo_field_name in request.files:
                photo_file = request.files[photo_field_name]
                
                # Only process if file was actually uploaded
                if photo_file and photo_file.filename:
                    # Validate file using FileUploadHandler
                    is_valid, error_message = file_upload_handler.validate_file(photo_file)
                    
                    if not is_valid:
                        return jsonify({
                            'status': 'error',
                            'message': f'Response {idx}: {error_message}'
                        }), 400
                    
                    # Save photo using FileUploadHandler
                    try:
                        photo_path = file_upload_handler.save_file(photo_file, username, community)
                    except IOError as e:
                        app.logger.error(f'File system error saving photo: {str(e)}')
                        return jsonify({
                            'status': 'error',
                            'message': f'Response {idx}: Internal server error - Failed to save photo'
                        }), 500
                    except Exception as e:
                        app.logger.exception('Unexpected error saving photo')
                        return jsonify({
                            'status': 'error',
                            'message': f'Response {idx}: Failed to save photo'
                        }), 400
            
            # Optional routing of this item's comment to Clinical / Ops
            route_to = (response.get('route_to') or '').strip().lower()
            if route_to not in settings_service.ROUTES:
                route_to = None

            # Create response object
            response_obj = {
                'question_id': question_id,
                'question_text': question_text,
                'condition': condition,
                'description': description,
                'photo_path': photo_path,
                'route_to': route_to,
                'answered_at': datetime.now().isoformat()
            }

            processed_responses.append(response_obj)
        
        # Create submission using InspectionService
        try:
            # Ad-hoc action items raised at the end of the visit. They're
            # follow-up tasks (not standards), so they never affect the score.
            manual_items = []
            try:
                raw_items = request.form.get('action_items')
                if raw_items:
                    parsed = json.loads(raw_items)
                    if isinstance(parsed, list):
                        for it in parsed[:20]:
                            if not isinstance(it, dict):
                                continue
                            manual_items.append({
                                'text': InputSanitizer.sanitize_description(it.get('text', '')),
                                'assigned_to': InputSanitizer.sanitize_string(it.get('assigned_to', ''), max_length=80),
                                'priority': InputSanitizer.sanitize_string(it.get('priority', 'medium'), max_length=10),
                                'photo': '',
                            })
            except (json.JSONDecodeError, TypeError) as e:
                app.logger.warning(f'Ignoring malformed action_items: {e}')

            # How many standards the form offered. Sent by the browser, but not
            # taken on trust: it is checked against the survey's real question
            # count so a doctored value can't make a partial visit look whole.
            standards_total = None
            try:
                claimed = int(request.form.get('standards_total') or 0)
            except (TypeError, ValueError):
                claimed = 0
            if claimed > 0:
                actual = 0
                if survey_type_id:
                    try:
                        actual = len(question_filter_service.get_questions_for_survey(
                            community, survey_type_id) or [])
                    except Exception:
                        # A survey type that no longer resolves shouldn't stop a
                        # visit being filed; fall back to what the form said.
                        actual = 0
                standards_total = actual or claimed

            # Notes from the visit. Free text about the visit as a whole, with
            # an optional photo. Not a standard and not a task, so it never
            # reaches the score — the only thing it changes is what the
            # community reads when the report arrives.
            visit_notes = InputSanitizer.sanitize_description(
                request.form.get('visit_notes', ''))[:2000]
            notes_photo = None
            nf = request.files.get('visit_notes_photo')
            if nf and nf.filename:
                ok_file, why = file_upload_handler.validate_file(nf)
                if ok_file:
                    try:
                        notes_photo = file_upload_handler.save_file(nf, username, community)
                    except Exception:
                        # A photo attached to a note is never worth failing a
                        # whole visit for — the walk is already done.
                        app.logger.exception('Could not save the visit notes photo')
                else:
                    app.logger.warning(f'Visit notes photo rejected: {why}')

            submission = inspection_service.create_submission(
                username=username,
                community=community,
                responses=processed_responses,
                survey_type_id=survey_type_id,
                inspector_name=(session.get('display_name') or resolve_display_name(username)),
                action_items=manual_items,
                standards_total=standards_total,
                notes=visit_notes,
                notes_photo=notes_photo
            )
            
            # Clear survey type from session after successful submission
            session.pop('survey_type_id', None)
            session.pop('survey_type_name', None)
            session.modified = True

            # Audit log
            # The event type stays as-is (it's a stored key other code filters
            # on); only the sentence people read changes.
            activity_service.log(username, 'inspection_submitted',
                                 f'Submitted a visit for {community}',
                                 meta={'community': community})
            
        except ValueError as e:
            return jsonify({
                'status': 'error',
                'message': str(e)
            }), 400
        except IOError as e:
            app.logger.error(f'File system error creating submission: {str(e)}')
            return jsonify({
                'status': 'error',
                'message': 'Internal server error: Failed to save inspection'
            }), 500
        
        # Fire off the post-visit summary email (best-effort; never blocks the
        # response or fails the submission if email is down/unconfigured).
        if email_service.enabled:
            try:
                survey_name = (survey_type_service.get_survey_type_name(submission.get('survey_type_id'))
                               if submission.get('survey_type_id') else None)
                community = submission.get('community')
                region = region_for_community(community)
                region_id = region.get('id') if region else None
                recipients = region_leader_emails(community)
                recipients += settings_service.recipients_for_inspection(
                    region_id, submission.get('inspector_name'))
                # criteria lookup so the email can show "must include to pass" per failed item
                criteria_map = {}
                for q in question_manager.get_all_active_questions():
                    crit = q.get('pass_criteria') or []
                    if q.get('id'):
                        criteria_map[q['id']] = crit
                    if q.get('text'):
                        criteria_map['t:' + q['text'].strip().lower()] = crit
                email_service.send_inspection_report(submission, recipients, survey_name, criteria_map)

                # The community gets its own, narrower email: what was found
                # here and what to do about it — no score, no comparisons.
                ed_emails = community_account_emails(community)
                if ed_emails:
                    responses = submission.get('responses') or []
                    failed = [r for r in responses if r.get('condition') == 'Fail']
                    passed = [r for r in responses if r.get('condition') == 'Pass']
                    open_items = [i for i in (submission.get('action_items') or [])
                                  if not i.get('resolved')]
                    done = len(passed) + len(failed)
                    email_service.send_community_findings(
                        ed_emails, community,
                        submission.get('inspector_name') or submission.get('username') or '',
                        (submission.get('submitted_at') or '')[:10],
                        failed, open_items, criteria_map,
                        notes=submission.get('notes', ''),
                        passed_items=passed,
                        score=round(len(passed) / done * 100) if done else None)

                # Route any item-level comments directed to a company-level team
                for route in settings_service.ROUTES:
                    items = [r for r in submission.get('responses', [])
                             if r.get('route_to') == route and (r.get('description') or '').strip()]
                    if items:
                        to = settings_service.recipients_for_route(route)
                        if to:
                            email_service.send_directed_comments(to, route, submission, items)
            except Exception as e:
                app.logger.error(f'Post-visit email step failed: {e}')

        # Return success with submission data
        return jsonify({
            'status': 'success',
            'submission': submission
        }), 201
        
    except Exception as e:
        # Unexpected errors
        app.logger.exception('Unexpected error submitting inspection')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while submitting inspection'
        }), 500


@app.route('/api/inspections', methods=['GET'])
@login_required
def get_inspections():
    """
    Get inspection submissions with community and survey type filtering
    
    Query Parameters:
        community (optional): Filter submissions by community name (admin only)
        survey_type (optional): Filter submissions by survey type ID
        
    Behavior:
        - For staff users: Automatically filters by their assigned community
        - For admin users: Returns all submissions, or filters by community parameter if provided
        - If survey_type is provided, filters submissions by that survey type
        
    Returns:
        200: JSON with status and submissions array
        400: Invalid survey type
        500: Internal server error
        
    Requirements: 9.1, Task 4.5
    """
    try:
        # Sanitize community filter from query parameter
        community_filter = request.args.get('community')
        if community_filter:
            community_filter = InputSanitizer.sanitize_community_name(community_filter)
        
        # Get survey type filter from query parameter
        survey_type_filter = request.args.get('survey_type')
        if survey_type_filter:
            survey_type_filter = InputSanitizer.sanitize_string(survey_type_filter, max_length=50)
            # Validate survey type
            if not survey_type_service.validate_survey_type(survey_type_filter):
                return jsonify({
                    'status': 'error',
                    'message': f'Invalid survey type: {survey_type_filter}'
                }), 400
        
        # Determine which submissions to return, by role (admin privileges see all)
        role = current_role()
        if is_admin():
            if community_filter:
                submissions = inspection_service.get_submissions_by_community(community_filter)
            else:
                submissions = inspection_service.get_all_submissions()
        elif is_leadership(role):
            # Regionals see submissions across their region's communities
            allowed = set(regional_communities())
            all_subs = inspection_service.get_all_submissions()
            if community_filter and community_filter in allowed:
                submissions = [s for s in all_subs if s.get('community') == community_filter]
            else:
                submissions = [s for s in all_subs if s.get('community') in allowed]
        else:
            # Community account — every site it covers, not just one.
            allowed_comms = set(session_communities())
            submissions = [s for s in inspection_service.get_all_submissions()
                           if s.get('community') in allowed_comms]
        
        # Apply survey type filter if provided
        if survey_type_filter:
            submissions = [
                sub for sub in submissions
                if sub.get('survey_type_id') == survey_type_filter
            ]

        # Enrich with a friendly inspector name. Prefer the name stored on the
        # submission (captured at visit time), else resolve from the username.
        # When photos live in S3, also attach a short-lived signed URL to each
        # response so the private objects can be displayed.
        enriched = []
        for sub in submissions:
            uname = sub.get('username', '')
            name = sub.get('inspector_name') or resolve_display_name(uname)
            new_sub = {**sub, 'inspector_name': name}
            if file_upload_handler.use_s3 and isinstance(sub.get('responses'), list):
                new_responses = []
                for resp in sub['responses']:
                    r = dict(resp)
                    if r.get('photo_path'):
                        r['photo_url'] = file_upload_handler.generate_presigned_url(r['photo_path'])
                    # Photo of the fix, uploaded when a failed standard is marked
                    # as addressed between visits.
                    if r.get('addressed_photo'):
                        r['addressed_photo_url'] = file_upload_handler.generate_presigned_url(r['addressed_photo'])
                    # Photos attached to comments need signing too.
                    if r.get('comments'):
                        r['comments'] = [
                            ({**c, 'photo_url': file_upload_handler.generate_presigned_url(c['photo'])}
                             if c.get('photo') else c)
                            for c in r['comments']]
                    new_responses.append(r)
                new_sub['responses'] = new_responses
                # Ad-hoc items carry comment photos too.
                if isinstance(sub.get('action_items'), list):
                    new_sub['action_items'] = [
                        ({**it, 'comments': [
                            ({**c, 'photo_url': file_upload_handler.generate_presigned_url(c['photo'])}
                             if c.get('photo') else c)
                            for c in it['comments']]}
                         if it.get('comments') else it)
                        for it in sub['action_items']]
            enriched.append(new_sub)

        return jsonify({
            'status': 'success',
            'submissions': enriched
        }), 200
        
    except Exception as e:
        # Log the error for debugging
        app.logger.exception('Error retrieving inspections')
        return jsonify({
            'status': 'error',
            'message': 'Internal server error while retrieving inspections'
        }), 500


# ==================== REPORT EXPORTS (CSV / XLSX / PDF) ====================

def _scoped_submissions_for_export():
    """Inspection submissions visible to the current user, role-scoped exactly
    like the Reports view (admin = all, regional = their region, staff = their
    community), each enriched with a friendly inspector name."""
    role = current_role()
    if is_admin():
        submissions = inspection_service.get_all_submissions()
    elif is_leadership(role):
        allowed = set(regional_communities())
        submissions = [s for s in inspection_service.get_all_submissions()
                       if s.get('community') in allowed]
    else:
        allowed_comms = set(session_communities())
        submissions = [s for s in inspection_service.get_all_submissions()
                       if s.get('community') in allowed_comms]
    out = []
    for sub in submissions:
        name = sub.get('inspector_name') or resolve_display_name(sub.get('username', ''))
        out.append({**sub, 'inspector_name': name})
    return out


# Columns shared by every export format.
_EXPORT_HEADERS = ['Community', 'Region', 'Survey Type', 'Inspector',
                   'Submitted', 'Standard', 'Result', 'Comment']


def _export_rows():
    """Flatten scoped submissions into one row per question response."""
    rows = []
    for sub in _scoped_submissions_for_export():
        community = sub.get('community', '') or ''
        region = region_for_community(community)
        region_name = region.get('name', '') if region else ''
        survey_name = survey_type_service.get_survey_type_name(sub.get('survey_type_id')) or ''
        inspector = sub.get('inspector_name', '') or ''
        submitted = (sub.get('submitted_at', '') or '')[:19].replace('T', ' ')
        responses = sub.get('responses') or []
        if not responses:
            rows.append([community, region_name, survey_name, inspector, submitted, '', '', ''])
            continue
        for r in responses:
            comment = (r.get('description', '') or '').replace('\r', ' ').replace('\n', ' ')
            # A failed standard can be marked as addressed between visits. The
            # result stays "Fail" (the visit is a faithful record); we only note
            # the follow-up alongside it.
            if r.get('addressed'):
                when = (r.get('addressed_at') or '')[:10]
                who = r.get('addressed_by') or ''
                fix = (r.get('addressed_note') or '').replace('\r', ' ').replace('\n', ' ')
                tag = 'Addressed'
                if when:
                    tag += f" {when}"
                if who:
                    tag += f" by {who}"
                comment = f"{comment} · {tag}".strip(' ·')
                if fix:
                    comment += f" — {fix}"
            rows.append([
                community, region_name, survey_name, inspector, submitted,
                r.get('question_text', '') or '',
                r.get('condition', '') or '',
                comment,
            ])
        # Ad-hoc action items raised on the visit. They're follow-up tasks, not
        # standards, so the Result column marks them as such (they never score).
        for it in (sub.get('action_items') or []):
            status = 'Done' if it.get('resolved') else 'Open'
            note = (it.get('resolution_note') or '').replace('\r', ' ').replace('\n', ' ')
            detail = []
            if it.get('assigned_to'):
                detail.append(f"For: {it['assigned_to']}")
            detail.append(status)
            if status == 'Done' and note:
                detail.append(f"Fix: {note}")
            rows.append([
                community, region_name, survey_name, inspector, submitted,
                (it.get('text', '') or '').replace('\r', ' ').replace('\n', ' '),
                f"Action item — {(it.get('priority') or 'medium').capitalize()}",
                ' · '.join(detail),
            ])
    return rows


def _export_filename(ext):
    return f"atlas-inspections-{datetime.now().strftime('%Y%m%d')}.{ext}"


def _export_summary():
    """Aggregate stats for the summary section of every export, computed from the
    same role-scoped submissions as the detail rows."""
    subs = _scoped_submissions_for_export()
    total_visits = len(subs)
    passes = fails = total_responses = 0
    fails_addressed = 0
    actions_total = actions_open = 0
    by_type = {}        # survey_type_id -> count of responses
    performers = {}     # inspector -> {visits, last}
    for sub in subs:
        stid = sub.get('survey_type_id')
        responses = sub.get('responses') or []
        for it in (sub.get('action_items') or []):
            actions_total += 1
            if not it.get('resolved'):
                actions_open += 1
        for r in responses:
            total_responses += 1
            cond = (r.get('condition') or '')
            if cond == 'Pass':
                passes += 1
            elif cond == 'Fail':
                fails += 1
                # Failures fixed between visits still count as failures here —
                # the score is a record of the visit — but we report them so the
                # follow-up work is visible.
                if r.get('addressed'):
                    fails_addressed += 1
            if stid:
                by_type[stid] = by_type.get(stid, 0) + 1
        name = sub.get('inspector_name') or 'Unknown'
        p = performers.setdefault(name, {'visits': 0, 'last': ''})
        p['visits'] += 1
        sa = (sub.get('submitted_at') or '')
        if sa > p['last']:
            p['last'] = sa

    by_survey_type = sorted(
        [(survey_type_service.get_survey_type_name(k) or k, v) for k, v in by_type.items()],
        key=lambda x: x[1], reverse=True)
    top_performers = sorted(
        [(n, d['visits'], (d['last'] or '')[:10]) for n, d in performers.items()],
        key=lambda x: x[1], reverse=True)[:10]

    pass_rate = round(passes / total_responses * 100) if total_responses else 0
    return {
        'total_visits': total_visits,
        'total_responses': total_responses,
        'passes': passes, 'fails': fails, 'pass_rate': pass_rate,
        'fails_addressed': fails_addressed,
        'actions_total': actions_total, 'actions_open': actions_open,
        'by_survey_type': by_survey_type,
        'top_performers': top_performers,
    }


@app.route('/api/reports/export.csv')
@login_required
def export_reports_csv():
    """Download the inspection report data as CSV (role-scoped)."""
    import csv
    import io
    buf = io.StringIO()
    writer = csv.writer(buf)

    # --- Summary block ---
    s = _export_summary()
    writer.writerow(['Atlas Senior Living — Inspection Report'])
    writer.writerow(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M')])
    writer.writerow([])
    writer.writerow(['Total visits', s['total_visits']])
    writer.writerow(['Total responses', s['total_responses']])
    writer.writerow(['Pass', s['passes']])
    writer.writerow(['Fail', s['fails']])
    writer.writerow(['Fail — addressed since the visit', s['fails_addressed']])
    writer.writerow(['Pass rate', f"{s['pass_rate']}%"])
    writer.writerow(['Action items raised', s['actions_total']])
    writer.writerow(['Action items still open', s['actions_open']])
    writer.writerow([])
    writer.writerow(['Survey type', 'Responses'])
    for name, count in s['by_survey_type']:
        writer.writerow([name, count])
    writer.writerow([])
    writer.writerow(['Top performers', 'Visits', 'Last visit'])
    for name, visits, last in s['top_performers']:
        writer.writerow([name, visits, last])
    writer.writerow([])
    writer.writerow(['DETAIL'])

    # --- Detail table ---
    writer.writerow(_EXPORT_HEADERS)
    writer.writerows(_export_rows())
    from flask import Response
    return Response(
        buf.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{_export_filename("csv")}"'},
    )


@app.route('/api/reports/export.xlsx')
@login_required
def export_reports_xlsx():
    """Download the inspection report data as a styled Excel workbook."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill('solid', fgColor='00285C')
    header_font = Font(bold=True, color='FFFFFF')

    wb = Workbook()

    # ---------- Summary sheet ----------
    summ = wb.active
    summ.title = 'Summary'
    s = _export_summary()
    title_font = Font(bold=True, size=14, color='00285C')
    label_font = Font(bold=True)

    summ['A1'] = 'Atlas Senior Living — Inspection Report'
    summ['A1'].font = title_font
    summ['A2'] = f"Generated {datetime.now().strftime('%B %d, %Y %H:%M')}"
    summ['A2'].font = Font(italic=True, color='6B7280')

    kpis = [('Total visits', s['total_visits']), ('Total responses', s['total_responses']),
            ('Pass', s['passes']), ('Fail', s['fails']),
            ('Fail — addressed since the visit', s['fails_addressed']),
            ('Pass rate', f"{s['pass_rate']}%"),
            ('Action items raised', s['actions_total']), ('Action items still open', s['actions_open'])]
    row = 4
    for label, val in kpis:
        summ.cell(row=row, column=1, value=label).font = label_font
        summ.cell(row=row, column=2, value=val)
        row += 1

    row += 1
    summ.cell(row=row, column=1, value='Survey Type Breakdown').font = title_font
    row += 1
    summ.cell(row=row, column=1, value='Survey type').font = header_font
    summ.cell(row=row, column=1).fill = header_fill
    summ.cell(row=row, column=2, value='Responses').font = header_font
    summ.cell(row=row, column=2).fill = header_fill
    row += 1
    for name, count in s['by_survey_type']:
        summ.cell(row=row, column=1, value=name)
        summ.cell(row=row, column=2, value=count)
        row += 1

    row += 1
    summ.cell(row=row, column=1, value='Top Performers').font = title_font
    row += 1
    for col, h in enumerate(['Team member', 'Visits', 'Last visit'], start=1):
        c = summ.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    row += 1
    for name, visits, last in s['top_performers']:
        summ.cell(row=row, column=1, value=name)
        summ.cell(row=row, column=2, value=visits)
        summ.cell(row=row, column=3, value=last)
        row += 1

    summ.column_dimensions['A'].width = 30
    summ.column_dimensions['B'].width = 14
    summ.column_dimensions['C'].width = 16

    # ---------- Detail sheet ----------
    ws = wb.create_sheet('Inspections')
    for col, name in enumerate(_EXPORT_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical='center')

    rows = _export_rows()
    for r in rows:
        ws.append(r)

    # Color the Result cells (Pass green, Fail red).
    pass_font = Font(color='0F8A5F', bold=True)
    fail_font = Font(color='D13212', bold=True)
    for i in range(len(rows)):
        cell = ws.cell(row=i + 2, column=7)  # Result column
        if (cell.value or '') == 'Pass':
            cell.font = pass_font
        elif (cell.value or '') == 'Fail':
            cell.font = fail_font

    widths = [26, 16, 20, 20, 19, 44, 9, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A2'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out, as_attachment=True, download_name=_export_filename('xlsx'),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/reports/export.pdf')
@login_required
def export_reports_pdf():
    """Download the inspection report data as a PDF table (role-scoped)."""
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)

    rows = _export_rows()
    styles = getSampleStyleSheet()
    cell = ParagraphStyle('cell', parent=styles['Normal'], fontSize=7.5, leading=9)
    head = ParagraphStyle('head', parent=styles['Normal'], fontSize=8,
                          leading=10, textColor=colors.white, fontName='Helvetica-Bold')

    # Build a Paragraph-wrapped table so long text wraps instead of overflowing.
    table_data = [[Paragraph(h, head) for h in _EXPORT_HEADERS]]
    for r in rows:
        table_data.append([Paragraph((str(v) if v is not None else ''), cell) for v in r])

    # Widths sum to ~10.1in, within the landscape-letter printable area
    # (11in - 0.8in margins = 10.2in), so the table never overflows the left edge.
    col_widths = [1.6 * inch, 0.8 * inch, 1.1 * inch, 1.0 * inch, 0.95 * inch,
                  2.1 * inch, 0.5 * inch, 2.05 * inch]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=0.4 * inch, rightMargin=0.4 * inch,
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    summ = _export_summary()
    h2 = ParagraphStyle('h2', parent=styles['Heading2'], textColor=colors.HexColor('#00285c'),
                        spaceBefore=10, spaceAfter=6)
    story = [
        Paragraph('Atlas Senior Living — Inspection Report', styles['Title']),
        Paragraph(f"Generated {datetime.now().strftime('%B %d, %Y %H:%M')} · {len(rows)} rows",
                  styles['Normal']),
        Spacer(1, 8),
    ]

    # ---- Summary section ----
    kpi_data = [[
        Paragraph('<b>Total visits</b>', cell), Paragraph('<b>Total responses</b>', cell),
        Paragraph('<b>Pass</b>', cell), Paragraph('<b>Fail</b>', cell),
        Paragraph('<b>Pass rate</b>', cell), Paragraph('<b>Action items</b>', cell),
    ], [
        Paragraph(str(summ['total_visits']), cell), Paragraph(str(summ['total_responses']), cell),
        Paragraph(f"<font color='#0f8a5f'>{summ['passes']}</font>", cell),
        Paragraph(
            f"<font color='#d13212'>{summ['fails']}</font>"
            + (f" <font color='#6b7280' size='8'>({summ['fails_addressed']} addressed)</font>"
               if summ.get('fails_addressed') else ''), cell),
        Paragraph(f"{summ['pass_rate']}%", cell),
        Paragraph(f"{summ['actions_total']} <font color='#6b7280' size='8'>({summ['actions_open']} open)</font>", cell),
    ]]
    kpi_tbl = Table(kpi_data, colWidths=[1.33 * inch] * 6)
    kpi_tbl.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d9dfe8')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eef2f7')),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(kpi_tbl)

    if summ['by_survey_type']:
        story.append(Paragraph('Survey Type Breakdown', h2))
        st_data = [[Paragraph('<b>Survey type</b>', cell), Paragraph('<b>Responses</b>', cell)]]
        st_data += [[Paragraph(n, cell), Paragraph(str(c), cell)] for n, c in summ['by_survey_type']]
        st_tbl = Table(st_data, colWidths=[3.5 * inch, 1.2 * inch])
        st_tbl.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d9dfe8')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eef2f7')),
        ]))
        story.append(st_tbl)

    if summ['top_performers']:
        story.append(Paragraph('Top Performers', h2))
        tp_data = [[Paragraph('<b>Team member</b>', cell), Paragraph('<b>Visits</b>', cell),
                    Paragraph('<b>Last visit</b>', cell)]]
        tp_data += [[Paragraph(n, cell), Paragraph(str(v), cell), Paragraph(l, cell)]
                    for n, v, l in summ['top_performers']]
        tp_tbl = Table(tp_data, colWidths=[3 * inch, 1 * inch, 1.4 * inch])
        tp_tbl.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d9dfe8')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eef2f7')),
        ]))
        story.append(tp_tbl)

    story.append(Paragraph('Detailed Responses', h2))
    if rows:
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.hAlign = 'LEFT'
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00285c')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d9dfe8')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f6f8fb')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(tbl)
    else:
        story.append(Paragraph('No inspection data available.', styles['Normal']))

    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=_export_filename('pdf'),
                     mimetype='application/pdf')


# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors"""
    return jsonify({'error': 'Not found'}), 404


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    return jsonify({'error': 'Internal server error'}), 500


@app.errorhandler(401)
def unauthorized(error):
    """Handle 401 errors - redirect to login"""
    return redirect(url_for('login')), 401


# ==================== MAIN ====================

if __name__ == '__main__':
    # Local development server only. In production the app is served by gunicorn
    # (see deploy/), which imports `app` directly and ignores this block.
    port = int(os.environ.get('PORT', 5001))
    debug = os.environ.get('FLASK_DEBUG', '0') == '1'  # off by default; set FLASK_DEBUG=1 for local dev
    app.run(host='0.0.0.0', port=port, debug=debug)

# ═══════════════════════════════════════════════════════════════════════
# TRASPASO DE SESIÓN DESDE ATLERTS
# ═══════════════════════════════════════════════════════════════════════
#
# Pegar en app.py. Necesita `import os` y `import requests` arriba
# (requests ya suele estar; si no: pip install requests).
#
# CÓMO FUNCIONA
#   1. La app de Atlerts pide un código de un solo uso a nuestro servidor.
#   2. Abre  https://standards.atlasseniorliving.net/sso?code=XXXX
#   3. Esta ruta canjea el código SERVIDOR CONTRA SERVIDOR y recibe el
#      correo verificado de esa persona.
#   4. Busca la cuenta con ese correo y crea la MISMA sesión que /api/login.
#
# POR QUÉ EL CANJE Y NO EL CORREO DIRECTO EN LA URL: cualquiera podría
# escribir ?email=keith@… y entrar como quien quisiera. El código no dice
# nada por sí mismo; hay que preguntárselo a Atlerts, y para eso hace falta
# el secreto compartido que solo tiene este servidor.


ATLERTS_REDEEM_URL = (
    "https://us-central1-atlertsapp.cloudfunctions.net/redeemExcellenceHandoff"
)

# ⚠️ Variable de entorno, NUNCA escrita aquí ni subida a git.
#    Tiene que ser exactamente el mismo valor que EXCELLENCE_SSO_SECRET
#    en los secretos de Firebase.
ATLERTS_SSO_SECRET = os.environ.get("ATLERTS_SSO_SECRET", "")


def _atlerts_account_by_email(email):
    """Busca una cuenta por su correo. Devuelve (username, account) o None.

    El diccionario `account` tiene la MISMA forma que devuelve
    authenticate_user, para que la sesión salga idéntica a la del login
    normal. Si aquí se inventara la forma, alguien acabaría con el
    region_id vacío y sin entender por qué no ve su región.

    Las cuentas de USERS_DB (admin/staff) NO se buscan: no guardan correo
    —resolve_account_context devuelve email: None— así que esas personas
    siguen entrando con su contraseña, como hasta ahora.
    """
    email = (email or "").strip().lower()
    if not email:
        return None

    # --- Cuentas creadas desde el panel (users.json) ---
    #
    # 👉 Si tu user_service llama distinto a "listar todas", cambia SOLO
    #    esta línea. Es lo único que no pude ver desde aquí.
    customs = user_service.get_all()

    items = customs.items() if isinstance(customs, dict) else [
        (u.get("username"), u) for u in (customs or [])
    ]
    for username, custom in items:
        if not username or not custom:
            continue
        if (custom.get("email") or "").strip().lower() != email:
            continue
        return username, {
            "role": custom.get("role", "staff"),
            "community": custom.get("community"),
            "communities": custom.get("communities")
                or ([custom.get("community")] if custom.get("community") else []),
            "region_id": custom.get("region_id"),
            "display_name": profile_service.get_display_name(username)
                or custom.get("display_name") or username,
        }

    # --- Cuentas regionales (por persona) ---
    for username, acct in (get_regional_accounts() or {}).items():
        if (acct.get("email") or "").strip().lower() != email:
            continue
        return username, {
            "role": "regional",
            "community": None,
            "region_id": acct.get("region_id"),
            "display_name": acct.get("display_name"),
        }

    return None


@app.route("/sso")
def atlerts_sso():
    """Entrada desde Atlerts. Ante cualquier duda, login de siempre.

    ⚠️ NUNCA crea cuentas. Solo empareja con las que ya existen. Si esta
    ruta pudiera crear usuarios, Atlerts se convertiría sin querer en una
    fábrica de cuentas de Excellence y nadie controlaría quién entra.

    Y nunca muestra un error: quien no tenga cuenta aquí —o la tenga con
    otro correo— ve la pantalla de acceso normal, que es lo que ha visto
    siempre, en vez de un mensaje que no puede resolver.
    """
    code = (request.args.get("code") or "").strip()
    if not code or not ATLERTS_SSO_SECRET:
        return redirect("/login")

    try:
        resp = requests.post(
            ATLERTS_REDEEM_URL,
            headers={"X-Atlerts-Sso-Secret": ATLERTS_SSO_SECRET},
            json={"code": code},
            timeout=6,
        )
    except Exception as e:
        app.logger.warning("SSO: no se pudo canjear el código: %s", e)
        return redirect("/login")

    if resp.status_code != 200:
        # Caducado, ya usado o inexistente. Atlerts no distingue entre esos
        # casos a propósito, y aquí tampoco hace falta.
        return redirect("/login")

    payload = resp.json() or {}
    found = _atlerts_account_by_email(payload.get("email"))
    if not found:
        app.logger.info("SSO: sin cuenta de Excellence para ese correo")
        return redirect("/login")

    username, account = found

    # A partir de aquí, IDÉNTICO a /api/login. Si algún día cambia allí,
    # tiene que cambiar aquí: dos sesiones distintas para la misma persona
    # es un error que solo aparece en la pantalla más rara de la app.
    session["user"] = username
    session["community"] = account["community"]
    session["communities"] = account.get("communities") or (
        [account["community"]] if account["community"] else []
    )
    session["role"] = account["role"]
    session["region_id"] = account["region_id"]
    session["display_name"] = account["display_name"]
    session.permanent = True
    session["admin_extra"] = profile_service.get_admin_extra(username)
    must_change = profile_service.get_must_change(username)
    session["must_change"] = bool(must_change)

    presence_service.record_login(username)
    activity_service.log(
        username, "login", "Signed in from Atlerts",
        meta={"ip": _client_ip()},
    )

    return redirect("/change-password" if must_change else "/")
